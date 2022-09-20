import json
import re
from copy import deepcopy
from http import HTTPStatus
import csv
import threading

import pydash
from openpyxl import Workbook, load_workbook
from django.contrib.auth import authenticate, login, logout
from django.db.models import Q
from django.forms import model_to_dict
from django.shortcuts import render
from django.http import HttpResponse, Http404, HttpResponseRedirect, JsonResponse
from django.urls import reverse
from django.db import IntegrityError
from django.core.mail import EmailMessage
from django.template.loader import get_template, render_to_string
from csv import reader
from io import BytesIO
import PyPDF2
from datetime import datetime
from decimal import *
from .models import Job, JobFolder, Type, Manufacturer, EquipmentFolder, Model, ModelFolder, Equipment, TestEquipment, CableTestData, \
    UserProperties, JobNotes, EquipmentNotes, EquipmentLink, TypeFolder, TypeTestStandards, TypeTestGuide, ModelTestGuide, JobSite, \
    JobSiteNotes, BusContactTestData, JobSiteFolder, Company, FeedbackFile, FeedbackNote, WorkingNote, TestSheet, TypeNotes, ModelNotes, Well, WellNotes, MaintEvent, MaintFile, MaintNotes, STATUS_PENDING


def create_profile(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})

    #get user
    try:
        #if properties already exist, get them
        properties = UserProperties.objects.get(user=request.user)
        # user_base = request.user
    except UserProperties.DoesNotExist:
        #if they dont already exist, create new user properties
        properties=UserProperties(user=request.user)
        properties.save()
    except:
        render(request, "jobs/error.html", {"message": "User profile load error. Contact Admin. 210-303-0471 or albert@blueskysw.com."})
        
    #extract data from form
    if "first_name" in request.POST:
        request.user.first_name = request.POST["first_name"]       
    if "last_name" in request.POST:
        request.user.last_name = request.POST["last_name"] 
    if "email" in request.POST:
        request.user.email = request.POST["email"]   
    request.user.save()

    if "company_key" in request.POST :    
        
        #get company and associate user to company
        if request.POST["company_key"] and request.POST["company_key"].strip():
            try:
                company = Company.objects.get(company_key=request.POST["company_key"])
                properties.company = company
            except Company.DoesNotExist:
                return render(request, "jobs/error.html", {"message": "Invalid Company Key. Return to profile and try again."})
        
    if "phone" in request.POST:
        properties.phone_number = request.POST["phone"]#adds phone number
    properties.save()
    return HttpResponseRedirect(reverse("profile"))
    

def create_job(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})

    #get user properties
    user_properties=UserProperties.objects.get(user=request.user)
    #get user company
    try:
        company=user_properties.company
    except:
        company=None
    #verify uniqueness of job within this company
    if Job.objects.filter(job_number=request.POST["job_number"], company = company):
        return render(request, "jobs/error.html", {"message": "Job Number is In Use. Enter A Unique One."})    
    
    #fill in all properties that were submitted in the form
    if "job_name" in request.POST:
        # extract job_name from form
        job_name = request.POST["job_name"].strip()
        #create new Job objects
        job=Job.objects.create_job(job_name, request.user)
        job_id=job.id
    else:
        return render(request, "jobs/error.html", {"message": "Valid job name was not entered."})

    if company is not None:
        job.company=company
    if "site_pk" in request.POST and (request.POST["site_pk"]!=""):    
        job.job_site = JobSite.objects.get(pk=request.POST["site_pk"])  
    elif "new_site" in request.POST and ("new_site"!=""):  
        site=JobSite(name = request.POST["new_site"].strip())  
        site.save()  
        site.company = company
        site.save()
        job.job_site = site  
    else:  
        return render(request, "jobs/error.html", {"message": "Job site validation error"})  
    if "job_number" in request.POST:
        job.job_number = request.POST["job_number"]
    if "customer_name" in request.POST:
        job.customer_name = request.POST["customer_name"]
    if "site_contact" in request.POST:
        job.site_contact = request.POST["site_contact"]
    if "site_contact_info" in request.POST:
        job.site_contact_info = request.POST["site_contact_info"]
    if "project_manager" in request.POST:
        job.project_manager = request.POST["project_manager"]
    if "start_date" in request.POST:
        if request.POST["start_date"]!="":
            job.start_date = request.POST["start_date"]
    if "end_date" in request.POST:
        if request.POST["end_date"]!="":
            job.end_date = request.POST["end_date"]
    if "job_scope_details" in request.POST:
        job.job_scope_details = request.POST["job_scope_details"]
    if "work_schedule" in request.POST:
        job.work_schedule = request.POST["work_schedule"]
    if "food_accomodations" in request.POST:
        job.food_accomodations = request.POST["food_accomodations"]
    if "lodging_recommendations" in request.POST:
        job.lodging_recommendations = request.POST["lodging_recommendations"]
    if "weather_considerations" in request.POST:
        job.weather_considerations = request.POST["weather_considerations"]
    if "address" in request.POST:
        job.address = request.POST["address"]
    if "nav_link" in request.POST:
        job.nav_link = request.POST["nav_link"]
    if "site_navigation" in request.POST:
        job.site_navigation = request.POST["site_navigation"]
    if "safety_training_time" in request.POST:
        job.safety_training_time = request.POST["safety_training_time"]
    if "safety_training_location" in request.POST:
        job.safety_training_location = request.POST["safety_training_location"]
    if "escort_considerations" in request.POST:
        job.escort_considerations = request.POST["escort_considerations"]
    if "restricted_items" in request.POST:
        job.restricted_items = request.POST["restricted_items"]
    if "grounding_wire_size" in request.POST:
        job.grounding_wire_size = request.POST["grounding_wire_size"]
    if "grounding_clamp_style" in request.POST:
        job.grounding_clamp_style = request.POST["grounding_clamp_style"]
    if "grounding_cluster_quantity" in request.POST:
        if request.POST["grounding_cluster_quantity"]!="":
            job.grounding_cluster_quantity = int(request.POST["grounding_cluster_quantity"])
    if "other_ppe_requirements" in request.POST:
        job.other_ppe_requirements = request.POST["other_ppe_requirements"]
    if "switching_specifications" in request.POST:
        job.switching_specifications = request.POST["switching_specifications"]
    if "live_work_voltage" in request.POST:
        job.live_work_voltage = request.POST["live_work_voltage"]
    if "chemical_hazards" in request.POST:
        job.chemical_hazards = request.POST["chemical_hazards"]
    if "permit_requirements" in request.POST:
        job.permit_requirements = request.POST["permit_requirements"]
    if "extension_cords" in request.POST:
        job.extension_cords = request.POST["extension_cords"]
    if "generators" in request.POST:
        job.generators = request.POST["generators"]
    if "gasoline" in request.POST:
        if request.POST["gasoline"]!="":
            job.gasoline = int(request.POST["gasoline"])
    if "diesel" in request.POST:
        if request.POST["diesel"]!="":
            job.diesel = int(request.POST["diesel"])
    if "ladders" in request.POST:
        job.ladders = request.POST["ladders"]
    if "harness_lanyard" in request.POST:
        job.harness_lanyard = request.POST["harness_lanyard"]
    if "torque_wrenches" in request.POST:
        job.torque_wrenches = request.POST["torque_wrenches"]
    if "tables" in request.POST:
        if request.POST["tables"]!="":
            job.tables = int(request.POST["tables"])
    if "chairs" in request.POST:
        if request.POST["chairs"]!="":
            job.chairs = int(request.POST["chairs"])
    if "lifts" in request.POST:
        job.lifts = request.POST["lifts"]
    if "additional_tools" in request.POST:
        job.additional_tools = request.POST["additional_tools"]
    if "material" in request.POST:
        job.material = request.POST["material"]
    if "job_folder" in request.POST:
        job.job_folder = request.POST["job_folder"]
    if "quoted_price" in request.POST:
        if request.POST["quoted_price"]!="":
            job.quoted_price = int(request.POST["quoted_price"])



    #set all boolean values
    if "is_startup" in request.POST:
        job.is_startup=True
    if "is_preventative_maintenance" in request.POST:
        job.is_preventative_maintenance=True
    if "is_troubleshooting" in request.POST:
        job.is_troubleshooting=True
    if "is_warranty" in request.POST:
        job.is_warranty=True
    if "is_standard_testing" in request.POST:
        job.is_standard_testing=True
    if "is_neta_testing" in request.POST:
        job.is_neta_testing=True
    if "is_safety_training_required" in request.POST:
        job.is_safety_training_required=True
    if "is_hardhat" in request.POST:
        job.is_hardhat=True
    if "is_safety_glasses" in request.POST:
        job.is_safety_glasses=True
    if "is_safety_shoes" in request.POST:
        job.is_safety_shoes=True
    if "is_safety_vest" in request.POST:
        job.is_safety_vest=True
    if "is_safety_gloves" in request.POST:
        job.is_safety_gloves=True
    if "is_fr_clothes" in request.POST:
        job.is_fr_clothes=True
    if "is_h2s_monitor" in request.POST:
        job.is_h2s_monitor=True
    if "is_mv_voltage_detector" in request.POST:
        job.is_mv_voltage_detector=True
    if "is_insulated_gloves" in request.POST:
        job.is_insulated_gloves=True
    if "is_8cal_protection" in request.POST:
        job.is_8cal_protection=True
    if "is_40cal_protection" in request.POST:
        job.is_40cal_protection=True
    if "is_harness" in request.POST:
        job.is_harness=True
    if "is_grounding_cluster" in request.POST:
        job.is_grounding_cluster=True
    if "is_6ft_work" in request.POST:
        job.is_6ft_work=True
    if "is_switching_required" in request.POST:
        job.is_switching_required=True
    if "is_live_work_required" in request.POST:
        job.is_live_work_required=True
    if "is_ungaurded_holes" in request.POST:
        job.is_ungaurded_holes=True
    if "is_confined_space" in request.POST:
        job.is_confined_space=True
    if "is_time_and_materials" in request.POST:
        job.is_time_and_materials=True
    if "is_standard_handtools" in request.POST:
        job.is_standard_handtools=True
    if "is_bus_bender" in request.POST:
        job.is_bus_bender=True
    if "is_fork_lift" in request.POST:
        job.is_fork_lift=True
    if "is_trailer" in request.POST:
        job.is_trailer=True
    job.save()
    job.user_properties.add(user_properties)
    # return render(request, "jobs/new_job2.html", context)
    return HttpResponseRedirect(reverse("job", args=(job_id, ))+ "#equipment")

def create_job_quick(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    # get user properties
    user_properties = UserProperties.objects.get(user=request.user)
    print(user_properties)
    # get user company
    try:
        company = user_properties.company
    except:
        company = None
    # verify uniqueness of job within this company
    if Job.objects.filter(job_number=request.POST["job_number"], company=company):
        return render(request, "jobs/error.html", {"message": "Job Number is In Use. Enter A Unique One."})

    # fill in all properties that were submitted in the form
    if "job_name" in request.POST:
        # extract job_name from form
        job_name = request.POST["job_name"].strip()
        # create new Job objects
        job = Job.objects.create_job(job_name, request.user)
        job_id = job.id
    else:
        return render(request, "jobs/error.html", {"message": "Valid job name was not entered."})

    if company is not None:
        job.company = company
    if "site_pk" in request.POST and (request.POST["site_pk"]!=""):    
        job.job_site = JobSite.objects.get(pk=request.POST["site_pk"])  
    elif "new_site" in request.POST and ("new_site"!=""):  
        site=JobSite(name = request.POST["new_site"].strip())
        site.save()  
        site.company = company
        site.save()
        job.job_site = site  
    else:  
        return render(request, "jobs/error.html", {"message": "Job site validation error"})  
    if "job_number" in request.POST:
        job.job_number = request.POST["job_number"]
    if "customer_name" in request.POST:
        job.customer_name = request.POST["customer_name"]
    if "site_contact" in request.POST:
        job.site_contact = request.POST["site_contact"]
    if "site_contact_info" in request.POST:
        job.site_contact_info = request.POST["site_contact_info"]
    if "project_manager" in request.POST:
        job.project_manager = request.POST["project_manager"]
    if request.POST.get("start_date"):
        job.start_date = request.POST["start_date"]
    if request.POST.get("end_date"):
        job.end_date = request.POST["end_date"]
    if "address" in request.POST:
        job.address = request.POST["address"]
    if "nav_link" in request.POST:
        job.nav_link = request.POST["nav_link"]
    if "is_safety_training_required" in request.POST:
        job.is_safety_training_required = True
    if "safety_training_time" in request.POST:
        job.safety_training_time = request.POST["safety_training_time"]
    if "safety_training_time" in request.POST:
        job.safety_training_time = request.POST["safety_training_time"]

    job.is_hardhat = True
    job.is_safety_glasses = True
    job.is_safety_shoes = True
    job.is_safety_vest = True
    job.is_fr_clothes = True
    job.save()
    job.user_properties.add(user_properties)
    return HttpResponseRedirect(reverse("job", args=(job_id, )))

def add_job_note(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    job=Job.objects.get(pk=job_id)
    
    if "add_note" in request.POST:
        if request.POST["add_note"] and request.POST["add_note"].strip():
            job_note_posted = request.POST["add_note"]
            job_note = JobNotes.objects.create(note=job_note_posted, job=job, author=request.user)
            note_html = note_notification(job_note)
            send_notification(request.user.email, ("Paragon Job Note: \""+job_note.truncated), note_html)
            return HttpResponseRedirect(reverse("job", args=(job_id, ))+"#job_notes")
    return render(request, "jobs/error.html", {"message": "Failure! Note not added. Must Contain letters, numbers, or symbols"})

def add_jobsite_note(request, jobsite_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    site=JobSite.objects.get(pk=jobsite_id)    
    if "add_note" in request.POST:
        if request.POST["add_note"] and request.POST["add_note"].strip():
            note_posted = request.POST["add_note"]
            note = JobSiteNotes(note=note_posted, jobsite=site, author=request.user)
            note.save()
            return HttpResponseRedirect(reverse("job_site", args=(jobsite_id, )))
    return render(request, "jobs/error.html", {"message": "Failure! Note not added. Must Contain letters, numbers, or symbols"})


def add_note(request, note_type):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    user = request.user
    user_properties=UserProperties.objects.get(user=request.user)
    new_note = None
    if request.POST.get('add_note'):
        note_posted = request.POST["add_note"].strip()

        if note_type == 'model_note' and request.POST.get('model_id'):
             new_note = ModelNotes.objects.create(
                 note=note_posted,
                 model_id=request.POST['model_id'],
                 author=user,
             )

        if note_type == 'type_note' and request.POST.get('type_id'):
             new_note = TypeNotes.objects.create(
                 note=note_posted,
                 eq_type_id=request.POST['type_id'],
                 author=user,
                 company= user_properties.company,
             )
             if request.POST.get('is_private'):
                 new_note.is_private = True           
        if new_note:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    return render(request, "jobs/error.html", {"message": "Failure! Note not changed. Must Contain letters, numbers, or symbols"})


def edit_note(request, note_type, note_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    if note_type == 'job_note':
        note = JobNotes.objects.get(pk=note_id)
    if note_type == 'equipment_note':
        note = EquipmentNotes.objects.get(pk=note_id)
    if note_type == 'type_note':
        note = TypeNotes.objects.get(pk=note_id)
    if note_type == 'model_note':
        note = ModelNotes.objects.get(pk=note_id)
    input_id = "edit_"+note_type+"_input_"+str(note_id)
    if input_id in request.POST:
        if request.POST[input_id] and request.POST[input_id].strip():
            note.note=request.POST[input_id]
            note.save()
    
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def delete_note(request, note_type, note_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    if note_type == 'job_note':
        note = JobNotes.objects.get(pk=note_id)
    if note_type == 'equipment_note':
        note = EquipmentNotes.objects.get(pk=note_id)
    if note_type == 'type_note':
        note = TypeNotes.objects.get(pk=note_id)
    if note_type == 'model_note':
        note = ModelNotes.objects.get(pk=note_id)
    note.delete()
    note.sub_notes.all().delete()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def reply_note(request, note_type, note_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    subject_pre = None
    if note_type == 'job_note':
        parent_note = JobNotes.objects.get(pk=note_id)
        job = Job.objects.get(pk=parent_note.job.pk)
        note=JobNotes(parent_note=parent_note)
        note.job=job
        note.author=request.user
        subject_pre = "Paragon Job Note Reply: \""
    if note_type == 'equipment_note':
        parent_note = EquipmentNotes.objects.get(pk=note_id)
        equipment = Equipment.objects.get(pk=parent_note.equipment.pk)
        note=EquipmentNotes(parent_note=parent_note)
        note.equipment=equipment
        note.author=request.user
        subject_pre = "Paragon Eq. Note Reply: \""
    if note_type == 'type_note':
        is_notif
        parent_note = TypeNotes.objects.get(pk=note_id)
        eq_type = Type.objects.get(pk=parent_note.eq_type.pk)
        note=TypeNotes(parent_note=parent_note)
        note.eq_type=eq_type
        note.author=request.user
    if note_type == 'model_note':
        parent_note = ModelNotes.objects.get(pk=note_id)
        model = Model.objects.get(pk=parent_note.model.pk)
        note = ModelNotes(parent_note=parent_note)
        note.model=model
        note.author=request.user
        
    reply_id = "reply_"+note_type+"_input_"+str(note_id)

    if reply_id in request.POST:
        if request.POST[reply_id] and request.POST[reply_id].strip():
            reply_posted = request.POST[reply_id].strip()
            note.note=reply_posted
            note.save()
            note_html = note_notification(note)
            notification_threader(request.user.email, (subject_pre+note.truncated), note_html)            
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    return render(request, "jobs/error.html", {"message": "Failure! Note not changed. Must Contain letters, numbers, or symbols"})


def edit_job_test_equipments(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    job = Job.objects.get(pk=job_id)
    if "mandatory_test_equipment" in request.POST:
        mandatory_test_equipments = request.POST['mandatory_test_equipment'].split(',')
        mandatory_test_equipments = TestEquipment.objects.filter(name__in=mandatory_test_equipments)

        existing_test_equipments = [i[0] for i in job.test_equipments]
        for existing_test_equipment in existing_test_equipments:
            if existing_test_equipment not in mandatory_test_equipments:
                equipments = (job.equipment
                              .filter(Q(mandatory_test_equipment=existing_test_equipment) |
                                      Q(optional_test_equipment=existing_test_equipment),
                                      trashed=False))
                for equipment in equipments:
                    equipment.mandatory_test_equipment.remove(existing_test_equipment)
                    equipment.optional_test_equipment.remove(existing_test_equipment)

                job.test_equipment.remove(existing_test_equipment)

        for mandatory_test_equipment in mandatory_test_equipments:
            if mandatory_test_equipment not in existing_test_equipments:
                job.test_equipment.add(mandatory_test_equipment)

    job.save()
    return HttpResponseRedirect(reverse("job", args=(job_id, )))

def edit_jobsite(request, jobsite_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    site = JobSite.objects.get(pk=jobsite_id)    
    jobs = Job.objects.filter(job_site = site)
    if "name" in request.POST:
        site.name = request.POST["name"] 
    if "site_contact" in request.POST:
        site.site_contact = request.POST["site_contact"] 
    if "site_contact_info" in request.POST:
        site.site_contact_info = request.POST["site_contact_info"] 
    if "owner" in request.POST:
        site.owner = request.POST["owner"] 
    if "facility_manager" in request.POST:
        site.facility_manager = request.POST["facility_manager"]
    if "other_contacts" in request.POST:
        site.other_contacts = request.POST["other_contacts"] 
    if "safety_training_item_requirements" in request.POST:
        site.safety_training_item_requirements = request.POST["safety_training_item_requirements"] 
    if "safety_training_location" in request.POST:
        site.safety_training_location = request.POST["safety_training_location"] 
    if "safety_training_time" in request.POST:
        site.safety_training_time = request.POST["safety_training_time"] 
    if "safety_training_procedure" in request.POST:
        site.safety_training_procedure = request.POST["safety_training_procedure"]
    if "additional_ppe_requirements" in request.POST:
        site.additional_ppe_requirements = request.POST["additional_ppe_requirements"]         
    if "address" in request.POST:
        site.address = request.POST["address"] 
    if "nav_link" in request.POST:
        site.nav_link = request.POST["nav_link"] 
    if "site_navigation" in request.POST:
        site.site_navigation = request.POST["site_navigation"] 
    if "background_check_procedure" in request.POST:
        site.background_check_procedure = request.POST["background_check_procedure"] 
    if "documents_required_for_access" in request.POST:
        site.documents_required_for_access = request.POST["documents_required_for_access"] 
    if "entry_procedure" in request.POST:
        site.entry_procedure = request.POST["entry_procedure"] 
    if "driving_vehicles" in request.POST:
        site.driving_vehicles = request.POST["driving_vehicles"] 
    if "parking_considerations" in request.POST:
        site.parking_considerations = request.POST["parking_considerations"] 
    if "restricted_items" in request.POST:
        site.restricted_items = request.POST["restricted_items"] 
    if "access_requirements" in request.POST:
        site.access_requirements = request.POST["access_requirements"] 
    if "is_hardhat" in request.POST:
        site.is_hardhat= True
        jobs.update(is_hardhat = True)
    else:
        site.is_hardhat= False
    if "is_safety_glasses" in request.POST:
        site.is_safety_glasses=True
        jobs.update(is_safety_glasses = True)
    else:
        site.is_safety_glasses= False
    if "is_safety_shoes" in request.POST:
        site.is_safety_shoes= True
        jobs.update(is_safety_shoes = True)
    else:
        site.is_safety_shoes= False
    if "is_safety_vest" in request.POST:
        site.is_safety_vest=True
        jobs.update(is_safety_vest = True)
    else:
        site.is_safety_vest= False
    if "is_safety_gloves" in request.POST:
        site.is_safety_gloves=True
        jobs.update(is_safety_gloves = True)
    else:
        site.is_safety_gloves= False
    if "is_fr_clothes" in request.POST:
        site.is_fr_clothes=True
        jobs.update(is_fr_clothes = True)
    else:
        site.is_fr_clothes= False
    if "is_h2s_monitor" in request.POST:
        site.is_h2s_monitor=True
        jobs.update(is_h2s_monitor = True)
    else:
        site.is_h2s_monitor= False
    
    site.save()
    return HttpResponseRedirect(reverse("job_site", args=(jobsite_id, )))

def edit_job(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    job=Job.objects.get(pk=job_id)
     #fill in all properties that were submitted in the form

    if "address" in request.POST:
        job.job_site.address = request.POST["address"]
    if "company" in request.POST:
        job.company = request.POST["company"]
    if "job_site" in request.POST:
        job.job_site = request.POST["job_site"]
    if "costs" in request.POST:
        job.costs = int(request.POST["costs"])
    # if "user_properties" in request.POST:

    if "text_input_jobname" in request.POST:
        if request.POST["text_input_jobname"] and request.POST["text_input_jobname"].strip():
            job.job_name = request.POST["text_input_jobname"]
    if "job_number" in request.POST:
        job.job_number = request.POST["job_number"]
    if "customer_name" in request.POST:
        job.customer_name = request.POST["customer_name"]
    if "site_contact" in request.POST:
        job.site_contact = request.POST["site_contact"]
    if "site_contact_info" in request.POST:
        job.site_contact_info = request.POST["site_contact_info"]
    if "project_manager" in request.POST:
        job.project_manager = request.POST["project_manager"]
    if "start_date" in request.POST:
        if request.POST["start_date"]!="":
            job.start_date = request.POST["start_date"]
    if "end_date" in request.POST:
        if request.POST["end_date"]!="":
            job.end_date = request.POST["end_date"]
    if "job_scope_details" in request.POST:
        job.job_scope_details = request.POST["job_scope_details"]
    if "work_schedule_i" in request.POST:
        if request.POST["work_schedule_i"] and request.POST["work_schedule_i"].strip():
            job.work_schedule = request.POST["work_schedule_i"]
    if "edit_food" in request.POST:
        if request.POST["edit_food"] and request.POST["edit_food"].strip():
            job.food_accomodations = request.POST["edit_food"]
    if "edit_lodging" in request.POST:
        if request.POST["edit_lodging"] and request.POST["edit_lodging"].strip():
            job.lodging_recommendations = request.POST["edit_lodging"]
    if "food_accomodations" in request.POST:
            job.food_accomodations = request.POST["food_accomodations"].strip()
    if "lodging_recommendations" in request.POST:
            job.lodging_recommendations = request.POST["lodging_recommendations"].strip()
    if "weather_considerations" in request.POST:
            job.weather_considerations = request.POST["weather_considerations"].strip()
    if "text_input_address" in request.POST:
        if request.POST["text_input_address"] and request.POST["text_input_address"].strip():
            job.address = request.POST["text_input_address"]
    if "text_input_address2" in request.POST:
        if request.POST["text_input_address2"] and request.POST["text_input_address2"].strip():
            job.address = request.POST["text_input_address2"]
    if "edit_navlink" in request.POST:
        if request.POST["edit_navlink"] and request.POST["edit_navlink"].strip():
            job.nav_link = request.POST["edit_navlink"]
    if "edit_sitenav" in request.POST:
        if request.POST["edit_sitenav"] and request.POST["edit_sitenav"].strip():
            job.site_navigation = request.POST["edit_sitenav"]
    if "edit_special_te" in request.POST:
        if request.POST["edit_special_te"] and request.POST["edit_special_te"].strip():
            job.specialty_test_equipment = request.POST["edit_special_te"]
    if "edit_ladders" in request.POST:
        if request.POST["edit_ladders"] and request.POST["edit_ladders"].strip():
            job.ladders = request.POST["edit_ladders"]

    if "safety_training_time" in request.POST:
        job.safety_training_time = request.POST["safety_training_time"]
    if "safety_training_location" in request.POST:
        job.safety_training_location = request.POST["safety_training_location"]
    if "escort_considerations" in request.POST:
        job.escort_considerations = request.POST["escort_considerations"].strip()
    if "restricted_items" in request.POST:
        job.job_site.restricted_items = request.POST["restricted_items"].strip()
    if "access_requirements" in request.POST:
        job.job_site.access_requirements = request.POST["access_requirements"].strip()
    if "parking_considerations" in request.POST:
        job.job_site.parking_considerations = request.POST["parking_considerations"].strip()
    if "driving_vehicles" in request.POST:
        job.job_site.driving_vehicles = request.POST["driving_vehicles"].strip()
    if "entry_procedure" in request.POST:
        job.job_site.entry_procedure = request.POST["entry_procedure"].strip()
    if "documents_required_for_access" in request.POST:
        job.job_site.documents_required_for_access = request.POST["documents_required_for_access"].strip()
    if "background_check_procedure" in request.POST:
        job.job_site.background_check_procedure = request.POST["background_check_procedure"].strip()
    if "grounding_wire_size" in request.POST:
        job.grounding_wire_size = request.POST["grounding_wire_size"]
    if "grounding_clamp_style" in request.POST:
        job.grounding_clamp_style = request.POST["grounding_clamp_style"]
    if "grounding_cluster_quantity" in request.POST:
        if request.POST["grounding_cluster_quantity"]!="":
            job.grounding_cluster_quantity = int(request.POST["grounding_cluster_quantity"])
    if "other_ppe_requirements" in request.POST:
        job.other_ppe_requirements = request.POST["other_ppe_requirements"]
    if "switching_specifications" in request.POST:
        job.switching_specifications = request.POST["switching_specifications"]
    if "live_work_voltage" in request.POST:
        job.live_work_voltage = request.POST["live_work_voltage"]
    if "chemical_hazards" in request.POST:
        job.chemical_hazards = request.POST["chemical_hazards"]
    if "permit_requirements" in request.POST:
        job.permit_requirements = request.POST["permit_requirements"]
    if "extension_cords" in request.POST:
        job.extension_cords = request.POST["extension_cords"]
    if "generators" in request.POST:
        job.generators = request.POST["generators"]
    if "gasoline" in request.POST:
        if request.POST["gasoline"]!="":
            job.gasoline = int(request.POST["gasoline"])
    if "diesel" in request.POST:
        if request.POST["diesel"] != "":
            job.diesel = int(request.POST["diesel"])
    # if "ladders" in request.POST:
    #     job.ladders = request.POST["ladders"]
    if "harness_lanyard" in request.POST:
        job.harness_lanyard = request.POST["harness_lanyard"]
    if "torque_wrenches" in request.POST:
        job.torque_wrenches = request.POST["torque_wrenches"]
    if "tables" in request.POST:
        if request.POST["tables"]!="":
            job.tables = int(request.POST["tables"])
    if "chairs" in request.POST:
        if request.POST["chairs"]!="":
            job.chairs = int(request.POST["chairs"])
    if "lifts" in request.POST:
        job.lifts = request.POST["lifts"]
    if "additional_tools" in request.POST:
        job.additional_tools = request.POST["additional_tools"]
    if "material" in request.POST:
        job.material = request.POST["material"]
    # if "job_folder" in request.POST:
    #     job.job_folder = request.POST["job_folder"]
    if "quoted_price" in request.POST:
        if request.POST["quoted_price"]!="":
            job.quoted_price = int(request.POST["quoted_price"])



    # #set all boolean values
    if "is_startup" in request.POST:
        job.is_startup=request.POST['is_startup'] == 'True'
    if "is_preventative_maintenance" in request.POST:
        job.is_preventative_maintenance=request.POST['is_preventative_maintenance'] == 'True'
    if "is_troubleshooting" in request.POST:
        job.is_troubleshooting=request.POST['is_troubleshooting'] == 'True'
    if "is_warranty" in request.POST:
        job.is_warranty=request.POST['is_warranty'] == 'True'
    if "is_standard_testing" in request.POST:
        job.is_standard_testing=request.POST['is_standard_testing'] == 'True'
    if "is_neta_testing" in request.POST:
        job.is_neta_testing=request.POST['is_neta_testing'] == 'True'
    if "is_safety_training_required" in request.POST:
        job.is_safety_training_required=request.POST['is_safety_training_required'] == 'True'
    if "is_hardhat" in request.POST:
        job.is_hardhat=request.POST['is_hardhat'] == 'True'
    if "is_safety_glasses" in request.POST:
        job.is_safety_glasses=request.POST['is_safety_glasses'] == 'True'
    if "is_safety_shoes" in request.POST:
        job.is_safety_shoes=request.POST['is_safety_shoes'] == 'True'
    if "is_safety_vest" in request.POST:
        job.is_safety_vest=request.POST['is_safety_vest'] == 'True'
    if "is_safety_gloves" in request.POST:
        job.is_safety_gloves=request.POST['is_safety_gloves'] == 'True'
    if "is_fr_clothes" in request.POST:
        job.is_fr_clothes = request.POST['is_fr_clothes'] == 'True'
    if "is_h2s_monitor" in request.POST:
        job.is_h2s_monitor=request.POST['is_h2s_monitor'] == 'True'
    if "is_mv_voltage_detector" in request.POST:
        job.is_mv_voltage_detector=request.POST['is_mv_voltage_detector'] == 'True'
    if "is_insulated_gloves" in request.POST:
        job.is_insulated_gloves=request.POST['is_insulated_gloves'] == 'True'
    if "is_8cal_protection" in request.POST:
        job.is_8cal_protection=request.POST['is_8cal_protection'] == 'True'
    if "is_40cal_protection" in request.POST:
        job.is_40cal_protection=request.POST['is_40cal_protection'] == 'True'
    if "is_harness" in request.POST:
        job.is_harness=request.POST['is_harness'] == 'True'
    if "is_grounding_cluster" in request.POST:
        job.is_grounding_cluster=request.POST['is_grounding_cluster'] == 'True'
    if "is_6ft_work" in request.POST:
        job.is_6ft_work=request.POST['is_6ft_work'] == 'True'
    if "is_switching_required" in request.POST:
        job.is_switching_required=request.POST['is_switching_required'] == 'True'
    if "is_live_work_required" in request.POST:
        job.is_live_work_required=request.POST['is_live_work_required'] == 'True'
    if "is_ungaurded_holes" in request.POST:
        job.is_ungaurded_holes=request.POST['is_ungaurded_holes'] == 'True'
    if "is_confined_space" in request.POST:
        job.is_confined_space=request.POST['is_confined_space'] == 'True'
    if "is_time_and_materials" in request.POST:
        job.is_time_and_materials = request.POST['is_time_and_materials'] == 'True'
    if "edit_handtools_true" in request.POST:
        if request.POST["edit_handtools_true"]=="True":
            job.is_standard_handtools=True
    elif "is_standard_handtools" in request.POST:
        job.is_standard_handtools = request.POST["is_standard_handtools"] == "True"
    if "bus_bender" in request.POST:
        job.is_bus_bender = request.POST['bus_bender'] == 'True'
    if "is_fork_lift" in request.POST:
        job.is_fork_lift = request.POST['is_fork_lift'] == 'True'
    if "is_trailer" in request.POST:
        job.is_trailer = request.POST['is_trailer'] == 'True'

    if "trashed" in request.POST:
        job.trashed = request.POST['trashed'] == 'True'

    if "archived" in request.POST:
        job.archived = request.POST['archived'] == 'True'

    if "completion" in request.POST:
        job.completion = request.POST['completion'] == 'True'

    job.save()
    job.job_site.save()
    return HttpResponseRedirect(reverse("job", args=(job_id, )))

def add_job_files(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        job = Job.objects.get(pk=job_id)
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except Job.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid Job Selection."})
    if 'file_add' in request.FILES:
        file_add = request.FILES["file_add"]
        job_folder=JobFolder(job=job)
        job_folder.job_file = file_add
        job_folder.save()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def add_jobsite_files(request, jobsite_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        site = JobSite.objects.get(pk=jobsite_id)
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except JobSite.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "add jobsite files error. contact admin"})
    if 'file_add' in request.FILES:
        file_add = request.FILES["file_add"]
        site_file=JobSiteFolder(jobsite=site)
        site_file.jobsite_file = file_add
        site_file.save()
        if 'eq_file_name' in request.POST:
            if request.POST["eq_file_name"].strip() != "":
                site_file.file_name = request.POST["eq_file_name"]
                site_file.save()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    path("job_folder<int:job_id>/addfile", constructors.add_job_files, name="add_job_files"),



def remove_job_files(request, file_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        job_file = JobFolder.objects.get(pk=file_id)
        job_id=job_file.job.id
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except JobFolder.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid File Selection. Contact Admin"})
    job_file.job_file.delete(save=False) #delete file in S3
    job_file.delete() #delete file in django
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
#need to definie function to compile test results into single pdf
# def compile_test_results(request, job_id):
#     x = HttpResponseRedirect(reverse("job_folder", args=(job_id, )))
#     try:
#         job=Job.objects.get(pk=job_id)
#     except KeyError:
#         return render(request, "jobs/error.html", {"message": "No Selection"})
#     except Job.DoesNotExist:
#         return render(request, "jobs/error.html", {"message": "Invalid File Selection. Contact Admin"})
#     pdfs=[]
#     merged_pdf = PdfFileMerger()
#     for equipment in job.equipment.all():
#         try 
#         try:
#             if equipment.test_results:
#                 PyPDF2.PdfFileReader(open(equipment.test_results, "rb"))

#         except PyPDF2.utils.PdfReadError:
#         else:
#             pass


#add fsr to a job
def create_job_site(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    
    #fill in all properties that were submitted in the form
    if "site_name" in request.POST:
            #verify uniqueness of jobsite
        if JobSite.objects.filter(name=request.POST["site_name"]):
            return render(request, "jobs/error.html", {"message": "Job Site Name is In Use. Enter A Unique One."})
    
        # extract job_name from form
        site_name = request.POST["site_name"].strip()
        #create new Job objects
        jobsite=JobSite(name=site_name)
        jobsite.save()
        jobsite_id=jobsite.id
    else:
        return render(request, "jobs/error.html", {"message": "Must Enter Job Site Name"})
    user_properties = UserProperties.objects.get(user=request.user)
    jobsite.company = user_properties.company
    if "owner" in request.POST:
        jobsite.owner = request.POST["owner"].strip()
    if "facility_manager" in request.POST:
        jobsite.facility_manager = request.POST["facility_manager"]
    if "address" in request.POST:
        jobsite.address = request.POST["address"]
    if "nav_link" in request.POST:
        jobsite.nav_link = request.POST["nav_link"]
    if "site_navigation" in request.POST:
        jobsite.site_navigation = request.POST["site_navigation"]
    if "site_contact" in request.POST:
        jobsite.site_contact = request.POST["site_contact"]
    if "site_contact_info" in request.POST:
        jobsite.site_contact_info = request.POST["site_contact_info"]
    if "other_contacts" in request.POST:
        jobsite.other_contacts = request.POST["other_contacts"]
    if "food_accomodations" in request.POST:
        jobsite.food_accomodations = request.POST["food_accomodations"]
    if "lodging_recommendations" in request.POST:
        jobsite.lodging_recommendations = request.POST["lodging_recommendations"]
    if "is_safety_training" in request.POST:
        jobsite.is_safety_training = True
    if "safety_training_item_requirements" in request.POST:
        jobsite.safety_training_item_requirements = request.POST["safety_training_item_requirements"]
    if "safety_training_time" in request.POST:
        jobsite.safety_training_time = request.POST["safety_training_time"]
    if "safety_training_location" in request.POST:
        jobsite.safety_training_location = request.POST["safety_training_location"]
    if "safety_training_procedure" in request.POST:
        jobsite.safety_training_procedure = request.POST["safety_training_procedure"]
    if "safety_rules" in request.POST:
        jobsite.safety_rules = request.POST["safety_rules"]
    if "is_hardhat" in request.POST:
        jobsite.is_hardhat = True
    if "is_safety_glasses" in request.POST:
        jobsite.is_safety_glasses = True
    if "is_safety_shoes" in request.POST:
        jobsite.is_safety_shoes = True
    if "is_safety_vest" in request.POST:
        jobsite.is_safety_vest = True
    if "is_safety_gloves" in request.POST:
        jobsite.is_safety_gloves = True
    if "is_fr_clothes" in request.POST:
        jobsite.is_fr_clothes = True
    if "is_h2s_monitor" in request.POST:
        jobsite.is_h2s_monitor = True
    if "additional_ppe_requirements" in request.POST:
        jobsite.additional_ppe_requirements  = request.POST["additional_ppe_requirements"]
    if "documents_required_for_access" in request.POST:
        jobsite.documents_required_for_access = request.POST["documents_required_for_access"]
    if "background_checks" in request.POST:
        jobsite.background_checks = True
    if "background_check_procedure" in request.POST:
        jobsite.background_check_procedure = request.POST["background_check_procedure"]
    if "access_requirements" in request.POST:
        jobsite.access_requirements = request.POST["access_requirements"]
    if "entry_procedure" in request.POST:
        jobsite.entry_procedure = request.POST["entry_procedure"]
    if "restricted_items" in request.POST:
        jobsite.restricted_items = request.POST["restricted_items"]
    if "driving_vehicles" in request.POST:
        jobsite.driving_vehicles = request.POST["driving_vehicles"]
    if "parking_considerations" in request.POST:
        jobsite.parking_considerations = request.POST["parking_considerations"]
    if "other_site_rules" in request.POST:
        jobsite.other_site_rules = request.POST["other_site_rules"]
    jobsite.save()
    return HttpResponseRedirect(reverse("job_sites"))

def add_fsr(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        fsr_id = int(request.POST["fsr"])
        fsr = UserProperties.objects.get(pk=fsr_id)
        job = Job.objects.get(pk=job_id)
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except UserProperties.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid FSR Selection."})
    except Job.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid Job Selection."})
    job.user_properties.add(fsr)

    return HttpResponseRedirect(reverse("job", args=(job_id, ))+"#fsr_add_form")

def remove_fsr(request, job_id, fsr_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
       job = Job.objects.get(pk=job_id)
       fsr = UserProperties.objects.get(pk=fsr_id)
    except UserProperties.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "FSR Removal Error."})
    except Job.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Associated Job not Identified."})
    job.user_properties.remove(fsr)
    return HttpResponseRedirect(reverse("job", args=(job_id, ))+"#fsr_add_form")

def add_equipment_files(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        equipment = Equipment.objects.get(pk=equipment_id)
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except Equipment.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid Equipment Selection."})
    if 'file_add' in request.FILES:
        file_add = request.FILES["file_add"]
        equipment_folder=EquipmentFolder(equipment=equipment)
        if 'eq_file_name' in request.POST:
            equipment_folder.file_name = request.POST["eq_file_name"]
        equipment_folder.save()
        equipment_folder.equipment_file = file_add        
        equipment_folder.author=request.user
        equipment_folder.save()
    return HttpResponseRedirect(reverse("equipment", args=(equipment_id, ))+"#eq_files_row")

def remove_equipment_files(request, file_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        equipment_file = EquipmentFolder.objects.get(pk=file_id)
        equipment_id=equipment_file.equipment.id
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except EquipmentFolder.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid File Selection. Contact Admin"})
    equipment_file.equipment_file.delete(save=False) #delete file in S3
    equipment_file.delete() #delete file in django
    return HttpResponseRedirect(reverse("equipment", args=(equipment_id, ))+"#eq_files_row")

def add_equipment_links(request, equipment_id):
    print("asldkjflksdjfl;kjfd98498416")
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        equipment = Equipment.objects.get(pk=equipment_id)
    except KeyError:
        return render(request, "jobs/error.html", {"message": "Add Link Error, Contact Admin"})
    except Equipment.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Add Link Error DNE, Contact Admin"})
    if 'link_url' in request.POST:
        link_url = request.POST["link_url"]
        equipment_link = EquipmentLink(equipment=equipment)
        equipment_link.link_url = link_url
        equipment_link.save()
        if 'eq_link_name' in request.POST:
            equipment_link.link_name = request.POST["eq_link_name"]
        equipment_link.save()
        equipment_link.author=request.user
        equipment_link.save()
    return HttpResponseRedirect(reverse("equipment", args=(equipment_id, ))+"#eq_files_row")

# def add_eq_to_jobsite_from_job(request, equipment_id):
#     try:
#         equipment = Equipment.objects.get(pk=equipment_id)
#     except KeyError:
#         return render(request, "jobs/error.html", {"message": "No Selection"})
#     except Equipment.DoesNotExist:
#         return render(request, "jobs/error.html", {"message": "Invalid Equipment Selection."})
#     try:
#         job = Job.objects.get(equipment = equipment)
#         job_site = job.job_site        
#     except KeyError:
#         return render(request, "jobs/error.html", {"message": "Contact admin. Error Code: [job retrieval error] "})
#     except Job.DoesNotExist:
#         return render(request, "jobs/error.html", {"message": "Contact admin. Error Code: [ retrieval error]"})

#     #check whether equipment with this site id is already on the jobsite. This should never have been possible to add so is a secondary safety check. 
#     if Equipment.objects.filter(job_site = job_site).filter(site_id = equipment.site_id).exists():        
#         return render(request, "jobs/error.html", {"message": "Contact admin. Error Code: [ site_equipment integrity error. duplicate(s)]"})
    

def add_site_equipment(request, job_site_id=None):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
        
    job_site = None
    try:
        job_site=JobSite.objects.get(pk = job_site_id)
        #extract equipment type from form
        
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except JobSite.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Contact Admin. Invalid JobSite Selection."})
    
    #get the type, make, and model from form
    parent_equipment = None
    if "parent_equipment_id" in request.POST:
        if request.POST["parent_equipment_id"]!="":
            try:
                parent_equipment_id = request.POST["parent_equipment_id"]
                parent_equipment = Equipment.objects.get(pk=parent_equipment_id)
                site_parent_equipment = parent_equipment.equipment_mold
            except:
                return render(request, "jobs/error.html", {"message": "Error Parent equipment not loaded. Contact Admin"})

    type_name=request.POST["eq_type"]
    equipment_type = Type.objects.get(name=type_name)
    model=request.POST["model"]
    equipment_model = Model.objects.get(pk=model)
    manufacturer_name = request.POST["manufacturer"]
    equipment_manufacturer=Manufacturer.objects.get(name=manufacturer_name)

    #get quanitity of equipment to add
    quantity = request.POST["quantity"]
    q=int(quantity)
    #get site_id of equipment
    site_id = request.POST["site_id"]
    #check if this site_id already exists in this jobsite
    if q<2:
        if Equipment.objects.filter(site_id=site_id, job_site=job_site).exists():
            return render(request, "jobs/error.html", {"message": "Already Exists."})

    #get scope, manual, notes, and test sheet
    if "test_sheet" in request.FILES:
        test_sheet=request.FILES["test_sheet"]
    else:
        test_sheet = False

    scope=request.POST["scope"]
    # pull test equipment into csv list
    mandatory_equipment_list=request.POST["mandatory_test_equipment"]
    optional_equipment_list=request.POST["optional_test_equipment"]
    if mandatory_equipment_list:
        mandatory_equipment_list=mandatory_equipment_list.split(',')
    else:
        mandatory_equipment_list=[]
   
    if optional_equipment_list:
        optional_equipment_list=optional_equipment_list.split(',')
    else:
        optional_equipment_list=[]
    
    x=0
    for each in range(q):
            
        x=x+1 #increments siteID by 1
        prog=0
        #create new objects by site ID
        new_equipment = Equipment.objects.first()
        if q==1:
            if Equipment.objects.filter(job_site = job_site, site_id = site_id).exists():
                return render(request, "jobs/error.html", {"message": "Job Site equipment add failure, Site ID already exists on this job site"})
            else:
                new_equipment=Equipment.objects.create_equipment(site_id)
        else:            
            site_id_it = site_id + " - "+str(x)
            #the below loop checks if there is already equipment with the provided site id and appends a dash and number. 
            #If the dash and number have already been added, it adds another. Up to 19 copies, at which point, user must rename some of the existing copies.
            for level in range(20):
                if Equipment.objects.filter(job_site = job_site, site_id = site_id_it).exists():
                    if level == 19:
                        return render(request, "jobs/error.html", {"message": "Error, Equipment NOT added. Max Copies reached at 19 copies. You must change the names of existing equipment with site ID: "+site_id})
                    site_id_it += " - "+str(x)
                else:
                    break
            new_equipment=Equipment.objects.create_equipment(site_id_it)

        #link type, model, and manfacturer
        if equipment_type.is_test_sheet == True:
            new_equipment.is_insulation_resistance=equipment_type.is_insulation_resistance
            new_equipment.is_contact_resistance=equipment_type.is_contact_resistance
            new_equipment.is_trip_unit=equipment_type.is_trip_unit
            new_equipment.is_primary_injection=equipment_type.is_primary_injection
            new_equipment.is_secondary_injection=equipment_type.is_secondary_injection
            new_equipment.is_xfmr_insulation_resistance=equipment_type.is_xfmr_insulation_resistance
            new_equipment.is_power_fused=equipment_type.is_power_fused
            new_equipment.is_breaker=equipment_type.is_breaker
            new_equipment.is_hipot=equipment_type.is_hipot
            new_equipment.is_inspection=equipment_type.is_inspection
            new_equipment.is_transformer=equipment_type.is_transformer
            new_equipment.is_winding_resistance=equipment_type.is_winding_resistance
            new_equipment.is_ttr=equipment_type.is_ttr
            new_equipment.is_liquid_type=equipment_type.is_liquid_type
            new_equipment.is_cable=equipment_type.is_cable
            new_equipment.is_cable_vlf_withstand_test=equipment_type.is_cable_vlf_withstand_test
            new_equipment.save()
            
            try:
                blank_test_sheet=TestSheet(eq=new_equipment)      
                blank_test_sheet.save()
            except:
                blank_test_sheet = None
                equipment_type.is_test_sheet = False
                equipment_type.save()
                return render(request, "jobs/error.html", {"message": "Error generating new test sheet. 'is_test_sheet' set to 'False' for eq type."})
        
        new_equipment.save()
        new_equipment.job_site = job_site 
        new_equipment.equipment_type = equipment_type
        new_equipment.equipment_model = equipment_model
        
        if parent_equipment:
            new_equipment.parent_equipment = parent_equipment
        new_equipment.scope = scope
        if test_sheet != False:
            new_equipment.test_sheet_template=test_sheet
        elif equipment_model.model_test_sheet:
            new_equipment.test_sheet_template=equipment_model.model_test_sheet
        elif equipment_type.test_sheet:
            new_equipment.test_sheet_template=equipment_type.test_sheet

        #fill in all properties that were submitted in the form
        new_equipment.save()

        if mandatory_equipment_list:
            for equipment_name in mandatory_equipment_list:
                
                try:
                    equipment=TestEquipment.objects.get(name=equipment_name)
                    new_equipment.mandatory_test_equipment.add(equipment.pk)
                    
                except TestEquipment.DoesNotExist:
                    return HttpResponseRedirect(reverse("profile"))
                    
        if optional_equipment_list:
            for equipment_name in optional_equipment_list:
                try:
                    equipment=TestEquipment.objects.get(name=equipment_name)
                    new_equipment.optional_test_equipment.add(equipment.pk)
                    
                except TestEquipment.DoesNotExist:
                    return HttpResponseRedirect(reverse("profile"))

        new_equipment.save()


    
    return HttpResponseRedirect(reverse("job_site", args=(job_site.id,)) + "#equipment")

def note_notification(new_note):
    from django.conf import settings
    domain = None
    if settings.DEBUG:
        #domain for testing on localhost
        domain = 'http://127.0.0.1:8000'
    else:
        #domain for production
        domain = 'https://paragon.blueskysw.com/'
    
    
    is_eq = False
    is_job = False
    if isinstance(new_note, EquipmentNotes):
        is_eq = True
        print('eq')
    elif isinstance(new_note, JobNotes):
        is_job = True
        print('jobnote')
    else:
        print("EMAILING ERROR IN NOTE_NOTIFICATION")
        return None
    context = {
        'is_eq': is_eq,
        'is_job':is_job,
        'domain': domain,
        'note': new_note
    }
    body = render_to_string('jobs/email_notification.html',context)
    return body
    
def send_notification(user_email, subject, contents):
    print('lvl 2')
    email = EmailMessage(
        subject,
        contents,
        'info@blueskysw.com',
        [user_email],
        )
    email.content_subtype = 'html'
    email.send(fail_silently=False)
    print('lvl 2 done')


def notification_threader(user_email, subject, contents):
    print('start')
    t1 = threading.Thread(target=send_notification, args=[user_email, subject, contents],daemon=True)
    t1.start()
    print('finish')

def add_equipment(request, job_id=None):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        job = Job.objects.get(pk=job_id)    
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except Job.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid Job Selection."})

    parent_equipment = None
    if "parent_equipment_id" in request.POST:
        if request.POST["parent_equipment_id"]!="":
            try:
                parent_equipment_id = request.POST["parent_equipment_id"]
                parent_equipment = Equipment.objects.get(pk=parent_equipment_id)
            except:
                return render(request, "jobs/error.html", {"message": "Error Parent equipment not loaded. Contact Admin"})

    type_name=request.POST["eq_type"]
    equipment_type = Type.objects.get(name=type_name)
    model=request.POST["model"]
    equipment_model = Model.objects.get(pk=model)
    manufacturer_name = request.POST["manufacturer"]
    equipment_manufacturer=Manufacturer.objects.get(name=manufacturer_name)
    user_properties = UserProperties.objects.get(user=request.user)
    
    #get quanitity of equipment to add
    quantity = request.POST["quantity"]
    q=int(quantity)
    #get site_id of equipment
    site_id = request.POST["site_id"]
    #check if this site_id already exists in this job or in jobsite
    
    job_check = job.equipment.filter(site_id=site_id)
    if job_check:
        return render(request, "jobs/error.html", {"message": "Already exists. Must select a unique site_id"})
    if Equipment.objects.filter(site_id=site_id, job_site=job.job_site).exists():        
        return render(request, "jobs/error.html", {"message": "Already Exists. Add this eq from the Job Site"})
    #get scope, manual, notes, and test sheet    
    if "test_sheet" in request.FILES:
        test_sheet=request.FILES["test_sheet"]
    else:
        test_sheet = False

    scope=request.POST["scope"]
    # pull test equipment into csv list
    mandatory_equipment_list=request.POST["mandatory_test_equipment"]
    optional_equipment_list=request.POST["optional_test_equipment"]
    if mandatory_equipment_list:
        mandatory_equipment_list=mandatory_equipment_list.split(',')
    else:
        mandatory_equipment_list=[]
   
    if optional_equipment_list:
        optional_equipment_list=optional_equipment_list.split(',')
    else:
        optional_equipment_list=[]
    
    x=0
    for each in range(q):
        x=x+1 #increments siteID by 1
        prog=0
        #create new objects by site ID
        new_equipment = Equipment.objects.first()
        if q==1:
            if Equipment.objects.filter(equipments = job, site_id = site_id).exists():
                return render(request, "jobs/error.html", {"message": "Error: Job equipment add failure, equipment is already added. Each Site ID must be unique to distinguish each one."})
            else:
                if Equipment.objects.filter(job_site = job.job_site, site_id = site_id).exists():
                    return render(request, "jobs/error.html", {"message": "Error: Job equipment add failure, equipment is already added to the job site. Add it from there."})
                new_equipment=Equipment.objects.create_equipment(site_id)
        else:  
            return render(request, "jobs/error.html", {"message": "Error: To add multiple equipment at once, open the job site and add them there. Then add them to the job from the job site." })
               
        try:                
            blank_test_sheet=TestSheet(eq=new_equipment)
            blank_test_sheet.save()   
        except:
            blank_test_sheet = None
            equipment_type.is_test_sheet = False
            equipment_type.save()
            return render(request, "jobs/error.html", {"message": "Error generating new test sheet. 'is_test_sheet' set to 'False' for eq type."}) 
            
        #link type, model, and manfacturer
        if equipment_type.is_test_sheet == True:
            new_equipment.is_insulation_resistance=equipment_type.is_insulation_resistance
            new_equipment.is_contact_resistance=equipment_type.is_contact_resistance
            new_equipment.is_trip_unit=equipment_type.is_trip_unit
            new_equipment.is_primary_injection=equipment_type.is_primary_injection
            new_equipment.is_secondary_injection=equipment_type.is_secondary_injection
            new_equipment.is_xfmr_insulation_resistance=equipment_type.is_xfmr_insulation_resistance
            new_equipment.is_power_fused=equipment_type.is_power_fused
            new_equipment.is_breaker=equipment_type.is_breaker
            new_equipment.is_hipot=equipment_type.is_hipot
            new_equipment.is_inspection=equipment_type.is_inspection
            new_equipment.is_transformer=equipment_type.is_transformer
            new_equipment.is_winding_resistance=equipment_type.is_winding_resistance
            new_equipment.is_ttr=equipment_type.is_ttr
            new_equipment.is_liquid_type=equipment_type.is_liquid_type
            new_equipment.is_cable=equipment_type.is_cable
            new_equipment.is_cable_vlf_withstand_test=equipment_type.is_cable_vlf_withstand_test
            new_equipment.save()

        new_equipment.save()
        new_eq_key = new_equipment.pk
        site_equipment=new_equipment
        site_equipment.pk = None
        site_equipment.save()
        site_test_sheet=TestSheet(eq=new_equipment)
        site_test_sheet.save()

        new_equipment = Equipment.objects.get(pk=new_eq_key)
        job.equipment.add(new_equipment)
        job.save()
        #add to notifications
    # try:
        print(user_properties.equipment_notifications.all())
        user_properties.equipment_notifications.add(new_equipment)
        user_properties.save()
        users_to_notify = job.job_notifications_user.all()
        subject = "Paragon Equipment Added To" + job.job_name
        contents = "Equipment with site I.D. <b><u>" + new_equipment.site_id + "</u></b> has been added to the job <u>" + job.job_name + "</u> by " + str(user_properties)
        for user_i in users_to_notify:
            notification_threader(user_i.user.email, subject, contents)
    # except:
    #     print("failed notification")
        
        site_equipment.job_site = job.job_site 

        new_equipment.equipment_type = equipment_type
        site_equipment.equipment_type = equipment_type
        new_equipment.equipment_model = equipment_model
        site_equipment.equipment_model = equipment_model
        if parent_equipment:
            new_equipment.parent_equipment = parent_equipment
            site_equipment.parent_equipment = new_equipment.parent_equipment.equipment_mold
        new_equipment.scope = scope
        if test_sheet:
            new_equipment.test_sheet_template=test_sheet
        elif equipment_model.model_test_sheet:
            new_equipment.test_sheet_template=equipment_model.model_test_sheet
        elif equipment_type.test_sheet:
            new_equipment.test_sheet_template=equipment_type.test_sheet

        #fill in all properties that were submitted in the form
        new_equipment.save()
        site_equipment.job_site = job.job_site
        site_equipment.save()

        if mandatory_equipment_list:
            for equipment_name in mandatory_equipment_list:
                
                try:
                    equipment=TestEquipment.objects.get(name=equipment_name)
                    new_equipment.mandatory_test_equipment.add(equipment.pk)
                    if job_id:
                        job.test_equipment.remove(equipment.pk)
                    
                except TestEquipment.DoesNotExist:
                    return HttpResponseRedirect(reverse("profile"))
                    
        if optional_equipment_list:
            for equipment_name in optional_equipment_list:
                try:
                    equipment=TestEquipment.objects.get(name=equipment_name)
                    new_equipment.mandatory_test_equipment.add(equipment.pk)
                    if job_id:
                        job.test_equipment.remove(equipment.pk)
                    
                except TestEquipment.DoesNotExist:
                    return HttpResponseRedirect(reverse("profile"))

        new_equipment.save()


        new_equipment.equipment_mold = site_equipment
        new_equipment.save()

    return HttpResponseRedirect(reverse("job", args=(job_id, ))+"#equipment")

def download_xlsx(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})

    wb = load_workbook('MV Breaker TS.xlsx')
    ws = wb.active
    
    equipment=Equipment.objects.get(pk=equipment_id)
    test_sheet = TestSheet.objects.get(eq=equipment)
    job=Job.objects.get(equipment=equipment)
    siteid=""
    serial=""
    eqlocation=""
    mfr=""
    eqmodel=""
    eqtype=""
    framerating=""
    mountstyle=""
    controlv=""
    controldc="AC"
    tripv=""
    tripdc="AC"
    tester=""
    testdate="2022-01-01"
    phphv=""
    phgv=""
    llv=""
    ab=""
    bc=""
    ca=""
    ag=""
    bg=""
    cg=""
    aa=""
    bb=""
    cc=""
    contactamps=""
    a=""
    b=""
    c=""
    hipotv=""
    hab=""
    hbc=""
    hca=""
    hag=""
    hbg=""
    hcg=""
    haa=""
    hbb=""
    hcc=""
    if equipment:
        siteid=equipment.site_id
        if equipment.serial_number:
            serial=equipment.serial_number
        if equipment.equipment_location:
            eqlocation=equipment.equipment_location
        if equipment.equipment_model:
            eqmodel=equipment.equipment_model.name
            if equipment.equipment_model.model_manufacturer:
                mfr=equipment.equipment_model.model_manufacturer.name
            if equipment.equipment_model.model_type:
                eqtype=equipment.equipment_model.model_type.name
        if test_sheet.mount_style:
            mountstyle=test_sheet.mount_style
        if test_sheet.frame_size:
            framerating=test_sheet.frame_size
        if test_sheet.control_voltage:
            controlv=test_sheet.control_voltage
        if test_sheet.is_dc_control_voltage:
            ws['C12'].value = 'Volts DC'    
        else:
            ws['C12'].value = 'Volts AC'
        if test_sheet.trip_coil_voltage:
            tripv=test_sheet.trip_coil_voltage
        if test_sheet.is_dc_trip_voltage:
            ws['C13'].value = 'Volts DC'    
        else:
            ws['C13'].value = 'Volts AC'
        if test_sheet.testers:
            tester=test_sheet.testers
        if test_sheet.date_tested:
            testdate=test_sheet.date_tested
        if test_sheet.insulation_resistance_ph_to_ph_test_voltage:
            phphv=test_sheet.insulation_resistance_ph_to_ph_test_voltage
        if test_sheet.insulation_resistance_ph_to_gr_test_voltage:
            phgv=test_sheet.insulation_resistance_ph_to_gr_test_voltage
        if test_sheet.insulation_resistance_ln_to_ld_test_voltage:
            llv=test_sheet.insulation_resistance_ln_to_ld_test_voltage
        if test_sheet.insulation_resistance_ph_to_ph_a_b:
            ab=test_sheet.insulation_resistance_ph_to_ph_a_b
        if test_sheet.insulation_resistance_ph_to_ph_a_b_units:
            ws['G5'].value = test_sheet.insulation_resistance_ph_to_ph_a_b_units
        if test_sheet.insulation_resistance_ph_to_ph_b_c:
            bc=test_sheet.insulation_resistance_ph_to_ph_b_c
        if test_sheet.insulation_resistance_ph_to_ph_b_c_units:
            ws['I5'].value = test_sheet.insulation_resistance_ph_to_ph_b_c_units
        if test_sheet.insulation_resistance_ph_to_ph_c_a:
            ca=test_sheet.insulation_resistance_ph_to_ph_c_a
        if test_sheet.insulation_resistance_ph_to_ph_c_a_units:
            ws['K5'].value = test_sheet.insulation_resistance_ph_to_ph_c_a_units
        if test_sheet.insulation_resistance_ph_to_gr_a_g:
            ag=test_sheet.insulation_resistance_ph_to_gr_a_g
        if test_sheet.insulation_resistance_ph_to_gr_a_g_units:
            ws['G7'].value = test_sheet.insulation_resistance_ph_to_gr_a_g_units
        if test_sheet.insulation_resistance_ph_to_gr_b_g:
            bg=test_sheet.insulation_resistance_ph_to_gr_b_g
        if test_sheet.insulation_resistance_ph_to_gr_b_g_units:
            ws['I7'].value = test_sheet.insulation_resistance_ph_to_gr_b_g_units
        if test_sheet.insulation_resistance_ph_to_gr_c_g:
            cg=test_sheet.insulation_resistance_ph_to_gr_c_g
        if test_sheet.insulation_resistance_ph_to_gr_c_g_units:
            ws['K7'].value = test_sheet.insulation_resistance_ph_to_gr_c_g_units
        if test_sheet.insulation_resistance_ln_to_ld_a:
            aa=test_sheet.insulation_resistance_ln_to_ld_a
        if test_sheet.insulation_resistance_ln_to_ld_a_units:
            ws['G9'].value = test_sheet.insulation_resistance_ln_to_ld_a_units
        if test_sheet.insulation_resistance_ln_to_ld_b:
            bb=test_sheet.insulation_resistance_ln_to_ld_b
        if test_sheet.insulation_resistance_ln_to_ld_b_units:
            ws['I9'].value = test_sheet.insulation_resistance_ln_to_ld_b_units
        if test_sheet.insulation_resistance_ln_to_ld_c:
            cc=test_sheet.insulation_resistance_ln_to_ld_c
        if test_sheet.insulation_resistance_ln_to_ld_c_units:
            ws['K9'].value = test_sheet.insulation_resistance_ln_to_ld_c_units
        if test_sheet.contact_resistance_current:
            contactamps=test_sheet.contact_resistance_current
        if test_sheet.contact_resistance_a:
            a=test_sheet.contact_resistance_a
        if test_sheet.contact_resistance_a_units:
            ws['G13'].value = test_sheet.contact_resistance_a_units
        if test_sheet.contact_resistance_b:
            b=test_sheet.contact_resistance_b
        if test_sheet.contact_resistance_b_units:
            ws['I13'].value = test_sheet.contact_resistance_b_units
        if test_sheet.contact_resistance_c:
            c=test_sheet.contact_resistance_c
        if test_sheet.contact_resistance_c_units:
            ws['K13'].value = test_sheet.contact_resistance_c_units
        if test_sheet.hipot_test_voltage:
            hipotv=test_sheet.hipot_test_voltage
        if test_sheet.hipot_ptp_ab:
            hab=test_sheet.hipot_ptp_ab
        if test_sheet.hipot_ptp_bc:
            hbc=test_sheet.hipot_ptp_bc
        if test_sheet.hipot_ptp_ca:
            hca=test_sheet.hipot_ptp_ca
        if test_sheet.hipot_ptg_a:
            hag=test_sheet.hipot_ptg_a
        if test_sheet.hipot_ptg_b:
            hbg=test_sheet.hipot_ptg_b
        if test_sheet.hipot_ptg_c:
            hcg=test_sheet.hipot_ptg_c
        if test_sheet.hipot_ltl_a:
            haa=test_sheet.hipot_ltl_a
        if test_sheet.hipot_ltl_b:
            hbb=test_sheet.hipot_ltl_b
        if test_sheet.hipot_ltl_c:
            hcc=test_sheet.hipot_ltl_c

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="MV Breaker TS.xlsx"'

    ws['A1'].value = 'Job: '+job.job_name
    ws['F5'].value = ab
    ws['B4'].value = siteid
    ws['B5'].value = serial
    ws['B6'].value = eqlocation
    ws['B7'].value = mfr
    ws['B8'].value = eqtype
    ws['B9'].value = eqmodel
    ws['B10'].value = framerating
    ws['B11'].value = mountstyle
    ws['B12'].value = controlv
    ws['B13'].value = tripv
    ws['B14'].value = tester
    ws['B15'].value = testdate
    ws['E5'].value = phphv
    ws['E7'].value = phgv
    ws['E9'].value = llv
    ws['F5'].value = ab
    ws['H5'].value = bc
    ws['J5'].value = ca
    ws['F7'].value = ag
    ws['H7'].value = bg
    ws['J7'].value = cg
    ws['F9'].value = aa
    ws['H9'].value = bb
    ws['J9'].value = cc
    ws['E13'].value = contactamps
    ws['F13'].value = a
    ws['H13'].value = b
    ws['J13'].value = c
    ws['E17'].value = hipotv
    ws['F17'].value = hab
    ws['H17'].value = hbc
    ws['J17'].value = hca
    ws['F19'].value = hag
    ws['H19'].value = hbg
    ws['J19'].value = hcg
    ws['F21'].value = haa
    ws['H21'].value = hbb
    ws['J21'].value = hcc
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888
    # ws['88'].value = 8888

    wb.save(response)


    return response
def upload_sync_xlsx(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    wb = None
    blank = ''
    is_blank = False
    is_value_error = False
    value_error = ''

    if "xlsx_file" in request.FILES:
        wb = load_workbook(request.FILES["xlsx_file"])
        ws = wb.active
        eq=Equipment.objects.get(pk=equipment_id)
        test_sheet = TestSheet.objects.get(eq=eq)
        job=Job.objects.get(equipment=eq)
                
        siteid=""
        serial=""
        eqlocation=""
        mfr=""
        eqmodel=""
        eqtype=""
        framerating=""
        mountstyle=""
        controlv=""
        controldc="AC"
        tripv=""
        tripdc="AC"
        tester=""
        testdate="2022-01-01"
        phphv=""
        phgv=""
        llv=""
        ab=""
        bc=""
        ca=""
        ag=""
        bg=""
        cg=""
        aa=""
        bb=""
        cc=""
        contactamps=""
        a=""
        b=""
        c=""
        hipotv=""
        hab=""
        hbc=""
        hca=""
        hag=""
        hbg=""
        hcg=""
        haa=""
        hbb=""
        hcc=""
        if test_sheet:

            if ws['B5'].value:
                try:
                    if eq.equipment_mold:
                        eq.equipment_mold.serial_number = ws['B5'].value
                        eq.serial_number = ws['B5'].value
                    else:
                        eq.serial_number = ws['B5'].value
                except:
                    value_error=value_error + 'Serial Number, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Serial Number, '
                is_blank = True
            if ws['B6'].value:
                try:
                    if eq.equipment_mold:
                        eq.equipment_mold.equipment_location = ws['B6'].value
                        eq.equipment_location = ws['B6'].value
                    else:
                        eq.equipment_location = ws['B6'].value
                except:
                    value_error=value_error + 'Location, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Location, '
                is_blank = True
            if ws['B10'].value:
                try:
                    eq.equipment_mold.sheet_eq.frame_size = ws['B10'].value
                    test_sheet.frame_size = ws['B10'].value
                except:
                    value_error=value_error + 'Frame Rating, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Frame Rating, '
                is_blank = True
            if ws['B11'].value:
                try:
                    eq.equipment_mold.sheet_eq.mount_style = ws['B11'].value
                    test_sheet.mount_style = ws['B11'].value

                except:
                    value_error=value_error + 'Mount Style, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Mount Style, '
                is_blank = True
            if ws['B12'].value:
                try:
                    eq.equipment_mold.sheet_eq.control_voltage = ws['B12'].value
                    test_sheet.control_voltage = ws['B12'].value
                except:
                    value_error=value_error + 'Control Voltage, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Control Voltage, '
                is_blank = True
            if ws['C12'].value:
                try:
                    if str(ws['C12'].value) == 'Volts DC':
                        print(1)
                        eq.equipment_mold.sheet_eq.is_dc_control_voltage = True
                        test_sheet.is_dc_control_voltage = True
                        print(eq.equipment_mold.sheet_eq.is_dc_control_voltage)
                        print(test_sheet.is_dc_control_voltage)
                    else:
                        print(2)
                        eq.equipment_mold.sheet_eq.is_dc_control_voltage = False
                        test_sheet.is_dc_control_voltage = False
                except:
                    value_error=value_error + 'Control Voltage Volt Type, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Control Voltage Volt Type, '
                is_blank = True
            if ws['B13'].value:
                try:
                    eq.equipment_mold.sheet_eq.trip_coil_voltage = ws['B13'].value
                    test_sheet.trip_coil_voltage = ws['B13'].value
                except:
                    value_error=value_error + 'Trip Voltage, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Trip Voltage, '
                is_blank = True
            if ws['C13'].value:
                try:
                    if ws['C13'].value == 'Volts DC':
                        eq.equipment_mold.sheet_eq.is_dc_trip_voltage = True
                        test_sheet.is_dc_trip_voltage = True
                    else:
                        eq.equipment_mold.sheet_eq.is_dc_trip_voltage = False
                        test_sheet.is_dc_trip_voltage = False
                except:
                    value_error=value_error + 'Trip Voltage Volt Type, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Trip Voltage Volt Type, '
                is_blank = True

            if ws['B14'].value:
                try:
                    test_sheet.testers = ws['B14'].value
                except:
                    value_error=value_error + 'Testers, '
                    is_value_error = True
                    pass           
            else:
                blank = blank + 'Testers, '
                is_blank = True
            if ws['E5'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_ph_test_voltage = ws['E5'].value     
                except:
                    value_error=value_error + 'Insulation Resistance Ph-Ph @volts, '
                    is_value_error = True
                    pass           
            else:
                blank = blank + 'Insulation Resistance Ph-Ph @volts, '
                is_blank = True
            if ws['F5'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_ph_a_b = ws['F5'].value     
                except:
                    value_error=value_error + 'Insulation Resistance A-B, '
                    is_value_error = True
                    pass           
            else:
                blank = blank + 'Insulation Resistance A-B, '
                is_blank = True
            if ws['H5'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_ph_b_c = ws['H5'].value     
                except:
                    value_error=value_error + 'Insulation Resistance B-C, '
                    is_value_error = True
                    pass           
            else:
                blank = blank + 'Insulation Resistance B-C, '
                is_blank = True
            if ws['J5'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_ph_c_a = ws['J5'].value
                except:
                    value_error=value_error + 'Insulation Resistance C-A, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance C-A, '
                is_blank = True
            if ws['G5'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_ph_a_b_units = ws['G5'].value
                except:
                    value_error=value_error + 'Insulation Resistance A-B Units, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance A-B Units, '
                is_blank = True
            if ws['I5'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_ph_b_c_units = ws['I5'].value
                except:
                    value_error=value_error + 'Insulation Resistance B-C Units, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance B-C Units, '
                is_blank = True
            if ws['K5'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_ph_c_a_units = ws['K5'].value
                except:
                    value_error=value_error + 'Insulation Resistance C-A Units, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance C-A Units, '
                is_blank = True
            if ws['E9'].value:
                try:
                    test_sheet.insulation_resistance_ln_to_ld_test_voltage = ws['E9'].value
                except:
                    value_error=value_error + 'Insulation Resistance Line-Load @volts, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance Line-Load @volts, '
                is_blank = True
            if ws['F9'].value:
                try:
                    test_sheet.insulation_resistance_ln_to_ld_a = ws['F9'].value
                except:
                    value_error=value_error + 'Insulation Resistance A-A, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance A-A, '
                is_blank = True
            if ws['G9'].value:
                try:
                    test_sheet.insulation_resistance_ln_to_ld_a_units = ws['G9'].value
                except:
                    value_error=value_error + 'Insulation Resistance A-A Units, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance A-A Units, '
                is_blank = True
            if ws['H9'].value:
                try:
                    test_sheet.insulation_resistance_ln_to_ld_b = ws['H9'].value
                except:
                    value_error=value_error + 'Insulation Resistance B-B, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance B-B, '
                is_blank = True
            if ws['I9'].value:
                try:
                    test_sheet.insulation_resistance_ln_to_ld_b_units = ws['I9'].value
                except:
                    value_error=value_error + 'Insulation Resistance A-A Units, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance A-A Units, '
                is_blank = True
            if ws['J9'].value:
                try:
                    test_sheet.insulation_resistance_ln_to_ld_c = ws['J9'].value
                except:
                    value_error=value_error + 'Insulation Resistance C-C, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance C-C, '
                is_blank = True
            if ws['K9'].value:
                try:
                    test_sheet.insulation_resistance_ln_to_ld_c_units = ws['K9'].value
                except:
                    value_error=value_error + 'Insulation Resistance C-C Units, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance C-C Units, '
                is_blank = True
            if ws['E7'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_gr_test_voltage = ws['E7'].value
                except:
                    value_error=value_error + 'Insulation Resistance Ph-G @volts, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance Ph-G @volts, '
                is_blank = True
            if ws['F7'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_gr_a_g = ws['F7'].value
                except:
                    value_error=value_error + 'Insulation Resistance A-G, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance A-A, '
                is_blank = True
            if ws['G7'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_gr_a_g_units = ws['G7'].value
                except:
                    value_error=value_error + 'Insulation Resistance A-G Units, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance A-G Units, '
                is_blank = True
            if ws['H7'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_gr_b_g = ws['H7'].value
                except:
                    value_error=value_error + 'Insulation Resistance B-G, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance B-G, '
                is_blank = True
            if ws['I7'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_gr_b_g_units = ws['I7'].value
                except:
                    value_error=value_error + 'Insulation Resistance B-G Units, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance B-G Units, '
                is_blank = True
            if ws['J7'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_gr_c_g = ws['J7'].value
                except:
                    value_error=value_error + 'Insulation Resistance C-G, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance C-G, '
                is_blank = True
            if ws['K7'].value:
                try:
                    test_sheet.insulation_resistance_ph_to_gr_c_g_units = ws['K7'].value
                except:
                    value_error=value_error + 'Insulation Resistance C-G Units, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Insulation Resistance C-G Units, '
                is_blank = True
            if ws['E13'].value:
                try:
                    test_sheet.contact_resistance_current = ws['E13'].value
                except:
                    value_error=value_error + 'Contact Resistance Current, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Contact Resistance Current, '
                is_blank = True
            if ws['F13'].value:
                try:
                    test_sheet.contact_resistance_a = ws['F13'].value
                except:
                    value_error=value_error + 'Contact Resistance A, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Contact Resistance A, '
                is_blank = True
            if ws['G13'].value:
                try:
                    test_sheet.contact_resistance_a_units = ws['G13'].value
                except:
                    value_error=value_error + 'Contact Resistance A Units, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Contact Resistance A Units, '
                is_blank = True
            if ws['H13'].value:
                try:
                    test_sheet.contact_resistance_b = ws['H13'].value
                except:
                    value_error=value_error + 'Contact Resistance B, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Contact Resistance B, '
                is_blank = True
            if ws['I13'].value:
                try:
                    test_sheet.contact_resistance_b_units = ws['I13'].value
                except:
                    value_error=value_error + 'Contact Resistance B Units, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Contact Resistance B Units, '
                is_blank = True
            if ws['J13'].value:
                try:
                    test_sheet.contact_resistance_c = ws['J13'].value
                except:
                    value_error=value_error + 'Contact Resistance C, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Contact Resistance C, '
                is_blank = True
            if ws['K13'].value:
                try:
                    test_sheet.contact_resistance_c_units = ws['K13'].value
                except:
                    value_error=value_error + 'Contact Resistance C Units, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'Contact Resistance C Units, '
                is_blank = True
            if ws['E17'].value:
                try:
                    test_sheet.hipot_test_voltage = ws['E17'].value     
                except:
                    value_error=value_error + 'HiPot Ph-Ph @volts, '
                    is_value_error = True
                    pass           
            else:
                blank = blank + 'HiPot Ph-Ph @volts, '
                is_blank = True
            if ws['F17'].value:
                try:
                    test_sheet.hipot_ptp_ab = ws['F17'].value     
                except:
                    value_error=value_error + 'HiPot A-B, '
                    is_value_error = True
                    pass           
            else:
                blank = blank + 'HiPot A-B, '
                is_blank = True
            if ws['H17'].value:
                try:
                    test_sheet.hipot_ptp_bc = ws['H17'].value     
                except:
                    value_error=value_error + 'HiPot B-C, '
                    is_value_error = True
                    pass           
            else:
                blank = blank + 'HiPot B-C, '
                is_blank = True
            if ws['J17'].value:
                try:
                    test_sheet.hipot_ptp_ca = ws['J17'].value
                except:
                    value_error=value_error + 'HiPot C-A, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'HiPot C-A, '
                is_blank = True
            if ws['F21'].value:
                try:
                    test_sheet.hipot_ltl_a = ws['F21'].value
                except:
                    value_error=value_error + 'HiPot A-A, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'HiPot A-A, '
                is_blank = True
            if ws['H21'].value:
                try:
                    test_sheet.hipot_ltl_b = ws['H21'].value
                except:
                    value_error=value_error + 'HiPot B-B, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'HiPot B-B, '
                is_blank = True
            if ws['J21'].value:
                try:
                    test_sheet.hipot_ltl_c = ws['J21'].value
                except:
                    value_error=value_error + 'HiPot C-C, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'HiPot C-C, '
                is_blank = True
            if ws['F19'].value:
                try:
                    test_sheet.hipot_ptg_a = ws['F19'].value
                except:
                    value_error=value_error + 'HiPot A-G, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'HiPot A-A, '
                is_blank = True
            if ws['H19'].value:
                try:
                    test_sheet.hipot_ptg_b = ws['H19'].value
                except:
                    value_error=value_error + 'HiPot B-G, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'HiPot B-G, '
                is_blank = True
            if ws['J19'].value:
                try:
                    test_sheet.hipot_ptg_c = ws['J19'].value
                except:
                    value_error=value_error + 'HiPot C-G, '
                    is_value_error = True
                    pass
            else:
                blank = blank + 'HiPot C-G, '
                is_blank = True





            test_sheet.save()
            eq.save()           
            eq.equipment_mold.save() 
            eq.equipment_mold.sheet_eq.save()
    blank_message = 'These spreadsheet fields were empty and were NOT synchronized: '+blank
    value_error_message = 'WARNING: Some fields did NOT synchronize because of the input type; consider entering them manually: '+value_error
    is_blank = False #need to remove this once a handler or alert system is created to deal with blank cells
    if is_blank or is_value_error:
        errormessage = 'Unknown Error: Contact Admin'
        if is_blank and is_value_error:
            errormessage = blank_message+"\n"+value_error_message
        elif is_blank and not is_value_error:
            errormessage = blank_message
        elif is_value_error and not is_blank:
            errormessage = value_error_message
        return render(request, "jobs/error.html", {"message": errormessage})


    return HttpResponseRedirect(reverse("equipment", args=(equipment_id, ))) 
            # if equipment.equipment_location:
            #     eqlocation=equipment.equipment_location
            # if equipment.equipment_model:
            #     eqmodel=equipment.equipment_model.name
            #     if equipment.equipment_model.model_manufacturer:
            #         mfr=equipment.equipment_model.model_manufacturer.name
            #     if equipment.equipment_model.model_type:
            #         eqtype=equipment.equipment_model.model_type.name
            # if test_sheet.mount_style:
            #     mountstyle=test_sheet.mount_style
            # if test_sheet.frame_size:
            #     framerating=test_sheet.frame_size
            # if test_sheet.control_voltage:
            #     controlv=test_sheet.control_voltage
            # if test_sheet.is_dc_control_voltage:
            #     ws['C12'].value = 'Volts DC'    
            # else:
            #     ws['C12'].value = 'Volts AC'
            # if test_sheet.trip_coil_voltage:
            #     tripv=test_sheet.trip_coil_voltage
            # if test_sheet.is_dc_trip_voltage:
            #     ws['C13'].value = 'Volts DC'    
            # else:
            #     ws['C13'].value = 'Volts AC'
            # if test_sheet.testers:
            #     tester=test_sheet.testers
            # if test_sheet.date_tested:
            #     testdate=test_sheet.date_tested
            # if test_sheet.insulation_resistance_ph_to_ph_test_voltage:
            #     phphv=test_sheet.insulation_resistance_ph_to_ph_test_voltage
            # if test_sheet.insulation_resistance_ph_to_gr_test_voltage:
            #     phgv=test_sheet.insulation_resistance_ph_to_gr_test_voltage
            # if test_sheet.insulation_resistance_ln_to_ld_test_voltage:
            #     llv=test_sheet.insulation_resistance_ln_to_ld_test_voltage
            # if test_sheet.insulation_resistance_ph_to_ph_a_b:
            #     ab=test_sheet.insulation_resistance_ph_to_ph_a_b
            # if test_sheet.insulation_resistance_ph_to_ph_a_b_units:
            #     ws['G5'].value = test_sheet.insulation_resistance_ph_to_ph_a_b_units
            # if test_sheet.insulation_resistance_ph_to_ph_b_c:
            #     bc=test_sheet.insulation_resistance_ph_to_ph_b_c
            # if test_sheet.insulation_resistance_ph_to_ph_b_c_units:
            #     ws['I5'].value = test_sheet.insulation_resistance_ph_to_ph_b_c_units
            # if test_sheet.insulation_resistance_ph_to_ph_c_a:
            #     ca=test_sheet.insulation_resistance_ph_to_ph_c_a
            # if test_sheet.insulation_resistance_ph_to_ph_c_a_units:
            #     ws['K5'].value = test_sheet.insulation_resistance_ph_to_ph_c_a_units
            # if test_sheet.insulation_resistance_ph_to_gr_a_g:
            #     ag=test_sheet.insulation_resistance_ph_to_gr_a_g
            # if test_sheet.insulation_resistance_ph_to_gr_a_g_units:
            #     ws['G7'].value = test_sheet.insulation_resistance_ph_to_gr_a_g_units
            # if test_sheet.insulation_resistance_ph_to_gr_b_g:
            #     bg=test_sheet.insulation_resistance_ph_to_gr_b_g
            # if test_sheet.insulation_resistance_ph_to_gr_b_g_units:
            #     ws['I7'].value = test_sheet.insulation_resistance_ph_to_gr_b_g_units
            # if test_sheet.insulation_resistance_ph_to_gr_c_g:
            #     cg=test_sheet.insulation_resistance_ph_to_gr_c_g
            # if test_sheet.insulation_resistance_ph_to_gr_c_g_units:
            #     ws['K7'].value = test_sheet.insulation_resistance_ph_to_gr_c_g_units
            # if test_sheet.insulation_resistance_ln_to_ld_a:
            #     aa=test_sheet.insulation_resistance_ln_to_ld_a
            # if test_sheet.insulation_resistance_ln_to_ld_a_units:
            #     ws['G9'].value = test_sheet.insulation_resistance_ln_to_ld_a_units
            # if test_sheet.insulation_resistance_ln_to_ld_b:
            #     bb=test_sheet.insulation_resistance_ln_to_ld_b
            # if test_sheet.insulation_resistance_ln_to_ld_b_units:
            #     ws['I9'].value = test_sheet.insulation_resistance_ln_to_ld_b_units
            # if test_sheet.insulation_resistance_ln_to_ld_c:
            #     cc=test_sheet.insulation_resistance_ln_to_ld_c
            # if test_sheet.insulation_resistance_ln_to_ld_c_units:
            #     ws['K9'].value = test_sheet.insulation_resistance_ln_to_ld_c_units
            # if test_sheet.contact_resistance_current:
            #     contactamps=test_sheet.contact_resistance_current
            # if test_sheet.contact_resistance_a:
            #     a=test_sheet.contact_resistance_a
            # if test_sheet.contact_resistance_a_units:
            #     ws['G13'].value = test_sheet.contact_resistance_a_units
            # if test_sheet.contact_resistance_b:
            #     b=test_sheet.contact_resistance_b
            # if test_sheet.contact_resistance_b_units:
            #     ws['I13'].value = test_sheet.contact_resistance_b_units
            # if test_sheet.contact_resistance_c:
            #     c=test_sheet.contact_resistance_c
            # if test_sheet.contact_resistance_c_units:
            #     ws['K13'].value = test_sheet.contact_resistance_c_units
            # if test_sheet.hipot_test_voltage:
            #     hipotv=test_sheet.hipot_test_voltage
            # if test_sheet.hipot_ptp_ab:
            #     hab=test_sheet.hipot_ptp_ab
            # if test_sheet.hipot_ptp_bc:
            #     hbc=test_sheet.hipot_ptp_bc
            # if test_sheet.hipot_ptp_ca:
            #     hca=test_sheet.hipot_ptp_ca
            # if test_sheet.hipot_ptg_a:
            #     hag=test_sheet.hipot_ptg_a
            # if test_sheet.hipot_ptg_b:
            #     hbg=test_sheet.hipot_ptg_b
            # if test_sheet.hipot_ptg_c:
            #     hcg=test_sheet.hipot_ptg_c
            # if test_sheet.hipot_ltl_a:
            #     haa=test_sheet.hipot_ltl_a
            # if test_sheet.hipot_ltl_b:
            #     hbb=test_sheet.hipot_ltl_b
            # if test_sheet.hipot_ltl_c:
            #     hcc=test_sheet.hipot_ltl_c

        # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # response['Content-Disposition'] = 'attachment; filename="MV Breaker TS.xlsx"'

        # ws['A1'].value = 'Job: '+job.job_name
        # ws['B4'].value = siteid
        # ws['B5'].value = serial
        # ws['B6'].value = eqlocation
        # ws['B7'].value = mfr
        # ws['B8'].value = eqtype
        # ws['B9'].value = eqmodel
        # ws['B10'].value = framerating
        # ws['B11'].value = mountstyle
        # ws['B12'].value = controlv
        # ws['B13'].value = tripv
        # ws['B14'].value = tester
        # ws['B15'].value = testdate
        # ws['E5'].value = phphv
        # ws['E7'].value = phgv
        # ws['E9'].value = llv
        # ws['F5'].value = ab
        # ws['H5'].value = bc
        # ws['J5'].value = ca
        # ws['F7'].value = ag
        # ws['H7'].value = bg
        # ws['J7'].value = cg
        # ws['F9'].value = aa
        # ws['H9'].value = bb
        # ws['J9'].value = cc
        # ws['E13'].value = contactamps
        # ws['F13'].value = a
        # ws['H13'].value = b
        # ws['J13'].value = c
        # ws['E17'].value = hipotv
        # ws['F17'].value = hab
        # ws['H17'].value = hbc
        # ws['J17'].value = hca
        # ws['F19'].value = hag
        # ws['H19'].value = hbg
        # ws['J19'].value = hcg
        # ws['F21'].value = haa
        # ws['H21'].value = hbb
        # ws['J21'].value = hcc

        # wb.save(response)


def edit_test_results(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    eq = Equipment.objects.get(pk=equipment_id)
    is_casting = False
    mold = None
    if eq.equipment_mold:
        is_casting = True    #this eq belongs to a job rather than a site so it is a casting
        mold = eq.equipment_mold.sheet_eq
    try:
        test_sheet = TestSheet.objects.get(eq = eq)       
    except:
        return render(request, "jobs/error.html", {"message": "Test sheet not found for this equipment. Please contact Admin."})

    if "date_tested" in request.POST:
        if request.POST["date_tested"]!="":
            try:
                test_sheet.date_tested = request.POST["date_tested"]    
            except:
                return render(request, "jobs/error.html", {"message": "other error"})
        else:
            test_sheet.date_tested = None
                
    if "date_manufactured" in request.POST:
        if request.POST["date_manufactured"]!="":
            try:
                test_sheet.date_manufactured = request.POST["date_manufactured"]    
            except:
                return render(request, "jobs/error.html", {"message": "date formatting error"})
        else:
            test_sheet.date_manufactured = None
                
    if "testers" in request.POST:
        test_sheet.testers = request.POST["testers"]        
    insul_toggle= request.POST.get("insulation_toggle", False)
    if insul_toggle:
        eq.is_insulation_resistance = True
    else:
        eq.is_insulation_resistance = False
    if "fuse_toggle" in request.POST:
        if is_casting:
            eq.equipment_mold.is_power_fused = True
        eq.is_power_fused = True
    else:
        if is_casting:
            eq.equipment_mold.is_power_fused = False
        eq.is_power_fused = False
    if "contact_toggle" in request.POST:
        eq.is_contact_resistance = True
    else:
        eq.is_contact_resistance = False
    if "trip_unit_toggle" in request.POST:
        if is_casting:
            eq.equipment_mold.is_trip_unit = True
        eq.is_trip_unit = True
    else:
        if is_casting:
            eq.equipment_mold.is_trip_unit = False
        eq.is_trip_unit = False
    if "pi_toggle" in request.POST:
        eq.is_primary_injection = True
    else:
        eq.is_primary_injection = False
    if "si_toggle" in request.POST:
        eq.is_secondary_injection = True
    else:
        eq.is_secondary_injection = False

    if "hipot_toggle" in request.POST:
        eq.is_hipot = True
    else:
        eq.is_hipot = False
    if "wr_toggle" in request.POST:
        eq.is_winding_resistance = True
    else:
        eq.is_winding_resistance = False
    if "ttr_toggle" in request.POST:
        eq.is_ttr = True
    else:
        eq.is_ttr = False
    if "xfmr_ir_toggle" in request.POST:
        eq.is_xfmr_insulation_resistance = True
    else:
        eq.is_xfmr_insulation_resistance = False
    #     is_insulation_resistance=mo
    # is_contact_resistance=model
    # is_trip_unit=models.Boolean
    # is_primary_injection=models
    # is_secondary_injection=mode
    # is_power_fused=models.Boole
    # is_breaker=models.BooleanFi
    # is_hipot=models.BooleanFiel
    # is_inspection=models.Boolea
    # is_transformer=models.Boole
    # is_winding_resistance=model
    # is_liquid_type=models.Boole
    # is_cable=models.BooleanFiel
    # is_cable_vlf_withstand_test
    
    if "cable_hipot_quantity" in request.POST:
        if request.POST["cable_hipot_quantity"]!="":
            test_sheet.cable_hipot_quantity = request.POST["cable_hipot_quantity"]
            if int(request.POST["cable_hipot_quantity"])>test_sheet.cable_test_data.count():
                for _ in range(int(request.POST["cable_hipot_quantity"])-test_sheet.cable_test_data.count()):
                    new_ctd = CableTestData(test_sheet = test_sheet)
                    new_ctd.save()
        else:
            test_sheet.cable_hipot_quantity = None  
    
    if "bus_contact_resistance_quantity" in request.POST:
        if request.POST["bus_contact_resistance_quantity"]!="":
            test_sheet.bus_contact_resistance_quantity = request.POST["bus_contact_resistance_quantity"]
            if int(request.POST["bus_contact_resistance_quantity"])>test_sheet.bus_test_data.count():
                for _ in range(int(request.POST["bus_contact_resistance_quantity"])-test_sheet.bus_test_data.count()):
                    new_ctd = BusContactTestData(test_sheet = test_sheet)
                    new_ctd.save()
        else:
            test_sheet.bus_contact_resistance_quantity = None  
    
    for test in test_sheet.bus_test_data.all():
        if "starting_section_"+str(test.pk) in request.POST:
            if request.POST["starting_section_"+str(test.pk)]!="":
                test.starting_section = request.POST["starting_section_"+str(test.pk)]
            else:
                test.starting_section = None  
        if "ending_section_"+str(test.pk) in request.POST:
            if request.POST["ending_section_"+str(test.pk)]!="":
                test.ending_section = request.POST["ending_section_"+str(test.pk)]
            else:
                test.ending_section = None  
        if "phase_a_"+str(test.pk) in request.POST:
            if request.POST["phase_a_"+str(test.pk)]!="":
                test.phase_a = request.POST["phase_a_"+str(test.pk)]
            else:
                test.phase_a = None  
        if "phase_b_"+str(test.pk) in request.POST:
            if request.POST["phase_b_"+str(test.pk)]!="":
                test.phase_b = request.POST["phase_b_"+str(test.pk)]
            else:
                test.phase_b = None  
        if "phase_c_"+str(test.pk) in request.POST:
            if request.POST["phase_c_"+str(test.pk)]!="":
                test.phase_c = request.POST["phase_c_"+str(test.pk)]
            else:
                test.phase_c = None  
        if "phase_a_units_"+str(test.pk) in request.POST:
            if request.POST["phase_a_units_"+str(test.pk)]!="":
                test.phase_a_units = request.POST["phase_a_units_"+str(test.pk)]
            else:
                test.phase_a_units = None  
        if "phase_b_units_"+str(test.pk) in request.POST:
            if request.POST["phase_b_units_"+str(test.pk)]!="":
                test.phase_b_units = request.POST["phase_b_units_"+str(test.pk)]
            else:
                test.phase_b_units = None  
        if "phase_c_units_"+str(test.pk) in request.POST:
            if request.POST["phase_c_units_"+str(test.pk)]!="":
                test.phase_c_units = request.POST["phase_c_units_"+str(test.pk)]
            else:
                test.phase_c_units = None  
        test.save()

        
    for test in test_sheet.cable_test_data.all():
        
        if "cable_hipot_time_"+str(test.pk) in request.POST:
            if request.POST["cable_hipot_time_"+str(test.pk)]!="":
                test.time = request.POST["cable_hipot_time_"+str(test.pk)]
            else:
                test.time = None  

        if "cable_hipot_volts_"+str(test.pk) in request.POST:
            if request.POST["cable_hipot_volts_"+str(test.pk)]!="":
                test.test_voltage = request.POST["cable_hipot_volts_"+str(test.pk)]
            else:
                test.test_voltage = None  

        if "cable_hipot_a_"+str(test.pk) in request.POST:
            if request.POST["cable_hipot_a_"+str(test.pk)]!="":
                test.phase_a = request.POST["cable_hipot_a_"+str(test.pk)]
            else:
                test.phase_a = None  

        if "cable_hipot_b_"+str(test.pk) in request.POST:
            if request.POST["cable_hipot_b_"+str(test.pk)]!="":
                test.phase_b = request.POST["cable_hipot_b_"+str(test.pk)]
            else:
                test.phase_b = None  

        if "cable_hipot_c_"+str(test.pk) in request.POST:
            if request.POST["cable_hipot_c_"+str(test.pk)]!="":
                test.phase_c = request.POST["cable_hipot_c_"+str(test.pk)]
            else:
                test.phase_c = None  
        test.save()

    
    if "cable_starting_point" in request.POST:
        if request.POST["cable_starting_point"]!="":
            test_sheet.cable_starting_point = request.POST["cable_starting_point"]
            if is_casting:
                mold.cable_starting_point = request.POST["cable_starting_point"]
        else:
            test_sheet.cable_starting_point = None
            if is_casting:
                mold.cable_starting_point = None
                      
    if "cable_ending_point" in request.POST:
        if request.POST["cable_ending_point"]!="":
            test_sheet.cable_ending_point = request.POST["cable_ending_point"]
            if is_casting:
                mold.cable_ending_point = request.POST["cable_ending_point"]
        else:
            test_sheet.cable_ending_point = None
            if is_casting:
                mold.cable_ending_point = None
                      
    if "cable_starting_termination_type" in request.POST:
        if request.POST["cable_starting_termination_type"]!="":
            test_sheet.cable_starting_termination_type = request.POST["cable_starting_termination_type"]
            if is_casting:
                mold.cable_starting_termination_type = request.POST["cable_starting_termination_type"]
        else:
            test_sheet.cable_starting_termination_type = None
            if is_casting:
                mold.cable_starting_termination_type = None
                      
    if "cable_ending_termination_type" in request.POST:
        if request.POST["cable_ending_termination_type"]!="":
            test_sheet.cable_ending_termination_type = request.POST["cable_ending_termination_type"]
            if is_casting:
                mold.cable_ending_termination_type = request.POST["cable_ending_termination_type"]
        else:
            test_sheet.cable_ending_termination_type = None
            if is_casting:
                mold.cable_ending_termination_type = None
                      
    if "cable_conductor_material" in request.POST:
        if request.POST["cable_conductor_material"]!="":
            test_sheet.cable_conductor_material = request.POST["cable_conductor_material"]
            if is_casting:
                mold.cable_conductor_material = request.POST["cable_conductor_material"]
        else:
            test_sheet.cable_conductor_material = None
            if is_casting:
                mold.cable_conductor_material = None
                      
    if "cable_size" in request.POST:
        if request.POST["cable_size"]!="":
            test_sheet.cable_size = request.POST["cable_size"]
            if is_casting:
                mold.cable_size = request.POST["cable_size"]
        else:
            test_sheet.cable_size = None
            if is_casting:
                mold.cable_size = None

    if "cable_size_units" in request.POST:
        if request.POST["cable_size_units"]!="":
            print(request.POST["cable_size_units"])
            test_sheet.cable_size_units = request.POST["cable_size_units"]
            if is_casting:
                mold.cable_size_units = request.POST["cable_size_units"]
        else:
            test_sheet.cable_size_units = None
            if is_casting:
                mold.cable_size_units = None
                      
    if "operating_cable_voltage" in request.POST:
        if request.POST["operating_cable_voltage"]!="":
            test_sheet.operating_cable_voltage = request.POST["operating_cable_voltage"]
            if is_casting:
                mold.operating_cable_voltage = request.POST["operating_cable_voltage"]
        else:
            test_sheet.operating_cable_voltage = None
            if is_casting:
                mold.operating_cable_voltage = None

    if "cable_voltage_rating" in request.POST:
        if request.POST["cable_voltage_rating"]!="":
            test_sheet.cable_voltage_rating = request.POST["cable_voltage_rating"]
            if is_casting:
                mold.cable_voltage_rating = request.POST["cable_voltage_rating"]
        else:
            test_sheet.cable_voltage_rating = None
            if is_casting:
                mold.cable_voltage_rating = None

    if "cable_insulation_rating" in request.POST:
        if request.POST["cable_insulation_rating"]!="":
            test_sheet.cable_insulation_rating = int(request.POST["cable_insulation_rating"])
            if is_casting:
                mold.cable_insulation_rating = int(request.POST["cable_insulation_rating"])
        else:
            test_sheet.cable_insulation_rating = None
            if is_casting:
                mold.cable_insulation_rating = None

    if "cable_insulation_material" in request.POST:
        if request.POST["cable_insulation_material"]!="":
            test_sheet.cable_insulation_material = request.POST["cable_insulation_material"]
            if is_casting:
                mold.cable_insulation_material = request.POST["cable_insulation_material"]
        else:
            test_sheet.cable_insulation_material = None
            if is_casting:
                mold.cable_insulation_material = None
                      
    if "cable_insulation_thickness" in request.POST:
        if request.POST["cable_insulation_thickness"]!="":
            test_sheet.cable_insulation_thickness = request.POST["cable_insulation_thickness"]
            if is_casting:
                mold.cable_insulation_thickness = request.POST["cable_insulation_thickness"]
        else:
            test_sheet.cable_insulation_thickness = None
            if is_casting:
                mold.cable_insulation_thickness = None
                         
    if "cable_insulation_thickness_units" in request.POST:
        if request.POST["cable_insulation_thickness_units"]!="":
            test_sheet.cable_insulation_thickness_units = request.POST["cable_insulation_thickness_units"]
            if is_casting:
                mold.cable_insulation_thickness_units = request.POST["cable_insulation_thickness_units"]
        else:
            test_sheet.cable_insulation_thickness_units = None
            if is_casting:
                mold.cable_insulation_thickness_units = None
                      
    if "cable_length" in request.POST:
        if request.POST["cable_length"]!="":
            test_sheet.cable_length = request.POST["cable_length"]
            if is_casting:
                mold.cable_length = request.POST["cable_length"]
        else:
            test_sheet.cable_length = None
            if is_casting:
                mold.cable_length = None
    
    if "cable_length_units" in request.POST:
        if request.POST["cable_length_units"]!="":
            test_sheet.cable_length_units = request.POST["cable_length_units"]
            if is_casting:
                mold.cable_length_units = request.POST["cable_length_units"]
        else:
            test_sheet.cable_length_units = None
            if is_casting:
                mold.cable_length_units = None

    
    
    if "is_cable_data_specs_match" in request.POST:
        test_sheet.is_cable_data_specs_match = True
    else:
        test_sheet.is_cable_data_specs_match = False

    if "is_no_physical_damage" in request.POST:
        test_sheet.is_no_physical_damage = True
    else:
        test_sheet.is_no_physical_damage = False

    if "is_connection_verification" in request.POST:
        test_sheet.is_connection_verification = True
    else:
        test_sheet.is_connection_verification = False

    if "is_compression_match" in request.POST:
        test_sheet.is_compression_match = True
    else:
        test_sheet.is_compression_match = False

    if "is_shield_supports_terminations" in request.POST:
        test_sheet.is_shield_supports_terminations = True
    else:
        test_sheet.is_shield_supports_terminations = False

    if "is_bend_radius" in request.POST:
        test_sheet.is_bend_radius = True
    else:
        test_sheet.is_bend_radius = False

    if "is_fireproofing" in request.POST:
        test_sheet.is_fireproofing = True
    else:
        test_sheet.is_fireproofing = False

    if "is_window_ct_correct" in request.POST:
        test_sheet.is_window_ct_correct = True
    else:
        test_sheet.is_window_ct_correct = False

    if "is_id_arrangments_correct" in request.POST:
        test_sheet.is_id_arrangments_correct = True
    else:
        test_sheet.is_id_arrangments_correct = False

    if "is_cable_jacket_insulation_ok" in request.POST:
        test_sheet.is_cable_jacket_insulation_ok = True
    else:
        test_sheet.is_cable_jacket_insulation_ok = False            
    if "operating_cable_voltage" in request.POST:
        if request.POST["operating_cable_voltage"]!="":
            test_sheet.operating_cable_voltage = request.POST["operating_cable_voltage"]
            if is_casting:
                mold.operating_cable_voltage = request.POST["operating_cable_voltage"]
        else:
            test_sheet.operating_cable_voltage = None
            if is_casting:
                mold.operating_cable_voltage = None
    if "system_voltage" in request.POST:
        if request.POST["system_voltage"]!="":
            test_sheet.system_voltage = request.POST["system_voltage"]
            if is_casting:
                mold.system_voltage = request.POST["system_voltage"]
        else:
            test_sheet.system_voltage = None
            if is_casting:
                mold.system_voltage = None            
    if "is_dc_system_voltage" in request.POST:
        if request.POST["is_dc_system_voltage"]=="DC":
            test_sheet.is_dc_system_voltage = True
            if is_casting:
                mold.is_dc_system_voltage = True
        else:
            test_sheet.is_dc_system_voltage = False
            if is_casting:
                mold.is_dc_system_voltage = False      
    if "interrupting_capacity" in request.POST:
        if request.POST["interrupting_capacity"]!="":
            test_sheet.interrupting_capacity = request.POST["interrupting_capacity"]
            if is_casting:
                mold.interrupting_capacity = request.POST["interrupting_capacity"]
        else:
            test_sheet.interrupting_capacity = None
            if is_casting:
                mold.interrupting_capacity = None           
    if "interrupting_voltage" in request.POST:
        if request.POST["interrupting_voltage"]!="":
            test_sheet.interrupting_voltage = request.POST["interrupting_voltage"]
            if is_casting:
                mold.interrupting_voltage = request.POST["interrupting_voltage"]
        else:
            test_sheet.interrupting_voltage = None
            if is_casting:
                mold.interrupting_voltage = None           
    if "equipment_voltage" in request.POST:
        if request.POST["equipment_voltage"]!="":
            test_sheet.equipment_voltage = request.POST["equipment_voltage"]
            if is_casting:
                mold.equipment_voltage = request.POST["equipment_voltage"]
        else:
            test_sheet.equipment_voltage = None
            if is_casting:
                mold.equipment_voltage = None             
    if "serial_number2" in request.POST:
        if request.POST["serial_number2"]!="":
            eq.serial_number = request.POST["serial_number2"]
            if is_casting:
                mold.eq.serial_number = request.POST["serial_number2"]  
        eq.save()
    if "equipment_location2" in request.POST:
        if request.POST["equipment_location2"]!="":
            eq.serial_number = request.POST["equipment_location2"]
            if is_casting:
                mold.eq.equipment_location = request.POST["equipment_location2"]  
        eq.save()
    if "is_dc_equipment_voltage" in request.POST:
        if request.POST["is_dc_equipment_voltage"]=="DC":
            test_sheet.is_dc_equipment_voltage = True
            if is_casting:
                mold.is_dc_equipment_voltage = True
        else:
            test_sheet.is_dc_equipment_voltage = False
            if is_casting:
                mold.is_dc_equipment_voltage = False               
    if "is_dc_control_voltage" in request.POST:
        if request.POST["is_dc_control_voltage"]=="DC":
            test_sheet.is_dc_control_voltage = True
            if is_casting:
                mold.is_dc_control_voltage = True
        else:
            test_sheet.is_dc_control_voltage = False
            if is_casting:
                mold.is_dc_control_voltage = False
                    
    if "control_voltage" in request.POST:
        if request.POST["control_voltage"]!="":
            test_sheet.control_voltage = request.POST["control_voltage"]
            if is_casting:
                mold.control_voltage = request.POST["control_voltage"]
        else:
            test_sheet.control_voltage = None
            if is_casting:
                mold.control_voltage = None
    
    if "trip_coil_voltage" in request.POST:
        if request.POST["trip_coil_voltage"]!="":
            test_sheet.trip_coil_voltage = request.POST["trip_coil_voltage"]
            if is_casting:
                mold.trip_coil_voltage = request.POST["trip_coil_voltage"]
        else:
            test_sheet.trip_coil_voltage = None
            if is_casting:
                mold.trip_coil_voltage = None
    if "is_dc_trip_voltage" in request.POST:
        if request.POST["is_dc_trip_voltage"]=="DC":
            test_sheet.is_dc_trip_voltage = True
            if is_casting:
                mold.is_dc_trip_voltage = True
        else:
            test_sheet.is_dc_trip_voltage = False
            if is_casting:
                mold.is_dc_trip_voltage = False
    if "insulation_resistance_ph_to_ph_test_voltage" in request.POST:
        if request.POST["insulation_resistance_ph_to_ph_test_voltage"]!="":
           test_sheet.insulation_resistance_ph_to_ph_test_voltage = request.POST["insulation_resistance_ph_to_ph_test_voltage"]
        else:
            test_sheet.insulation_resistance_ph_to_ph_test_voltage = None
    if "insulation_resistance_ph_to_ph_a_b" in request.POST:
        if request.POST["insulation_resistance_ph_to_ph_a_b"]!="":
            test_sheet.insulation_resistance_ph_to_ph_a_b = request.POST["insulation_resistance_ph_to_ph_a_b"]
        else:
            test_sheet.insulation_resistance_ph_to_ph_a_b = None     
    if "insulation_resistance_ph_to_ph_b_c" in request.POST:
        if request.POST["insulation_resistance_ph_to_ph_b_c"]!="":
           test_sheet.insulation_resistance_ph_to_ph_b_c = request.POST["insulation_resistance_ph_to_ph_b_c"]
        else:
            test_sheet.insulation_resistance_ph_to_ph_b_c = None 
    if "insulation_resistance_ph_to_ph_c_a" in request.POST:
        if request.POST["insulation_resistance_ph_to_ph_c_a"]!="":
           test_sheet.insulation_resistance_ph_to_ph_c_a = request.POST["insulation_resistance_ph_to_ph_c_a"]
        else:
            test_sheet.insulation_resistance_ph_to_ph_c_a = None 
    if "insulation_resistance_ln_to_ld_test_voltage" in request.POST:
        if request.POST["insulation_resistance_ln_to_ld_test_voltage"]!="":
           test_sheet.insulation_resistance_ln_to_ld_test_voltage = request.POST["insulation_resistance_ln_to_ld_test_voltage"]
        else:
            test_sheet.insulation_resistance_ln_to_ld_test_voltage = None 
    if "insulation_resistance_ln_to_ld_a" in request.POST:
        if request.POST["insulation_resistance_ln_to_ld_a"]!="":
           test_sheet.insulation_resistance_ln_to_ld_a = request.POST["insulation_resistance_ln_to_ld_a"]
        else:
            test_sheet.insulation_resistance_ln_to_ld_a = None 
    if "insulation_resistance_ln_to_ld_b" in request.POST:
        if request.POST["insulation_resistance_ln_to_ld_b"]!="":
           test_sheet.insulation_resistance_ln_to_ld_b = request.POST["insulation_resistance_ln_to_ld_b"]
        else:
            test_sheet.insulation_resistance_ln_to_ld_b = None 
    if "insulation_resistance_ln_to_ld_c" in request.POST:
        if request.POST["insulation_resistance_ln_to_ld_c"]!="":
           test_sheet.insulation_resistance_ln_to_ld_c = request.POST["insulation_resistance_ln_to_ld_c"]
        else:
            test_sheet.insulation_resistance_ln_to_ld_c = None 
    if "insulation_resistance_ph_to_gr_test_voltage" in request.POST:
        if request.POST["insulation_resistance_ph_to_gr_test_voltage"]!="":
           test_sheet.insulation_resistance_ph_to_gr_test_voltage = request.POST["insulation_resistance_ph_to_gr_test_voltage"]
        else:
            test_sheet.insulation_resistance_ph_to_gr_test_voltage = None 
    if "insulation_resistance_ph_to_gr_a_g" in request.POST:
        if request.POST["insulation_resistance_ph_to_gr_a_g"]!="":
           test_sheet.insulation_resistance_ph_to_gr_a_g = request.POST["insulation_resistance_ph_to_gr_a_g"]
        else:
            test_sheet.insulation_resistance_ph_to_gr_a_g = None 
    if "insulation_resistance_ph_to_gr_b_g" in request.POST:
        if request.POST["insulation_resistance_ph_to_gr_b_g"]!="":
           test_sheet.insulation_resistance_ph_to_gr_b_g = request.POST["insulation_resistance_ph_to_gr_b_g"]
        else:
            test_sheet.insulation_resistance_ph_to_gr_b_g = None 
    if "insulation_resistance_ph_to_gr_c_g" in request.POST:
        if request.POST["insulation_resistance_ph_to_gr_c_g"]!="":
           test_sheet.insulation_resistance_ph_to_gr_c_g = request.POST["insulation_resistance_ph_to_gr_c_g"]
        else:
            test_sheet.insulation_resistance_ph_to_gr_c_g = None 
    if "insulation_resistance_ph_to_ph_a_b_units" in request.POST:
        if request.POST["insulation_resistance_ph_to_ph_a_b_units"]!="":
            test_sheet.insulation_resistance_ph_to_ph_a_b_units = request.POST["insulation_resistance_ph_to_ph_a_b_units"]
        else:
            test_sheet.insulation_resistance_ph_to_ph_a_b_units = None 
    if "insulation_resistance_ph_to_ph_b_c_units" in request.POST:
        if request.POST["insulation_resistance_ph_to_ph_b_c_units"]!="":
            test_sheet.insulation_resistance_ph_to_ph_b_c_units = request.POST["insulation_resistance_ph_to_ph_b_c_units"]
        else:
            test_sheet.insulation_resistance_ph_to_ph_b_c_units = None 
    if "insulation_resistance_ph_to_ph_c_a_units" in request.POST:
        if request.POST["insulation_resistance_ph_to_ph_c_a_units"]!="":
            test_sheet.insulation_resistance_ph_to_ph_c_a_units = request.POST["insulation_resistance_ph_to_ph_c_a_units"]
        else:
            test_sheet.insulation_resistance_ph_to_ph_c_a_units = None  
    if "insulation_resistance_ln_to_ld_a_units" in request.POST:
        if request.POST["insulation_resistance_ln_to_ld_a_units"]!="":
            test_sheet.insulation_resistance_ln_to_ld_a_units = request.POST["insulation_resistance_ln_to_ld_a_units"]
        else:
            test_sheet.insulation_resistance_ln_to_ld_a_units = None  
    if "insulation_resistance_ln_to_ld_b_units" in request.POST:
        if request.POST["insulation_resistance_ln_to_ld_b_units"]!="":
            test_sheet.insulation_resistance_ln_to_ld_b_units = request.POST["insulation_resistance_ln_to_ld_b_units"]
        else:
            test_sheet.insulation_resistance_ln_to_ld_b_units = None  
    if "insulation_resistance_ln_to_ld_c_units" in request.POST:
        if request.POST["insulation_resistance_ln_to_ld_c_units"]!="":
            test_sheet.insulation_resistance_ln_to_ld_c_units = request.POST["insulation_resistance_ln_to_ld_c_units"]
        else:
            test_sheet.insulation_resistance_ln_to_ld_c_units = None  
    if "insulation_resistance_ph_to_gr_a_g_units" in request.POST:
        if request.POST["insulation_resistance_ph_to_gr_a_g_units"]!="":
            test_sheet.insulation_resistance_ph_to_gr_a_g_units = request.POST["insulation_resistance_ph_to_gr_a_g_units"]
        else:
            test_sheet.insulation_resistance_ph_to_gr_a_g_units = None  
    if "insulation_resistance_ph_to_gr_b_g_units" in request.POST:
        if request.POST["insulation_resistance_ph_to_gr_b_g_units"]!="":
            test_sheet.insulation_resistance_ph_to_gr_b_g_units = request.POST["insulation_resistance_ph_to_gr_b_g_units"]
        else:
            test_sheet.insulation_resistance_ph_to_gr_b_g_units = None  
    if "insulation_resistance_ph_to_gr_c_g_units" in request.POST:
        if request.POST["insulation_resistance_ph_to_gr_c_g_units"]!="":
            test_sheet.insulation_resistance_ph_to_gr_c_g_units = request.POST["insulation_resistance_ph_to_gr_c_g_units"]
        else:
            test_sheet.insulation_resistance_ph_to_gr_c_g_units = None 
    if "contact_resistance_current" in request.POST:
        if request.POST["contact_resistance_current"]!="":
           test_sheet.contact_resistance_current = request.POST["contact_resistance_current"]
        else:
            test_sheet.contact_resistance_current = None 
    if "contact_resistance_a" in request.POST:
        if request.POST["contact_resistance_a"]!="":
           test_sheet.contact_resistance_a = request.POST["contact_resistance_a"]
        else:
            test_sheet.contact_resistance_a = None 
    if "contact_resistance_b" in request.POST:
        if request.POST["contact_resistance_b"]!="":
           test_sheet.contact_resistance_b = request.POST["contact_resistance_b"]
        else:
            test_sheet.contact_resistance_b = None 
    if "contact_resistance_c" in request.POST:
        if request.POST["contact_resistance_c"]!="":
           test_sheet.contact_resistance_c = request.POST["contact_resistance_c"]
        else:
            test_sheet.contact_resistance_c = None 
    if "contact_resistance_a_units" in request.POST:
        if request.POST["contact_resistance_a_units"]!="":
            test_sheet.contact_resistance_a_units = request.POST["contact_resistance_a_units"]
        else:
            test_sheet.contact_resistance_a_units = None   
    if "contact_resistance_b_units" in request.POST:
        if request.POST["contact_resistance_b_units"]!="":
            test_sheet.contact_resistance_b_units = request.POST["contact_resistance_b_units"]
        else:
            test_sheet.contact_resistance_b_units = None 
    if "contact_resistance_c_units" in request.POST:
        if request.POST["contact_resistance_c_units"]!="":
            test_sheet.contact_resistance_c_units = request.POST["contact_resistance_c_units"]
        else:
            test_sheet.contact_resistance_c_units = None
    if "control_wiring_insulation_resistance" in request.POST:
        if request.POST["control_wiring_insulation_resistance"]!="":
           test_sheet.control_wiring_insulation_resistance = request.POST["control_wiring_insulation_resistance"]
        else:
            test_sheet.control_wiring_insulation_resistance = None 
    if "insulation_resistance_hi_g" in request.POST:
        if request.POST["insulation_resistance_hi_g"]!="":
            test_sheet.insulation_resistance_hi_g = request.POST["insulation_resistance_hi_g"]
        else:
            test_sheet.insulation_resistance_hi_g = None  
    if "insulation_resistance_lo_g" in request.POST:
        if request.POST["insulation_resistance_lo_g"]!="":
            test_sheet.insulation_resistance_lo_g = request.POST["insulation_resistance_lo_g"]
        else:
            test_sheet.insulation_resistance_lo_g = None  
    if "insulation_resistance_hi_lo" in request.POST:
        if request.POST["insulation_resistance_hi_lo"]!="":
            test_sheet.insulation_resistance_hi_lo = request.POST["insulation_resistance_hi_lo"]
        else:
            test_sheet.insulation_resistance_hi_lo = None  
    if "insulation_resistance_hi_g_test_voltage" in request.POST:
        if request.POST["insulation_resistance_hi_g_test_voltage"]!="":
            test_sheet.insulation_resistance_hi_g_test_voltage = request.POST["insulation_resistance_hi_g_test_voltage"]
        else:
            test_sheet.insulation_resistance_hi_g_test_voltage = None  
    if "insulation_resistance_lo_g_test_voltage" in request.POST:
        if request.POST["insulation_resistance_lo_g_test_voltage"]!="":
            test_sheet.insulation_resistance_lo_g_test_voltage = request.POST["insulation_resistance_lo_g_test_voltage"]
        else:
            test_sheet.insulation_resistance_lo_g_test_voltage = None  
    if "insulation_resistance_hi_lo_test_voltage" in request.POST:
        if request.POST["insulation_resistance_hi_lo_test_voltage"]!="":
            test_sheet.insulation_resistance_hi_lo_test_voltage = request.POST["insulation_resistance_hi_lo_test_voltage"]
        else:
            test_sheet.insulation_resistance_hi_lo_test_voltage = None    
    if "insulation_resistance_hi_lo_units" in request.POST:
        if request.POST["insulation_resistance_hi_lo_units"]!="":
            test_sheet.insulation_resistance_hi_lo_units = request.POST["insulation_resistance_hi_lo_units"]
        else:
            test_sheet.insulation_resistance_hi_lo_units = None
    if "insulation_resistance_lo_g_units" in request.POST:
        if request.POST["insulation_resistance_lo_g_units"]!="":
            test_sheet.insulation_resistance_lo_g_units = request.POST["insulation_resistance_lo_g_units"]
        else:
            test_sheet.insulation_resistance_lo_g_units = None
    if "insulation_resistance_hi_g_units" in request.POST:
        if request.POST["insulation_resistance_hi_g_units"]!="":
            test_sheet.insulation_resistance_hi_g_units = request.POST["insulation_resistance_hi_g_units"]
        else:
            test_sheet.insulation_resistance_hi_g_units = None
    if "fuse_manufacturer" in request.POST:
        if request.POST["fuse_manufacturer"]!="":
           test_sheet.fuse_manufacturer = request.POST["fuse_manufacturer"]
        else:
            test_sheet.fuse_manufacturer = None 
    if "fuse_type" in request.POST:
        if request.POST["fuse_type"]!="":
           test_sheet.fuse_type = request.POST["fuse_type"]
        else:
            test_sheet.fuse_type = None 
    if "fuse_size" in request.POST:
        if request.POST["fuse_size"]!="":
           test_sheet.fuse_size = request.POST["fuse_size"]
        else:
            test_sheet.fuse_size = None 
    if "fuse_resistance_a" in request.POST:
        if request.POST["fuse_resistance_a"]!="":
            test_sheet.fuse_resistance_a = request.POST["fuse_resistance_a"]
        else:
            test_sheet.fuse_resistance_a = None  
    if "fuse_resistance_b" in request.POST:
        if request.POST["fuse_resistance_b"]!="":
            test_sheet.fuse_resistance_b = request.POST["fuse_resistance_b"]
        else:
            test_sheet.fuse_resistance_b = None  
    if "fuse_resistance_c" in request.POST:
        if request.POST["fuse_resistance_c"]!="":
            test_sheet.fuse_resistance_c = request.POST["fuse_resistance_c"]
        else:
            test_sheet.fuse_resistance_c = None  
    if "fuse_resistance_a_units" in request.POST:
        if request.POST["fuse_resistance_a_units"]!="":
            test_sheet.fuse_resistance_a_units = request.POST["fuse_resistance_a_units"]
        else:
            test_sheet.fuse_resistance_a_units = None  
    if "fuse_resistance_b_units" in request.POST:
        if request.POST["fuse_resistance_b_units"]!="":
            test_sheet.fuse_resistance_b_units = request.POST["fuse_resistance_b_units"]
        else:
            test_sheet.fuse_resistance_b_units = None  
    if "fuse_resistance_c_units" in request.POST:
        if request.POST["fuse_resistance_c_units"]!="":
            test_sheet.fuse_resistance_c_units = request.POST["fuse_resistance_c_units"]
        else:
            test_sheet.fuse_resistance_c_units = None 
    if "fuse_resistance_secondary_a" in request.POST:
        if request.POST["fuse_resistance_secondary_a"]!="":
            test_sheet.fuse_resistance_secondary_a = request.POST["fuse_resistance_secondary_a"]
        else:
            test_sheet.fuse_resistance_secondary_a = None  
    if "fuse_resistance_secondary_b" in request.POST:
        if request.POST["fuse_resistance_secondary_b"]!="":
            test_sheet.fuse_resistance_secondary_b = request.POST["fuse_resistance_secondary_b"]
        else:
            test_sheet.fuse_resistance_secondary_b = None  
    if "fuse_resistance_secondary_c" in request.POST:
        if request.POST["fuse_resistance_secondary_c"]!="":
            test_sheet.fuse_resistance_secondary_c = request.POST["fuse_resistance_secondary_c"]
        else:
            test_sheet.fuse_resistance_secondary_c = None  
    if "fuse_resistance_secondary_a_units" in request.POST:
        if request.POST["fuse_resistance_secondary_a_units"]!="":
            test_sheet.fuse_resistance_secondary_a_units = request.POST["fuse_resistance_secondary_a_units"]
        else:
            test_sheet.fuse_resistance_secondary_a_units = None  
    if "fuse_resistance_secondary_b_units" in request.POST:
        if request.POST["fuse_resistance_secondary_b_units"]!="":
            test_sheet.fuse_resistance_secondary_b_units = request.POST["fuse_resistance_secondary_b_units"]
        else:
            test_sheet.fuse_resistance_secondary_b_units = None  
    if "fuse_resistance_secondary_c_units" in request.POST:
        if request.POST["fuse_resistance_secondary_c_units"]!="":
            test_sheet.fuse_resistance_secondary_c_units = request.POST["fuse_resistance_secondary_c_units"]
        else:
            test_sheet.fuse_resistance_secondary_c_units = None  
    if "operations_counter_af" in request.POST:
        if request.POST["operations_counter_af"]!="":
            test_sheet.operations_counter_af = request.POST["operations_counter_af"]
        else:
            test_sheet.operations_counter_af = None  
    if "operations_counter_al" in request.POST:
        if request.POST["operations_counter_al"]!="":
            test_sheet.operations_counter_al = request.POST["operations_counter_al"]
        else:
            test_sheet.operations_counter_al = None  
    # if "insulation_resistance_hi_g_units" in request.POST:
    #     if request.POST["insulation_resistance_hi_g_units"]!="":
    #         test_sheet.insulation_resistance_hi_g_units = request.POST["insulation_resistance_hi_g_units"]
    #     else:
    #         test_sheet.insulation_resistance_hi_g_units = None  
    # if "insulation_resistance_lo_g_units" in request.POST:
    #     if request.POST["insulation_resistance_lo_g_units"]!="":
    #         test_sheet.insulation_resistance_lo_g_units = request.POST["insulation_resistance_lo_g_units"]
    #     else:
    #         test_sheet.insulation_resistance_lo_g_units = None  
    # if "insulation_resistance_hi_lo_units" in request.POST:
    #     if request.POST["insulation_resistance_hi_lo_units"]!="":
    #         test_sheet.insulation_resistance_hi_lo_units = request.POST["insulation_resistance_hi_lo_units"]
    #     else:
    #         test_sheet.insulation_resistance_hi_lo_units = None 
            
    if "frame_size" in request.POST:
        if request.POST["frame_size"]!="":
           test_sheet.frame_size = request.POST["frame_size"]
           if is_casting:
               mold.frame_size = request.POST["frame_size"]
        else:
            if is_casting:
                mold.frame_size = None 
            test_sheet.frame_size = None 
    if "mount_style" in request.POST:
        test_sheet.mount_style = request.POST["mount_style"]
        if is_casting:
            mold.mount_style = request.POST["mount_style"]
    if "trip_unit_model" in request.POST:
        if is_casting:
            mold.trip_unit_model = request.POST["trip_unit_model"]
        test_sheet.trip_unit_model = request.POST["trip_unit_model"]
    if "trip_unit_manufacturer" in request.POST:
        test_sheet.trip_unit_manufacturer = request.POST["trip_unit_manufacturer"]
        if is_casting:
            mold.trip_unit_manufacturer = request.POST["trip_unit_manufacturer"]
    if "trip_unit_serial_number" in request.POST:
        if is_casting:
            mold.trip_unit_serial_number = request.POST["trip_unit_serial_number"]
        test_sheet.trip_unit_serial_number = request.POST["trip_unit_serial_number"]
    
    if "trip_unit_curve" in request.POST:        
        if request.POST["trip_unit_curve"]!="":
            test_sheet.trip_unit_curve = request.POST["trip_unit_curve"]
            if is_casting:
                mold.trip_unit_curve = request.POST["trip_unit_curve"]
        else:
            test_sheet.trip_unit_curve = None
            if is_casting:
                mold.trip_unit_curve = None
    if "trip_unit_rating_plug" in request.POST:
        if request.POST["trip_unit_rating_plug"]!="":            
            test_sheet.trip_unit_rating_plug = request.POST["trip_unit_rating_plug"]
        else:
            test_sheet.trip_unit_rating_plug = None 
    if "trip_unit_phase_ct_high" in request.POST:
        if request.POST["trip_unit_phase_ct_high"]!="":
            test_sheet.trip_unit_phase_ct_high = request.POST["trip_unit_phase_ct_high"]
            if is_casting:
                mold.trip_unit_phase_ct_high = request.POST["trip_unit_phase_ct_high"]
        else:
            if is_casting:
                mold.trip_unit_phase_ct_high = None
            test_sheet.trip_unit_phase_ct_high = None 
    if "trip_unit_phase_ct_low" in request.POST:
        if request.POST["trip_unit_phase_ct_low"]!="":
            test_sheet.trip_unit_phase_ct_low = request.POST["trip_unit_phase_ct_low"]
            if is_casting:
                mold.trip_unit_phase_ct_low = request.POST["trip_unit_phase_ct_low"]
        else:
            if is_casting:
                mold.trip_unit_phase_ct_low = None
            test_sheet.trip_unit_phase_ct_low = None 
    if "settings_af_ltpu" in request.POST:
        if request.POST["settings_af_ltpu"]!="":
           test_sheet.settings_af_ltpu = request.POST["settings_af_ltpu"]
        else:
            test_sheet.settings_af_ltpu = None 
    if "settings_af_ltd" in request.POST:
        if request.POST["settings_af_ltd"]!="":
           test_sheet.settings_af_ltd = request.POST["settings_af_ltd"]
        else:
            test_sheet.settings_af_ltd = None 
    if "settings_af_stpu" in request.POST:
        if request.POST["settings_af_stpu"]!="":
            test_sheet.settings_af_stpu = request.POST["settings_af_stpu"]
        else:
            test_sheet.settings_af_stpu = None 
    if "settings_af_std" in request.POST:
        if request.POST["settings_af_std"]!="":
           test_sheet.settings_af_std = request.POST["settings_af_std"]
        else:
            test_sheet.settings_af_std = None 
    if "settings_af_inst" in request.POST:
        if request.POST["settings_af_inst"]!="":
           test_sheet.settings_af_inst = request.POST["settings_af_inst"]
        else:
            test_sheet.settings_af_inst = None 
    if "settings_af_gfpu" in request.POST:
        if request.POST["settings_af_gfpu"]!="":
           test_sheet.settings_af_gfpu = request.POST["settings_af_gfpu"]
        else:
            test_sheet.settings_af_gfpu = None 
    if "settings_af_gfd" in request.POST:
        if request.POST["settings_af_gfd"]!="":
           test_sheet.settings_af_gfd = request.POST["settings_af_gfd"]
        else:
            test_sheet.settings_af_gfd = None 
    if "settings_al_ltpu" in request.POST:
        if request.POST["settings_al_ltpu"]!="":
            test_sheet.settings_al_ltpu = request.POST["settings_al_ltpu"]
        else:
            test_sheet.settings_al_ltpu = None 
    if "settings_al_ltd" in request.POST:
        if request.POST["settings_al_ltd"]!="":
           test_sheet.settings_al_ltd = request.POST["settings_al_ltd"]
        else:
            test_sheet.settings_al_ltd = None 
    if "settings_al_stpu" in request.POST:
        if request.POST["settings_al_stpu"]!="":
           test_sheet.settings_al_stpu = request.POST["settings_al_stpu"]
        else:
            test_sheet.settings_al_stpu = None 
    if "settings_al_std" in request.POST:
        if request.POST["settings_al_std"]!="":
           test_sheet.settings_al_std = request.POST["settings_al_std"]
        else:
            test_sheet.settings_al_std = None 
    if "settings_al_inst" in request.POST:
        if request.POST["settings_al_inst"]!="":
           test_sheet.settings_al_inst = request.POST["settings_al_inst"]
        else:
            test_sheet.settings_al_inst = None 
    if "settings_al_gfpu" in request.POST:
        if request.POST["settings_al_gfpu"]!="":
           test_sheet.settings_al_gfpu = request.POST["settings_al_gfpu"]
        else:
            test_sheet.settings_al_gfpu = None 
    if "settings_al_gfd" in request.POST:
        if request.POST["settings_al_gfd"]!="":
           test_sheet.settings_al_gfd = request.POST["settings_al_gfd"]
        else:
            test_sheet.settings_al_gfd = None 


    #switchgear booleans
    if "switchgear_nameplate_drawings" in request.POST:
        test_sheet.switchgear_nameplate_drawings = True
    else:
        test_sheet.switchgear_nameplate_drawings = False

    if "switchgear_inspect_cords_connectors" in request.POST:
        test_sheet.switchgear_inspect_cords_connectors = True
    else:
        test_sheet.switchgear_inspect_cords_connectors = False

    if "switchgear_anchorage_alignment" in request.POST:
        test_sheet.switchgear_anchorage_alignment = True
    else:
        test_sheet.switchgear_anchorage_alignment = False

    if "switchgear_clean" in request.POST:
        test_sheet.switchgear_clean = True
    else:
        test_sheet.switchgear_clean = False

    if "switchgear_fuse_cb_match_drawings" in request.POST:
        test_sheet.switchgear_fuse_cb_match_drawings = True
    else:
        test_sheet.switchgear_fuse_cb_match_drawings = False

    if "switchgear_ct_vt_ratios_match_drawings" in request.POST:
        test_sheet.switchgear_ct_vt_ratios_match_drawings = True
    else:
        test_sheet.switchgear_ct_vt_ratios_match_drawings = False

    if "switchgear_wiring_tight_secure" in request.POST:
        test_sheet.switchgear_wiring_tight_secure = True
    else:
        test_sheet.switchgear_wiring_tight_secure = False

    if "switchgear_connection_inspection" in request.POST:
        test_sheet.switchgear_connection_inspection = True
    else:
        test_sheet.switchgear_connection_inspection = False

    if "switchgear_op_sequence_correct" in request.POST:
        test_sheet.switchgear_op_sequence_correct = True
    else:
        test_sheet.switchgear_op_sequence_correct = False

    if "switchgear_moving_parts_lubricated" in request.POST:
        test_sheet.switchgear_moving_parts_lubricated = True
    else:
        test_sheet.switchgear_moving_parts_lubricated = False

    if "switchgear_insulators_no_damage" in request.POST:
        test_sheet.switchgear_insulators_no_damage = True
    else:
        test_sheet.switchgear_insulators_no_damage = False

    if "switchgear_barrier_installation_correct" in request.POST:
        test_sheet.switchgear_barrier_installation_correct = True
    else:
        test_sheet.switchgear_barrier_installation_correct = False

    if "switchgear_active_components_exercised" in request.POST:
        test_sheet.switchgear_active_components_exercised = True
    else:
        test_sheet.switchgear_active_components_exercised = False

    if "switchgear_indicating_devices" in request.POST:
        test_sheet.switchgear_indicating_devices = True
    else:
        test_sheet.switchgear_indicating_devices = False

    if "switchgear_filters_vents_clear" in request.POST:
        test_sheet.switchgear_filters_vents_clear = True
    else:
        test_sheet.switchgear_filters_vents_clear = False

    if "switchgear_instrument_transformers_inspected" in request.POST:
        test_sheet.switchgear_instrument_transformers_inspected = True
    else:
        test_sheet.switchgear_instrument_transformers_inspected = False

    if "switchgear_surge_arresters_inspected" in request.POST:
        test_sheet.switchgear_surge_arresters_inspected = True
    else:
        test_sheet.switchgear_surge_arresters_inspected = False

    if "switchgear_cpts_undamaged" in request.POST:
        test_sheet.switchgear_cpts_undamaged = True
    else:
        test_sheet.switchgear_cpts_undamaged = False

    if "switchgear_space_heaters" in request.POST:
        test_sheet.switchgear_space_heaters = True
    else:
        test_sheet.switchgear_space_heaters = False

    if "switchgear_phasing_verified" in request.POST:
        test_sheet.switchgear_phasing_verified = True
    else:
        test_sheet.switchgear_phasing_verified = False

    if "cpt_secondary_wiring_drawings" in request.POST:
        test_sheet.cpt_secondary_wiring_drawings = True
    else:
        test_sheet.cpt_secondary_wiring_drawings = False

    if "v_i_t_secondary_wiring_drawings" in request.POST:
        test_sheet.v_i_t_secondary_wiring_drawings = True
    else:
        test_sheet.v_i_t_secondary_wiring_drawings = False

    if "v_i_t_secondary_voltage_design" in request.POST:
        test_sheet.v_i_t_secondary_voltage_design = True
    else:
        test_sheet.v_i_t_secondary_voltage_design = False

    if "v_i_t_inspect_cords_connectors" in request.POST:
        test_sheet.v_i_t_inspect_cords_connectors = True
    else:
        test_sheet.v_i_t_inspect_cords_connectors = False

    if "v_i_t_nameplate_drawings" in request.POST:
        test_sheet.v_i_t_nameplate_drawings = True
    else:
        test_sheet.v_i_t_nameplate_drawings = False

    if "v_i_t_physical_mechanical_condition" in request.POST:
        test_sheet.v_i_t_physical_mechanical_condition = True
    else:
        test_sheet.v_i_t_physical_mechanical_condition = False

    if "v_i_t_correct_connection" in request.POST:
        test_sheet.v_i_t_correct_connection = True
    else:
        test_sheet.v_i_t_correct_connection = False

    if "v_i_t_clearances_primary_secondary" in request.POST:
        test_sheet.v_i_t_clearances_primary_secondary = True
    else:
        test_sheet.v_i_t_clearances_primary_secondary = False

    if "v_i_t_clean" in request.POST:
        test_sheet.v_i_t_clean = True
    else:
        test_sheet.v_i_t_clean = False

    if "v_i_t_connection_inspection" in request.POST:
        test_sheet.v_i_t_connection_inspection = True
    else:
        test_sheet.v_i_t_connection_inspection = False

    if "v_i_t_grounding_contact" in request.POST:
        test_sheet.v_i_t_grounding_contact = True
    else:
        test_sheet.v_i_t_grounding_contact = False

    if "v_i_t_grounding_contact" in request.POST:
        test_sheet.v_i_t_grounding_contact = True
    else:
        test_sheet.v_i_t_grounding_contact = False

    if "v_i_t_fuze_sizes" in request.POST:
        test_sheet.v_i_t_fuze_sizes = True
    else:
        test_sheet.v_i_t_fuze_sizes = False

    if "v_i_t_lubrication" in request.POST:
        test_sheet.v_i_t_lubrication = True
    else:
        test_sheet.v_i_t_lubrication = False

    if "v_i_t_as_left" in request.POST:
        test_sheet.v_i_t_as_left = True
    else:
        test_sheet.v_i_t_as_left = False

    
    #RELAY BOOLEAN FIELDS
    if "relay_tighten_case" in request.POST:
        test_sheet.relay_tighten_case = True
    else:
        test_sheet.relay_tighten_case = False
    if "relay_inspect_gasket" in request.POST:
        test_sheet.relay_inspect_gasket = True
    else:
        test_sheet.relay_inspect_gasket = False
    if "relay_inspect_shorting" in request.POST:
        test_sheet.relay_inspect_shorting = True
    else:
        test_sheet.relay_inspect_shorting = False
    if "relay_foreign_material" in request.POST:
        test_sheet.relay_foreign_material = True
    else:
        test_sheet.relay_foreign_material = False
    if "relay_verify_reset" in request.POST:
        test_sheet.relay_verify_reset = True
    else:
        test_sheet.relay_verify_reset = False
    if "relay_clean_glass" in request.POST:
        test_sheet.relay_clean_glass = True
    else:
        test_sheet.relay_clean_glass = False
    if "relay_inspect_disk_solts" in request.POST:
        test_sheet.relay_inspect_disk_solts = True
    else:
        test_sheet.relay_inspect_disk_solts = False
    if "relay_disk_clearance" in request.POST:
        test_sheet.relay_disk_clearance = True
    else:
        test_sheet.relay_disk_clearance = False
    if "relay_spring_convolutions" in request.POST:
        test_sheet.relay_spring_convolutions = True
    else:
        test_sheet.relay_spring_convolutions = False
    if "relay_disk_free_travel" in request.POST:
        test_sheet.relay_disk_free_travel = True
    else:
        test_sheet.relay_disk_free_travel = False
    if "relay_tightness_hardware" in request.POST:
        test_sheet.relay_tightness_hardware = True
    else:
        test_sheet.relay_tightness_hardware = False
    if "relay_burnish_contacts" in request.POST:
        test_sheet.relay_burnish_contacts = True
    else:
        test_sheet.relay_burnish_contacts = False
    if "relay_inspect_bearings" in request.POST:
        test_sheet.relay_inspect_bearings = True
    else:
        test_sheet.relay_inspect_bearings = False
    if "relay_verify_settings" in request.POST:
        test_sheet.relay_verify_settings = True
    else:
        test_sheet.relay_verify_settings = False

    #LV SWITCH BOOLEAN FIELDS
    if "switch_verify_mechanical_support" in request.POST:
        test_sheet.switch_verify_mechanical_support = True
    else:
        test_sheet.switch_verify_mechanical_support = False

    #MV AIR SWITCH BOOLEAN FIELDS
    if "switch_verify_expulsion_limiting_devices" in request.POST:
        test_sheet.switch_verify_expulsion_limiting_devices = True
    else:
        test_sheet.switch_verify_expulsion_limiting_devices = False
    if "switch_verify_motor_operator_limit" in request.POST:
        test_sheet.switch_verify_motor_operator_limit = True
    else:
        test_sheet.switch_verify_motor_operator_limit = False
    if "switch_alignment_travel_stops_arc_interrupter" in request.POST:
        test_sheet.switch_alignment_travel_stops_arc_interrupter = True
    else:
        test_sheet.switch_alignment_travel_stops_arc_interrupter = False
    if "fuse_sizes_types_drawings_ect" in request.POST:
        test_sheet.fuse_sizes_types_drawings_ect = True
    else:
        test_sheet.fuse_sizes_types_drawings_ect = False


    #MVVB BOOLEAN FIELDS
    if "mvvb_maintenance_tools_gauges" in request.POST:
        test_sheet.mvvb_maintenance_tools_gauges = True
    else:
        test_sheet.mvvb_maintenance_tools_gauges = False
    if "mvvb_mechanical_operation_tests" in request.POST:
        test_sheet.mvvb_mechanical_operation_tests = True
    else:
        test_sheet.mvvb_mechanical_operation_tests = False
    if "mvvb_critical_distances" in request.POST:
        test_sheet.mvvb_critical_distances = True
    else:
        test_sheet.mvvb_critical_distances = False
    if "mvvb_fit_alignment" in request.POST:
        test_sheet.mvvb_fit_alignment = True
    else:
        test_sheet.mvvb_fit_alignment = False
    if "mvvb_racking_mechanism" in request.POST:
        test_sheet.mvvb_racking_mechanism = True
    else:
        test_sheet.mvvb_racking_mechanism = False
    if "mvvb_current_carrying_parts" in request.POST:
        test_sheet.mvvb_current_carrying_parts = True
    else:
        test_sheet.mvvb_current_carrying_parts = False
    if "mvvb_contact_timing_test" in request.POST:
        test_sheet.mvvb_contact_timing_test = True
    else:
        test_sheet.mvvb_contact_timing_test = False
    if "mvvb_trip_close_coil_analysis" in request.POST:
        test_sheet.mvvb_trip_close_coil_analysis = True
    else:
        test_sheet.mvvb_trip_close_coil_analysis = False
    if "mvvb_mechanism_motion_analysis" in request.POST:
        test_sheet.mvvb_mechanism_motion_analysis = True
    else:
        test_sheet.mvvb_mechanism_motion_analysis = False

    #VT CT AND CPT BOOLEANS
    if "connections_with_system_requirements" in request.POST:
        test_sheet.connections_with_system_requirements = True
    else:
        test_sheet.connections_with_system_requirements = False
    if "grounding_connections_contact" in request.POST:
        test_sheet.grounding_connections_contact = True
    else:
        test_sheet.grounding_connections_contact = False
    if "clearances_primary_secondary_wiring" in request.POST:
        test_sheet.clearances_primary_secondary_wiring = True
    else:
        test_sheet.clearances_primary_secondary_wiring = False

    #SF6 BREAKER BOOLEAN FIELDS
    if "sf6_gas_sample" in request.POST:
        test_sheet.sf6_gas_sample = True
    else:
        test_sheet.sf6_gas_sample = False
    if "sf6_gas_leaks" in request.POST:
        test_sheet.sf6_gas_leaks = True
    else:
        test_sheet.sf6_gas_leaks = False

    #METER BOOLEAN FIELDS
    if "unit_grounded_manufacturer" in request.POST:
        test_sheet.unit_grounded_manufacturer = True
    else:
        test_sheet.unit_grounded_manufacturer = False
    if "unit_connected_manfacturer_drawings" in request.POST:
        test_sheet.unit_connected_manfacturer_drawings = True
    else:
        test_sheet.unit_connected_manfacturer_drawings = False
    if "parameters_ratios_ect" in request.POST:
        test_sheet.parameters_ratios_ect = True
    else:
        test_sheet.parameters_ratios_ect = False
    if "correct_aux_in_out" in request.POST:
        test_sheet.correct_aux_in_out = True
    else:
        test_sheet.correct_aux_in_out = False
    if "measurements_indications_consistant_standards" in request.POST:
        test_sheet.measurements_indications_consistant_standards = True
    else:
        test_sheet.measurements_indications_consistant_standards = False        

    #MEDIUM VOLTAGE AIR BREAKER BOOLEAN FIELDS
    if "mvab_arc_chutes_intact" in request.POST:
        test_sheet.mvab_arc_chutes_intact = True
    else:
        test_sheet.mvab_arc_chutes_intact = False
    if "mvab_check_binding_friction_ect" in request.POST:
        test_sheet.mvab_check_binding_friction_ect = True
    else:
        test_sheet.mvab_check_binding_friction_ect = False
    if "mvab_puffer_operation" in request.POST:
        test_sheet.mvab_puffer_operation = True
    else:
        test_sheet.mvab_puffer_operation = False

    #OIL BREAKER BOOLEAN FIELDS
    if "ofcb_oil_level_tanks_bushings" in request.POST:
        test_sheet.ofcb_oil_level_tanks_bushings = True
    else:
        test_sheet.ofcb_oil_level_tanks_bushings = False
    if "ofcb_vents_clear" in request.POST:
        test_sheet.ofcb_vents_clear = True
    else:
        test_sheet.ofcb_vents_clear = False
    if "ofcb_hydraulic_air_inspected" in request.POST:
        test_sheet.ofcb_hydraulic_air_inspected = True
    else:
        test_sheet.ofcb_hydraulic_air_inspected = False
    if "ofcb_alarms_pressure_switches_operators" in request.POST:
        test_sheet.ofcb_alarms_pressure_switches_operators = True
    else:
        test_sheet.ofcb_alarms_pressure_switches_operators = False
    if "ofcb_inspect_bottom_for_parts_debris" in request.POST:
        test_sheet.ofcb_inspect_bottom_for_parts_debris = True
    else:
        test_sheet.ofcb_inspect_bottom_for_parts_debris = False
    if "ofcb_lift_rod_ect" in request.POST:
        test_sheet.ofcb_lift_rod_ect = True
    else:
        test_sheet.ofcb_lift_rod_ect = False
    if "ofcb_contact_sequence" in request.POST:
        test_sheet.ofcb_contact_sequence = True
    else:
        test_sheet.ofcb_contact_sequence = False
    if "ofcb_refill_tanks" in request.POST:
        test_sheet.ofcb_refill_tanks = True
    else:
        test_sheet.ofcb_refill_tanks = False
    #DRY TRANSFORMER BOOLEAN FIELDS
    if "xfmr_resilient_mounts" in request.POST:
        test_sheet.xfmr_resilient_mounts = True
    else:
        test_sheet.xfmr_resilient_mounts = False
    if "xfmr_temp_indicators" in request.POST:
        test_sheet.xfmr_temp_indicators = True
    else:
        test_sheet.xfmr_temp_indicators = False
    if "xfmr_fans_oc_protection" in request.POST:
        test_sheet.xfmr_fans_oc_protection = True
    else:
        test_sheet.xfmr_fans_oc_protection = False
    if "xfmr_manufacture_inspections_mechanical_tests" in request.POST:
        test_sheet.xfmr_manufacture_inspections_mechanical_tests = True
    else:
        test_sheet.xfmr_manufacture_inspections_mechanical_tests = False
    if "xfmr_as_left_verified" in request.POST:
        test_sheet.xfmr_as_left_verified = True
    else:
        test_sheet.xfmr_as_left_verified = False
    if "xfmr_surge_arrestors_present" in request.POST:
        test_sheet.xfmr_surge_arrestors_present = True
    else:
        test_sheet.xfmr_surge_arrestors_present = False

    #OIL FILLED TRANSFORMER BOOLEAN FIELDS
    if "oil_xfmr_impact_recorder" in request.POST:
        test_sheet.oil_xfmr_impact_recorder = True
    else:
        test_sheet.oil_xfmr_impact_recorder = False
    if "oil_xfmr_dew_point" in request.POST:
        test_sheet.oil_xfmr_dew_point = True
    else:
        test_sheet.oil_xfmr_dew_point = False
    if "oil_xfmr_pcb_content_labeling" in request.POST:
        test_sheet.oil_xfmr_pcb_content_labeling = True
    else:
        test_sheet.oil_xfmr_pcb_content_labeling = False
    if "oil_xfmr_shipping_bracing" in request.POST:
        test_sheet.oil_xfmr_shipping_bracing = True
    else:
        test_sheet.oil_xfmr_shipping_bracing = False
    if "oil_xfmr_bushings_clean" in request.POST:
        test_sheet.oil_xfmr_bushings_clean = True
    else:
        test_sheet.oil_xfmr_bushings_clean = False
    if "oil_xfmr_alarm_temperature" in request.POST:
        test_sheet.oil_xfmr_alarm_temperature = True
    else:
        test_sheet.oil_xfmr_alarm_temperature = False
    if "oil_xfmr_alarm_control_trip_indicators_devices" in request.POST:
        test_sheet.oil_xfmr_alarm_control_trip_indicators_devices = True
    else:
        test_sheet.oil_xfmr_alarm_control_trip_indicators_devices = False
    if "oil_xfmr_verify_liquid_level" in request.POST:
        test_sheet.oil_xfmr_verify_liquid_level = True
    else:
        test_sheet.oil_xfmr_verify_liquid_level = False
    if "oil_xfmr_verify_valve_positions" in request.POST:
        test_sheet.oil_xfmr_verify_valve_positions = True
    else:
        test_sheet.oil_xfmr_verify_valve_positions = False
    if "oil_xfmr_verify_positive_pressure" in request.POST:
        test_sheet.oil_xfmr_verify_positive_pressure = True
    else:
        test_sheet.oil_xfmr_verify_positive_pressure = False
    if "oil_xfmr_load_tap_changer" in request.POST:
        test_sheet.oil_xfmr_load_tap_changer = True
    else:
        test_sheet.oil_xfmr_load_tap_changer = False
    if "oil_xfmr_denergized_as_left" in request.POST:
        test_sheet.oil_xfmr_denergized_as_left = True
    else:
        test_sheet.oil_xfmr_denergized_as_left = False

    #INSULATED/MOLDED-CASE BOOLEAN FIELDS
    if "insul_mol_lvcb_smooth_op" in request.POST:
        test_sheet.insul_mol_lvcb_smooth_op = True
    else:
        test_sheet.insul_mol_lvcb_smooth_op = False
    if "insul_mol_lvcb_inspect_mech_chutes" in request.POST:
        test_sheet.insul_mol_lvcb_inspect_mech_chutes = True
    else:
        test_sheet.insul_mol_lvcb_inspect_mech_chutes = False
    if "insul_mol_lvcb_adjustments_protective_settings" in request.POST:
        test_sheet.insul_mol_lvcb_adjustments_protective_settings = True
    else:
        test_sheet.insul_mol_lvcb_adjustments_protective_settings = False
        
    #LOW VOLTAGE CB BOOLEAN FIELDS
    if "lvcb_contacts_condition_alignment" in request.POST:
        test_sheet.lvcb_contacts_condition_alignment = True
    else:
        test_sheet.lvcb_contacts_condition_alignment = False
    if "lvcb_primary_secondary_dimensions_correct" in request.POST:
        test_sheet.lvcb_primary_secondary_dimensions_correct = True
    else:
        test_sheet.lvcb_primary_secondary_dimensions_correct = False
    if "lvcb_operator_mechanism_accordance_manufacturer_data" in request.POST:
        test_sheet.lvcb_operator_mechanism_accordance_manufacturer_data = True
    else:
        test_sheet.lvcb_operator_mechanism_accordance_manufacturer_data = False
    if "lvcb_cell_fit_element_alignment" in request.POST:
        test_sheet.lvcb_cell_fit_element_alignment = True
    else:
        test_sheet.lvcb_cell_fit_element_alignment = False


    #pi testing
    if "pi_setting_ltpu" in request.POST:
        if request.POST["pi_setting_ltpu"]!="":
            test_sheet.pi_setting_ltpu = request.POST["pi_setting_ltpu"]
        else:
            test_sheet.pi_setting_ltpu = None 
    if "pi_setting_ltd" in request.POST:
        if request.POST["pi_setting_ltd"]!="":
           test_sheet.pi_setting_ltd = request.POST["pi_setting_ltd"]
        else:
            test_sheet.pi_setting_ltd = None 
    if "pi_setting_stpu" in request.POST:
        if request.POST["pi_setting_stpu"]!="":
           test_sheet.pi_setting_stpu = request.POST["pi_setting_stpu"]
        else:
            test_sheet.pi_setting_stpu = None 
    if "pi_setting_std" in request.POST:
        if request.POST["pi_setting_std"]!="":
           test_sheet.pi_setting_std = request.POST["pi_setting_std"]
        else:
            test_sheet.pi_setting_std = None 
    if "pi_setting_inst" in request.POST:
        if request.POST["pi_setting_inst"]!="":
           test_sheet.pi_setting_inst = request.POST["pi_setting_inst"]
        else:
            test_sheet.pi_setting_inst = None 
    if "pi_setting_gfpu" in request.POST:
        if request.POST["pi_setting_gfpu"]!="":
           test_sheet.pi_setting_gfpu = request.POST["pi_setting_gfpu"]
        else:
            test_sheet.pi_setting_gfpu = None 
    if "pi_setting_gfd" in request.POST:
        if request.POST["pi_setting_gfd"]!="":
           test_sheet.pi_setting_gfd = request.POST["pi_setting_gfd"]
        else:
            test_sheet.pi_setting_gfd = None 
    if "pi_testamps_ltd" in request.POST:
        if request.POST["pi_testamps_ltd"]!="":
           test_sheet.pi_testamps_ltd = request.POST["pi_testamps_ltd"]
        else:
            test_sheet.pi_testamps_ltd = None 
    if "pi_testamps_std" in request.POST:
        if request.POST["pi_testamps_std"]!="":
           test_sheet.pi_testamps_std = request.POST["pi_testamps_std"]
        else:
            test_sheet.pi_testamps_std = None 
    if "pi_testamps_inst" in request.POST:
        if request.POST["pi_testamps_inst"]!="":
           test_sheet.pi_testamps_inst = request.POST["pi_testamps_inst"]
        else:
            test_sheet.pi_testamps_inst = None 
    if "pi_testamps_gfd" in request.POST:
        if request.POST["pi_testamps_gfd"]!="":
           test_sheet.pi_testamps_gfd = request.POST["pi_testamps_gfd"]
        else:
            test_sheet.pi_testamps_gfd = None 
    if "pi_xpu_ltd" in request.POST:
        if request.POST["pi_xpu_ltd"]!="":
           test_sheet.pi_xpu_ltd = request.POST["pi_xpu_ltd"]
        else:
            test_sheet.pi_xpu_ltd = None 
    if "pi_xpu_std" in request.POST:
        if request.POST["pi_xpu_std"]!="":
           test_sheet.pi_xpu_std = request.POST["pi_xpu_std"]
        else:
            test_sheet.pi_xpu_std = None 
    if "pi_xpu_inst" in request.POST:
        if request.POST["pi_xpu_inst"]!="":
           test_sheet.pi_xpu_inst = request.POST["pi_xpu_inst"]
        else:
            test_sheet.pi_xpu_inst = None 
    if "pi_xpu_gfd" in request.POST:
        if request.POST["pi_xpu_gfd"]!="":
           test_sheet.pi_xpu_gfd = request.POST["pi_xpu_gfd"]
        else:
            test_sheet.pi_xpu_gfd = None 
    if "a_af_ltpu" in request.POST:
        if request.POST["a_af_ltpu"]!="":
           test_sheet.a_af_ltpu = request.POST["a_af_ltpu"]
        else:
            test_sheet.a_af_ltpu = None 
    if "a_af_ltd" in request.POST:
        if request.POST["a_af_ltd"]!="":
           test_sheet.a_af_ltd = request.POST["a_af_ltd"]
        else:
            test_sheet.a_af_ltd = None 
    if "a_af_stpu" in request.POST:
        if request.POST["a_af_stpu"]!="":
           test_sheet.a_af_stpu = request.POST["a_af_stpu"]
        else:
            test_sheet.a_af_stpu = None 
    if "a_af_std" in request.POST:
        if request.POST["a_af_std"]!="":
           test_sheet.a_af_std = request.POST["a_af_std"]
        else:
            test_sheet.a_af_std = None 
    if "a_af_inst" in request.POST:
        if request.POST["a_af_inst"]!="":
           test_sheet.a_af_inst = request.POST["a_af_inst"]
        else:
            test_sheet.a_af_inst = None 
    if "a_af_gfpu" in request.POST:
        if request.POST["a_af_gfpu"]!="":
           test_sheet.a_af_gfpu = request.POST["a_af_gfpu"]
        else:
            test_sheet.a_af_gfpu = None 
    if "a_af_gfd" in request.POST:
        if request.POST["a_af_gfd"]!="":
           test_sheet.a_af_gfd = request.POST["a_af_gfd"]
        else:
            test_sheet.a_af_gfd = None 
    if "a_al_ltpu" in request.POST:
        if request.POST["a_al_ltpu"]!="":
           test_sheet.a_al_ltpu = request.POST["a_al_ltpu"]
        else:
            test_sheet.a_al_ltpu = None 
    if "a_al_ltd" in request.POST:
        if request.POST["a_al_ltd"]!="":
           test_sheet.a_al_ltd = request.POST["a_al_ltd"]
        else:
            test_sheet.a_al_ltd = None 
    if "a_al_stpu" in request.POST:
        if request.POST["a_al_stpu"]!="":
           test_sheet.a_al_stpu = request.POST["a_al_stpu"]
        else:
            test_sheet.a_al_stpu = None 
    if "a_al_std" in request.POST:
        if request.POST["a_al_std"]!="":
           test_sheet.a_al_std = request.POST["a_al_std"]
        else:
            test_sheet.a_al_std = None 
    if "a_al_inst" in request.POST:
        if request.POST["a_al_inst"]!="":
           test_sheet.a_al_inst = request.POST["a_al_inst"]
        else:
            test_sheet.a_al_inst = None 
    if "a_al_gfpu" in request.POST:
        if request.POST["a_al_gfpu"]!="":
           test_sheet.a_al_gfpu = request.POST["a_al_gfpu"]
        else:
            test_sheet.a_al_gfpu = None 
    if "a_al_gfd" in request.POST:
        if request.POST["a_al_gfd"]!="":
           test_sheet.a_al_gfd = request.POST["a_al_gfd"]
        else:
            test_sheet.a_al_gfd = None 
    if "b_af_ltpu" in request.POST:
        if request.POST["b_af_ltpu"]!="":
           test_sheet.b_af_ltpu = request.POST["b_af_ltpu"]
        else:
            test_sheet.b_af_ltpu = None 
    if "b_af_ltd" in request.POST:
        if request.POST["b_af_ltd"]!="":
           test_sheet.b_af_ltd = request.POST["b_af_ltd"]
        else:
            test_sheet.b_af_ltd = None 
    if "b_af_stpu" in request.POST:
        if request.POST["b_af_stpu"]!="":
           test_sheet.b_af_stpu = request.POST["b_af_stpu"]
        else:
            test_sheet.b_af_stpu = None 
    if "b_af_std" in request.POST:
        if request.POST["b_af_std"]!="":
           test_sheet.b_af_std = request.POST["b_af_std"]
        else:
            test_sheet.b_af_std = None 
    if "b_af_inst" in request.POST:
        if request.POST["b_af_inst"]!="":
           test_sheet.b_af_inst = request.POST["b_af_inst"]
        else:
            test_sheet.b_af_inst = None 
    if "b_af_gfpu" in request.POST:
        if request.POST["b_af_gfpu"]!="":
           test_sheet.b_af_gfpu = request.POST["b_af_gfpu"]
        else:
            test_sheet.b_af_gfpu = None 
    if "b_af_gfd" in request.POST:
        if request.POST["b_af_gfd"]!="":
           test_sheet.b_af_gfd = request.POST["b_af_gfd"]
        else:
            test_sheet.b_af_gfd = None 
    if "b_al_ltpu" in request.POST:
        if request.POST["b_al_ltpu"]!="":
           test_sheet.b_al_ltpu = request.POST["b_al_ltpu"]
        else:
            test_sheet.b_al_ltpu = None 
    if "b_al_ltd" in request.POST:
        if request.POST["b_al_ltd"]!="":
           test_sheet.b_al_ltd = request.POST["b_al_ltd"]
        else:
            test_sheet.b_al_ltd = None 
    if "b_al_stpu" in request.POST:
        if request.POST["b_al_stpu"]!="":
           test_sheet.b_al_stpu = request.POST["b_al_stpu"]
        else:
            test_sheet.b_al_stpu = None 
    if "b_al_std" in request.POST:
        if request.POST["b_al_std"]!="":
           test_sheet.b_al_std = request.POST["b_al_std"]
        else:
            test_sheet.b_al_std = None 
    if "b_al_inst" in request.POST:
        if request.POST["b_al_inst"]!="":
           test_sheet.b_al_inst = request.POST["b_al_inst"]
        else:
            test_sheet.b_al_inst = None 
    if "b_al_gfpu" in request.POST:
        if request.POST["b_al_gfpu"]!="":
           test_sheet.b_al_gfpu = request.POST["b_al_gfpu"]
        else:
            test_sheet.b_al_gfpu = None 
    if "b_al_gfd" in request.POST:
        if request.POST["b_al_gfd"]!="":
           test_sheet.b_al_gfd = request.POST["b_al_gfd"]
        else:
            test_sheet.b_al_gfd = None 
    if "c_af_ltpu" in request.POST:
        if request.POST["c_af_ltpu"]!="":
           test_sheet.c_af_ltpu = request.POST["c_af_ltpu"]
        else:
            test_sheet.c_af_ltpu = None 
    if "c_af_ltd" in request.POST:
        if request.POST["c_af_ltd"]!="":
           test_sheet.c_af_ltd = request.POST["c_af_ltd"]
        else:
            test_sheet.c_af_ltd = None 
    if "c_af_stpu" in request.POST:
        if request.POST["c_af_stpu"]!="":
           test_sheet.c_af_stpu = request.POST["c_af_stpu"]
        else:
            test_sheet.c_af_stpu = None 
    if "c_af_std" in request.POST:
        if request.POST["c_af_std"]!="":
           test_sheet.c_af_std = request.POST["c_af_std"]
        else:
            test_sheet.c_af_std = None 
    if "c_af_inst" in request.POST:
        if request.POST["c_af_inst"]!="":
           test_sheet.c_af_inst = request.POST["c_af_inst"]
        else:
            test_sheet.c_af_inst = None 
    if "c_af_gfpu" in request.POST:
        if request.POST["c_af_gfpu"]!="":
           test_sheet.c_af_gfpu = request.POST["c_af_gfpu"]
        else:
            test_sheet.c_af_gfpu = None 
    if "c_af_gfd" in request.POST:
        if request.POST["c_af_gfd"]!="":
           test_sheet.c_af_gfd = request.POST["c_af_gfd"]
        else:
            test_sheet.c_af_gfd = None 
    if "c_al_ltpu" in request.POST:
        if request.POST["c_al_ltpu"]!="":
           test_sheet.c_al_ltpu = request.POST["c_al_ltpu"]
        else:
            test_sheet.c_al_ltpu = None 
    if "c_al_ltd" in request.POST:
        if request.POST["c_al_ltd"]!="":
           test_sheet.c_al_ltd = request.POST["c_al_ltd"]
        else:
            test_sheet.c_al_ltd = None 
    if "c_al_stpu" in request.POST:
        if request.POST["c_al_stpu"]!="":
           test_sheet.c_al_stpu = request.POST["c_al_stpu"]
        else:
            test_sheet.c_al_stpu = None 
    if "c_al_std" in request.POST:
        if request.POST["c_al_std"]!="":
           test_sheet.c_al_std = request.POST["c_al_std"]
        else:
            test_sheet.c_al_std = None 
    if "c_al_inst" in request.POST:
        if request.POST["c_al_inst"]!="":
           test_sheet.c_al_inst = request.POST["c_al_inst"]
        else:
            test_sheet.c_al_inst = None 
    if "c_al_gfpu" in request.POST:
        if request.POST["c_al_gfpu"]!="":
           test_sheet.c_al_gfpu = request.POST["c_al_gfpu"]
        else:
            test_sheet.c_al_gfpu = None 
    if "c_al_gfd" in request.POST:
        if request.POST["c_al_gfd"]!="":
           test_sheet.c_al_gfd = request.POST["c_al_gfd"]
        else:
            test_sheet.c_al_gfd = None 
    if "min_af_ltpu" in request.POST:
        if request.POST["min_af_ltpu"]!="":
           test_sheet.min_af_ltpu = request.POST["min_af_ltpu"]
        else:
            test_sheet.min_af_ltpu = None 
    if "min_af_ltd" in request.POST:
        if request.POST["min_af_ltd"]!="":
           test_sheet.min_af_ltd = request.POST["min_af_ltd"]
        else:
            test_sheet.min_af_ltd = None 
    if "min_af_stpu" in request.POST:
        if request.POST["min_af_stpu"]!="":
           test_sheet.min_af_stpu = request.POST["min_af_stpu"]
        else:
            test_sheet.min_af_stpu = None 
    if "min_af_std" in request.POST:
        if request.POST["min_af_std"]!="":
           test_sheet.min_af_std = request.POST["min_af_std"]
        else:
            test_sheet.min_af_std = None 
    if "min_af_inst" in request.POST:
        if request.POST["min_af_inst"]!="":
           test_sheet.min_af_inst = request.POST["min_af_inst"]
        else:
            test_sheet.min_af_inst = None 
    if "min_af_gfpu" in request.POST:
        if request.POST["min_af_gfpu"]!="":
           test_sheet.min_af_gfpu = request.POST["min_af_gfpu"]
        else:
            test_sheet.min_af_gfpu = None 
    if "min_af_gfd" in request.POST:
        if request.POST["min_af_gfd"]!="":
           test_sheet.min_af_gfd = request.POST["min_af_gfd"]
        else:
            test_sheet.min_af_gfd = None 
    if "min_al_ltpu" in request.POST:
        if request.POST["min_al_ltpu"]!="":
           test_sheet.min_al_ltpu = request.POST["min_al_ltpu"]
        else:
            test_sheet.min_al_ltpu = None 
    if "min_al_ltd" in request.POST:
        if request.POST["min_al_ltd"]!="":
           test_sheet.min_al_ltd = request.POST["min_al_ltd"]
        else:
            test_sheet.min_al_ltd = None 
    if "min_al_stpu" in request.POST:
        if request.POST["min_al_stpu"]!="":
           test_sheet.min_al_stpu = request.POST["min_al_stpu"]
        else:
            test_sheet.min_al_stpu = None 
    if "min_al_std" in request.POST:
        if request.POST["min_al_std"]!="":
           test_sheet.min_al_std = request.POST["min_al_std"]
        else:
            test_sheet.min_al_std = None 
    if "min_al_inst" in request.POST:
        if request.POST["min_al_inst"]!="":
           test_sheet.min_al_inst = request.POST["min_al_inst"]
        else:
            test_sheet.min_al_inst = None 
    if "min_al_gfpu" in request.POST:
        if request.POST["min_al_gfpu"]!="":
           test_sheet.min_al_gfpu = request.POST["min_al_gfpu"]
        else:
            test_sheet.min_al_gfpu = None 
    if "min_al_gfd" in request.POST:
        if request.POST["min_al_gfd"]!="":
           test_sheet.min_al_gfd = request.POST["min_al_gfd"]
        else:
            test_sheet.min_al_gfd = None 
    if "max_af_ltpu" in request.POST:
        if request.POST["max_af_ltpu"]!="":
           test_sheet.max_af_ltpu = request.POST["max_af_ltpu"]
        else:
            test_sheet.max_af_ltpu = None 
    if "max_af_ltd" in request.POST:
        if request.POST["max_af_ltd"]!="":
           test_sheet.max_af_ltd = request.POST["max_af_ltd"]
        else:
            test_sheet.max_af_ltd = None 
    if "max_af_stpu" in request.POST:
        if request.POST["max_af_stpu"]!="":
           test_sheet.max_af_stpu = request.POST["max_af_stpu"]
        else:
            test_sheet.max_af_stpu = None 
    if "max_af_std" in request.POST:
        if request.POST["max_af_std"]!="":
           test_sheet.max_af_std = request.POST["max_af_std"]
        else:
            test_sheet.max_af_std = None 
    if "max_af_inst" in request.POST:
        if request.POST["max_af_inst"]!="":
           test_sheet.max_af_inst = request.POST["max_af_inst"]
        else:
            test_sheet.max_af_inst = None 
    if "max_af_gfpu" in request.POST:
        if request.POST["max_af_gfpu"]!="":
           test_sheet.max_af_gfpu = request.POST["max_af_gfpu"]
        else:
            test_sheet.max_af_gfpu = None 
    if "max_af_gfd" in request.POST:
        if request.POST["max_af_gfd"]!="":
           test_sheet.max_af_gfd = request.POST["max_af_gfd"]
        else:
            test_sheet.max_af_gfd = None 
    if "max_al_ltpu" in request.POST:
        if request.POST["max_al_ltpu"]!="":
           test_sheet.max_al_ltpu = request.POST["max_al_ltpu"]
        else:
            test_sheet.max_al_ltpu = None 
    if "max_al_ltd" in request.POST:
        if request.POST["max_al_ltd"]!="":
           test_sheet.max_al_ltd = request.POST["max_al_ltd"]
        else:
            test_sheet.max_al_ltd = None 
    if "max_al_stpu" in request.POST:
        if request.POST["max_al_stpu"]!="":
           test_sheet.max_al_stpu = request.POST["max_al_stpu"]
        else:
            test_sheet.max_al_stpu = None 
    if "max_al_std" in request.POST:
        if request.POST["max_al_std"]!="":
           test_sheet.max_al_std = request.POST["max_al_std"]
        else:
            test_sheet.max_al_std = None 
    if "max_al_inst" in request.POST:
        if request.POST["max_al_inst"]!="":
           test_sheet.max_al_inst = request.POST["max_al_inst"]
        else:
            test_sheet.max_al_inst = None 
    if "max_al_gfpu" in request.POST:
        if request.POST["max_al_gfpu"]!="":
           test_sheet.max_al_gfpu = request.POST["max_al_gfpu"]
        else:
            test_sheet.max_al_gfpu = None 
    if "max_al_gfd" in request.POST:
        if request.POST["max_al_gfd"]!="":
           test_sheet.max_al_gfd = request.POST["max_al_gfd"]
        else:
            test_sheet.max_al_gfd = None 
    if "si_pu_a" in request.POST:
        if request.POST["si_pu_a"]!="":
           test_sheet.si_pu_a = request.POST["si_pu_a"]
        else:
            test_sheet.si_pu_a = None 
    if "si_pu_b" in request.POST:
        if request.POST["si_pu_b"]!="":
           test_sheet.si_pu_b = request.POST["si_pu_b"]
        else:
            test_sheet.si_pu_b = None 
    if "si_pu_c" in request.POST:
        if request.POST["si_pu_c"]!="":
           test_sheet.si_pu_c = request.POST["si_pu_c"]
        else:
            test_sheet.si_pu_c = None 
    if "si_lt_current_a" in request.POST:
        if request.POST["si_lt_current_a"]!="":
           test_sheet.si_lt_current_a = request.POST["si_lt_current_a"]
        else:
            test_sheet.si_lt_current_a = None 
    if "si_lt_current_b" in request.POST:
        if request.POST["si_lt_current_b"]!="":
           test_sheet.si_lt_current_b = request.POST["si_lt_current_b"]
        else:
            test_sheet.si_lt_current_b = None 
    if "si_lt_current_c" in request.POST:
        if request.POST["si_lt_current_c"]!="":
           test_sheet.si_lt_current_c = request.POST["si_lt_current_c"]
        else:
            test_sheet.si_lt_current_c = None 
    if "si_lt_d_a" in request.POST:
        if request.POST["si_lt_d_a"]!="":
           test_sheet.si_lt_d_a = request.POST["si_lt_d_a"]
        else:
            test_sheet.si_lt_d_a = None 
    if "si_lt_d_b" in request.POST:
        if request.POST["si_lt_d_b"]!="":
           test_sheet.si_lt_d_b = request.POST["si_lt_d_b"]
        else:
            test_sheet.si_lt_d_b = None 
    if "si_lt_d_c" in request.POST:
        if request.POST["si_lt_d_c"]!="":
           test_sheet.si_lt_d_c = request.POST["si_lt_d_c"]
        else:
            test_sheet.si_lt_d_c = None 
    if "si_st_current_a" in request.POST:
        if request.POST["si_st_current_a"]!="":
           test_sheet.si_st_current_a = request.POST["si_st_current_a"]
        else:
            test_sheet.si_st_current_a = None 
    if "si_st_current_b" in request.POST:
        if request.POST["si_st_current_b"]!="":
           test_sheet.si_st_current_b = request.POST["si_st_current_b"]
        else:
            test_sheet.si_st_current_b = None 
    if "si_st_current_c" in request.POST:
        if request.POST["si_st_current_c"]!="":
           test_sheet.si_st_current_c = request.POST["si_st_current_c"]
        else:
            test_sheet.si_st_current_c = None 
    if "si_st_d_a" in request.POST:
        if request.POST["si_st_d_a"]!="":
           test_sheet.si_st_d_a = request.POST["si_st_d_a"]
        else:
            test_sheet.si_st_d_a = None 
    if "si_st_d_b" in request.POST:
        if request.POST["si_st_d_b"]!="":
           test_sheet.si_st_d_b = request.POST["si_st_d_b"]
        else:
            test_sheet.si_st_d_b = None 
    if "si_st_d_c" in request.POST:
        if request.POST["si_st_d_c"]!="":
           test_sheet.si_st_d_c = request.POST["si_st_d_c"]
        else:
            test_sheet.si_st_d_c = None 
    if "si_inst_current_a" in request.POST:
        if request.POST["si_inst_current_a"]!="":
           test_sheet.si_inst_current_a = request.POST["si_inst_current_a"]
        else:
            test_sheet.si_inst_current_a = None 
    if "si_inst_current_b" in request.POST:
        if request.POST["si_inst_current_b"]!="":
           test_sheet.si_inst_current_b = request.POST["si_inst_current_b"]
        else:
            test_sheet.si_inst_current_b = None 
    if "si_inst_current_c" in request.POST:
        if request.POST["si_inst_current_c"]!="":
           test_sheet.si_inst_current_c = request.POST["si_inst_current_c"]
        else:
            test_sheet.si_inst_current_c = None    
    if "ttr_upper_tolerance" in request.POST:
        if request.POST["ttr_upper_tolerance"]!="":
           test_sheet.ttr_upper_tolerance = request.POST["ttr_upper_tolerance"]
           if is_casting:
               mold.ttr_upper_tolerance = request.POST["ttr_upper_tolerance"]
        else:
            if is_casting:
                mold.ttr_upper_tolerance = None    
            test_sheet.ttr_upper_tolerance = None    
    if "ttr_lower_tolerance" in request.POST:
        if request.POST["ttr_lower_tolerance"]!="":
           test_sheet.ttr_lower_tolerance = -abs(Decimal(request.POST["ttr_lower_tolerance"])) 
           if is_casting:
               mold.ttr_lower_tolerance = -abs(Decimal(request.POST["ttr_lower_tolerance"])) 
        else:
           if is_casting:
               mold.ttr_lower_tolerance = None
           test_sheet.ttr_lower_tolerance = None
    if "tap_one_volts" in request.POST:
        if request.POST["tap_one_volts"]!="":
           test_sheet.tap_one_volts = request.POST["tap_one_volts"]
           if is_casting:
               mold.tap_one_volts = request.POST["tap_one_volts"]
        else:
            if is_casting:
                mold.tap_one_volts = None
            test_sheet.tap_one_volts = None
    if "tap_two_volts" in request.POST:
        if request.POST["tap_two_volts"]!="":
           test_sheet.tap_two_volts = request.POST["tap_two_volts"]
           if is_casting:
               mold.tap_two_volts = request.POST["tap_two_volts"]
        else:
            if is_casting:
                mold.tap_two_volts = None  
            test_sheet.tap_two_volts = None  
    if "tap_three_volts" in request.POST:
        if request.POST["tap_three_volts"]!="":
           test_sheet.tap_three_volts = request.POST["tap_three_volts"]
           if is_casting:
               mold.tap_three_volts = request.POST["tap_three_volts"]
        else:
            if is_casting:
                mold.tap_three_volts = None
            test_sheet.tap_three_volts = None
    if "tap_four_volts" in request.POST:
        if request.POST["tap_four_volts"]!="":
           test_sheet.tap_four_volts = request.POST["tap_four_volts"]
           if is_casting:
               mold.tap_four_volts = request.POST["tap_four_volts"]
        else:
            if is_casting:
                mold.tap_four_volts = None    
            test_sheet.tap_four_volts = None    

    if "tap_five_volts" in request.POST:
        if request.POST["tap_five_volts"]!="":
           test_sheet.tap_five_volts = request.POST["tap_five_volts"]
           if is_casting:
               mold.tap_five_volts = request.POST["tap_five_volts"]
        else:
            if is_casting:
                mold.tap_five_volts = None
            test_sheet.tap_five_volts = None
    if "tap_six_volts" in request.POST:
        if request.POST["tap_six_volts"]!="":
           test_sheet.tap_six_volts = request.POST["tap_six_volts"]
           if is_casting:
               mold.tap_six_volts = request.POST["tap_six_volts"]
        else:
            if is_casting:
                mold.tap_six_volts = None    
            test_sheet.tap_six_volts = None    

    if "tap_seven_volts" in request.POST:
        if request.POST["tap_seven_volts"]!="":
           test_sheet.tap_seven_volts = request.POST["tap_seven_volts"]
           if is_casting:
               mold.tap_seven_volts = request.POST["tap_seven_volts"]
        else:
            if is_casting:
                mold.tap_seven_volts = None
            test_sheet.tap_seven_volts = None
    if "tap_one_ratio" in request.POST:
        if request.POST["tap_one_ratio"]!="":
           test_sheet.tap_one_ratio = request.POST["tap_one_ratio"]
        else:
            test_sheet.tap_one_ratio = None    
    if "tap_two_ratio" in request.POST:
        if request.POST["tap_two_ratio"]!="":
           test_sheet.tap_two_ratio = request.POST["tap_two_ratio"]
        else:
            test_sheet.tap_two_ratio = None    
    if "tap_three_ratio" in request.POST:
        if request.POST["tap_three_ratio"]!="":
           test_sheet.tap_three_ratio = request.POST["tap_three_ratio"]
        else:
            test_sheet.tap_three_ratio = None    
    if "tap_four_ratio" in request.POST:
        if request.POST["tap_four_ratio"]!="":
           test_sheet.tap_four_ratio = request.POST["tap_four_ratio"]
        else:
            test_sheet.tap_four_ratio = None    
    if "tap_five_ratio" in request.POST:
        if request.POST["tap_five_ratio"]!="":
           test_sheet.tap_five_ratio = request.POST["tap_five_ratio"]
        else:
            test_sheet.tap_five_ratio = None    
    if "tap_six_ratio" in request.POST:
        if request.POST["tap_six_ratio"]!="":
           test_sheet.tap_six_ratio = request.POST["tap_six_ratio"]
        else:
            test_sheet.tap_six_ratio = None    
    if "tap_seven_ratio" in request.POST:
        if request.POST["tap_seven_ratio"]!="":
           test_sheet.tap_seven_ratio = request.POST["tap_seven_ratio"]
        else:
            test_sheet.tap_seven_ratio = None    
    if "tap_one_h12_x02_ttr" in request.POST:
        if request.POST["tap_one_h12_x02_ttr"]!="":
           test_sheet.tap_one_h12_x02_ttr = request.POST["tap_one_h12_x02_ttr"]
        else:
            test_sheet.tap_one_h12_x02_ttr = None    
    if "tap_two_h12_x02_ttr" in request.POST:
        if request.POST["tap_two_h12_x02_ttr"]!="":
           test_sheet.tap_two_h12_x02_ttr = request.POST["tap_two_h12_x02_ttr"]
        else:
            test_sheet.tap_three_h12_x02_ttr = None
    if "tap_three_h12_x02_ttr" in request.POST:
        if request.POST["tap_three_h12_x02_ttr"]!="":
           test_sheet.tap_three_h12_x02_ttr = request.POST["tap_three_h12_x02_ttr"]
        else:
            test_sheet.tap_three_h12_x02_ttr = None
    if "tap_four_h12_x02_ttr" in request.POST:
        if request.POST["tap_four_h12_x02_ttr"]!="":
           test_sheet.tap_four_h12_x02_ttr = request.POST["tap_four_h12_x02_ttr"]
        else:
            test_sheet.tap_four_h12_x02_ttr = None
    if "tap_five_h12_x02_ttr" in request.POST:
        if request.POST["tap_five_h12_x02_ttr"]!="":
           test_sheet.tap_five_h12_x02_ttr = request.POST["tap_five_h12_x02_ttr"]
        else:
            test_sheet.tap_five_h12_x02_ttr = None
    if "tap_six_h12_x02_ttr" in request.POST:
        if request.POST["tap_six_h12_x02_ttr"]!="":
           test_sheet.tap_six_h12_x02_ttr = request.POST["tap_six_h12_x02_ttr"]
        else:
            test_sheet.tap_six_h12_x02_ttr = None
    if "tap_seven_h12_x02_ttr" in request.POST:
        if request.POST["tap_seven_h12_x02_ttr"]!="":
           test_sheet.tap_seven_h12_x02_ttr = request.POST["tap_seven_h12_x02_ttr"]
        else:
            test_sheet.tap_seven_h12_x02_ttr = None
    if "tap_one_h23_x03_ttr" in request.POST:
        if request.POST["tap_one_h23_x03_ttr"]!="":
           test_sheet.tap_one_h23_x03_ttr = request.POST["tap_one_h23_x03_ttr"]
        else:
            test_sheet.tap_one_h23_x03_ttr = None    
    if "tap_two_h23_x03_ttr" in request.POST:
        if request.POST["tap_two_h23_x03_ttr"]!="":
           test_sheet.tap_two_h23_x03_ttr = request.POST["tap_two_h23_x03_ttr"]
        else:
            test_sheet.tap_two_h23_x03_ttr = None
    if "tap_three_h23_x03_ttr" in request.POST:
        if request.POST["tap_three_h23_x03_ttr"]!="":
           test_sheet.tap_three_h23_x03_ttr = request.POST["tap_three_h23_x03_ttr"]
        else:
            test_sheet.tap_three_h23_x03_ttr = None
    if "tap_four_h23_x03_ttr" in request.POST:
        if request.POST["tap_four_h23_x03_ttr"]!="":
           test_sheet.tap_four_h23_x03_ttr = request.POST["tap_four_h23_x03_ttr"]
        else:
            test_sheet.tap_four_h23_x03_ttr = None
    if "tap_five_h23_x03_ttr" in request.POST:
        if request.POST["tap_five_h23_x03_ttr"]!="":
           test_sheet.tap_five_h23_x03_ttr = request.POST["tap_five_h23_x03_ttr"]
        else:
            test_sheet.tap_five_h23_x03_ttr = None
    if "tap_six_h23_x03_ttr" in request.POST:
        if request.POST["tap_six_h23_x03_ttr"]!="":
           test_sheet.tap_six_h23_x03_ttr = request.POST["tap_six_h23_x03_ttr"]
        else:
            test_sheet.tap_six_h23_x03_ttr = None
    if "tap_seven_h23_x03_ttr" in request.POST:
        if request.POST["tap_seven_h23_x03_ttr"]!="":
           test_sheet.tap_seven_h23_x03_ttr = request.POST["tap_seven_h23_x03_ttr"]
        else:
            test_sheet.tap_seven_h23_x03_ttr = None
    if "tap_one_h31_x01_ttr" in request.POST:
        if request.POST["tap_one_h31_x01_ttr"]!="":
           test_sheet.tap_one_h31_x01_ttr = request.POST["tap_one_h31_x01_ttr"]
        else:
            test_sheet.tap_one_h31_x01_ttr = None   
    if "tap_two_h31_x01_ttr" in request.POST:
        if request.POST["tap_two_h31_x01_ttr"]!="":
           test_sheet.tap_two_h31_x01_ttr = request.POST["tap_two_h31_x01_ttr"]
        else:
            test_sheet.tap_two_h31_x01_ttr = None
    if "tap_three_h31_x01_ttr" in request.POST:
        if request.POST["tap_three_h31_x01_ttr"]!="":
           test_sheet.tap_three_h31_x01_ttr = request.POST["tap_three_h31_x01_ttr"]
        else:
            test_sheet.tap_three_h31_x01_ttr = None
    if "tap_four_h31_x01_ttr" in request.POST:
        if request.POST["tap_four_h31_x01_ttr"]!="":
           test_sheet.tap_four_h31_x01_ttr = request.POST["tap_four_h31_x01_ttr"]
        else:
            test_sheet.tap_four_h31_x01_ttr = None
    if "tap_five_h31_x01_ttr" in request.POST:
        if request.POST["tap_five_h31_x01_ttr"]!="":
           test_sheet.tap_five_h31_x01_ttr = request.POST["tap_five_h31_x01_ttr"]
        else:
            test_sheet.tap_five_h31_x01_ttr = None
    if "tap_six_h31_x01_ttr" in request.POST:
        if request.POST["tap_six_h31_x01_ttr"]!="":
           test_sheet.tap_six_h31_x01_ttr = request.POST["tap_six_h31_x01_ttr"]
        else:
            test_sheet.tap_six_h31_x01_ttr = None
    if "tap_seven_h31_x01_ttr" in request.POST:
        if request.POST["tap_seven_h31_x01_ttr"]!="":
           test_sheet.tap_seven_h31_x01_ttr = request.POST["tap_seven_h31_x01_ttr"]
        else:
            test_sheet.tap_seven_h31_x01_ttr = None
    if "wr_h1_h2" in request.POST:
        if request.POST["wr_h1_h2"]!="":
           test_sheet.wr_h1_h2 = request.POST["wr_h1_h2"]
        else:
            test_sheet.wr_h1_h2 = None
    if "wr_h2_h3" in request.POST:
        if request.POST["wr_h2_h3"]!="":
           test_sheet.wr_h2_h3 = request.POST["wr_h2_h3"]
        else:
            test_sheet.wr_h2_h3 = None
    if "wr_h3_h1" in request.POST:
        if request.POST["wr_h3_h1"]!="":
           test_sheet.wr_h3_h1 = request.POST["wr_h3_h1"]
        else:
            test_sheet.wr_h3_h1 = None
    if "wr_x0_x1" in request.POST:
        if request.POST["wr_x0_x1"]!="":
           test_sheet.wr_x0_x1 = request.POST["wr_x0_x1"]
        else:
            test_sheet.wr_x0_x1 = None
    if "wr_x0_x2" in request.POST:
        if request.POST["wr_x0_x2"]!="":
           test_sheet.wr_x0_x2 = request.POST["wr_x0_x2"]
        else:
            test_sheet.wr_x0_x2 = None
    if "wr_x0_x3" in request.POST:
        if request.POST["wr_x0_x3"]!="":
           test_sheet.wr_x0_x3 = request.POST["wr_x0_x3"]
        else:
            test_sheet.wr_x0_x3 = None
    if "hipot_test_voltage" in request.POST:
        if request.POST["hipot_test_voltage"]!="":
            test_sheet.hipot_test_voltage = request.POST["hipot_test_voltage"]
        else:
            test_sheet.hipot_test_voltage = None     
    if "hipot_ptp_ab" in request.POST:
        if request.POST["hipot_ptp_ab"]!="":
            test_sheet.hipot_ptp_ab = request.POST["hipot_ptp_ab"]
        else:
            test_sheet.hipot_ptp_ab = None     
    if "hipot_ptp_bc" in request.POST:
        if request.POST["hipot_ptp_bc"]!="":
            test_sheet.hipot_ptp_bc = request.POST["hipot_ptp_bc"]
        else:
            test_sheet.hipot_ptp_bc = None     
    if "hipot_ptp_ca" in request.POST:
        if request.POST["hipot_ptp_ca"]!="":
            test_sheet.hipot_ptp_ca = request.POST["hipot_ptp_ca"]
        else:
            test_sheet.hipot_ptp_ca = None     
    if "hipot_ltl_a" in request.POST:
        if request.POST["hipot_ltl_a"]!="":
            test_sheet.hipot_ltl_a = request.POST["hipot_ltl_a"]
        else:
            test_sheet.hipot_ltl_a = None     
    if "hipot_ltl_b" in request.POST:
        if request.POST["hipot_ltl_b"]!="":
            test_sheet.hipot_ltl_b = request.POST["hipot_ltl_b"]
        else:
            test_sheet.hipot_ltl_b = None     
    if "hipot_ltl_c" in request.POST:
        if request.POST["hipot_ltl_c"]!="":
            test_sheet.hipot_ltl_c = request.POST["hipot_ltl_c"]
        else:
            test_sheet.hipot_ltl_c = None     
    if "hipot_ptg_a" in request.POST:
        if request.POST["hipot_ptg_a"]!="":
            test_sheet.hipot_ptg_a = request.POST["hipot_ptg_a"]
        else:
            test_sheet.hipot_ptg_a = None     
    if "hipot_ptg_b" in request.POST:
        if request.POST["hipot_ptg_b"]!="":
            test_sheet.hipot_ptg_b = request.POST["hipot_ptg_b"]
        else:
            test_sheet.hipot_ptg_b = None     
    if "hipot_ptg_c" in request.POST:
        if request.POST["hipot_ptg_c"]!="":
            test_sheet.hipot_ptg_c = request.POST["hipot_ptg_c"]
        else:
            test_sheet.hipot_ptg_c = None     
    if "primary_voltage" in request.POST:
        if request.POST["primary_voltage"]!="":
           test_sheet.primary_voltage = request.POST["primary_voltage"]
           if is_casting:
               mold.primary_voltage = request.POST["primary_voltage"]
        else:
            if is_casting:
                mold.primary_voltage = None
            test_sheet.primary_voltage = None
    if "secondary_voltage" in request.POST:
        if request.POST["secondary_voltage"]!="":
           test_sheet.secondary_voltage = request.POST["secondary_voltage"]
           if is_casting:
               mold.secondary_voltage = request.POST["secondary_voltage"]
        else:
            if is_casting:
                mold.secondary_voltage = None
            test_sheet.secondary_voltage = None
    if "power_rating" in request.POST:
        if request.POST["power_rating"]!="":
           test_sheet.power_rating = request.POST["power_rating"]
           if is_casting:
               mold.power_rating = request.POST["power_rating"]
        else:
            if is_casting:
                mold.power_rating = None
            test_sheet.power_rating = None
    if "power_rating_units" in request.POST:
        if request.POST["power_rating_units"]!="":
           test_sheet.power_rating_units = request.POST["power_rating_units"]
           if is_casting:
               mold.power_rating_units = request.POST["power_rating_units"]
        else:
            if is_casting:
                mold.power_rating_units = None
            test_sheet.power_rating_units = None
    if "primary_winding_config" in request.POST:
        if request.POST["primary_winding_config"]!="":
           test_sheet.primary_winding_config = request.POST["primary_winding_config"]
           if is_casting:
               mold.primary_winding_config = request.POST["primary_winding_config"]
        else:
            if is_casting:
                mold.primary_winding_config = None
            test_sheet.primary_winding_config = None
    if "secondary_winding_config" in request.POST:
        if request.POST["secondary_winding_config"]!="":
           test_sheet.secondary_winding_config = request.POST["secondary_winding_config"]
           if is_casting:
               mold.secondary_winding_config = request.POST["secondary_winding_config"]
        else:
            if is_casting:
                mold.secondary_winding_config = None
            test_sheet.secondary_winding_config = None
    if "temp_rise" in request.POST:
        if request.POST["temp_rise"]!="":
            test_sheet.temp_rise = request.POST["temp_rise"]
            if is_casting:
                mold.temp_rise = request.POST["temp_rise"]
        else:
            if is_casting:
                mold.temp_rise = None
            test_sheet.temp_rise = None
    if "temp_rise_units" in request.POST:
        if request.POST["temp_rise_units"]!="":
            test_sheet.temp_rise_units = request.POST["temp_rise_units"]
            if is_casting:
                mold.temp_rise_units = request.POST["temp_rise_units"]
        else:
            if is_casting:
                mold.temp_rise_units = None
            test_sheet.temp_rise_units = None
    if "impedance" in request.POST:
        if request.POST["impedance"]!="":
            test_sheet.impedance = request.POST["impedance"]
            if is_casting:
                mold.impedance = request.POST["impedance"]
        else:
            if is_casting:
                mold.impedance = None
            test_sheet.impedance = None
    if "impedance_at" in request.POST:
        if request.POST["impedance_at"]!="":
            test_sheet.impedance_at = request.POST["impedance_at"]
            if is_casting:
                mold.impedance_at = request.POST["impedance_at"]
        else:
            if is_casting:
                mold.impedance_at = None
            test_sheet.impedance_at = None
    if "xfmr_class" in request.POST:
        if request.POST["xfmr_class"]!="":
            test_sheet.xfmr_class = request.POST["xfmr_class"]
            if is_casting:
                mold.xfmr_class = request.POST["xfmr_class"]
        else:
            if is_casting:
                mold.xfmr_class = None
            test_sheet.xfmr_class = None
    if "ambient_temp" in request.POST:
        if request.POST["ambient_temp"]!="":
            test_sheet.ambient_temp = request.POST["ambient_temp"]
        else:
            test_sheet.ambient_temp = None
    if "humidity" in request.POST:
        if request.POST["humidity"]!="":
            test_sheet.humidity = request.POST["humidity"]
        else:
            test_sheet.humidity = None
    if "ambient_temp_units" in request.POST:
        if request.POST["ambient_temp_units"]!="":
            test_sheet.ambient_temp_units = request.POST["ambient_temp_units"]
        else:
            test_sheet.ambient_temp_units = None
    if "tap_qty" in request.POST:
        if request.POST["tap_qty"]!="":
            test_sheet.tap_qty = request.POST["tap_qty"]
            if is_casting:
                mold.tap_qty = request.POST["tap_qty"]
        else:
            if is_casting:
                mold.tap_qty = None
            test_sheet.tap_qty = None
    if "tap_position" in request.POST:
        if request.POST["tap_position"]!="":
           test_sheet.tap_position = request.POST["tap_position"]
        else:
            test_sheet.tap_position = None
    if "fluid_capacity" in request.POST:
        if request.POST["fluid_capacity"]!="":
            test_sheet.fluid_capacity = request.POST["fluid_capacity"]
            if is_casting:
                mold.fluid_capacity = request.POST["fluid_capacity"]
        else:
            if is_casting:
                mold.fluid_capacity = None
            test_sheet.fluid_capacity = None
    if "fluid_capacity_units" in request.POST:
        if request.POST["fluid_capacity_units"]!="":
           test_sheet.fluid_capacity_units = request.POST["fluid_capacity_units"]
           if is_casting:
               mold.fluid_capacity_units = request.POST["fluid_capacity_units"]
        else:
            if is_casting:
                mold.fluid_capacity_units = None
            test_sheet.fluid_capacity_units = None
    if "liquid_level" in request.POST:
        if request.POST["liquid_level"]!="":
           test_sheet.liquid_level = request.POST["liquid_level"]
        else:
            test_sheet.liquid_level = None
    if "pressure" in request.POST:
        if request.POST["pressure"]!="":
           test_sheet.pressure = request.POST["pressure"]
        else:
            test_sheet.pressure = None
    if "pressure_units" in request.POST:
        if request.POST["pressure_units"]!="":
           test_sheet.pressure_units = request.POST["pressure_units"]
        else:
            test_sheet.pressure_units = None
    if "fluid_type" in request.POST:
        if request.POST["fluid_type"]!="":
            test_sheet.fluid_type = request.POST["fluid_type"]
            if is_casting:
                mold.fluid_type = request.POST["fluid_type"]
        else:
            test_sheet.fluid_type = None
            if is_casting:
                mold.fluid_type = None
    if "weight" in request.POST:
        if request.POST["weight"]!="":
            test_sheet.weight = request.POST["weight"]
            if is_casting:
                mold.weight = request.POST["weight"]
        else:
            test_sheet.weight = None
            if is_casting:
                mold.weight = None
    if "weight_units" in request.POST:
        if request.POST["weight_units"]!="":
            test_sheet.weight_units = request.POST["weight_units"]
            if is_casting:
                mold.weight_units = request.POST["weight_units"]
        else:
            test_sheet.weight_units = None
            if is_casting:
                mold.weight_units = None
                
    test_sheet.save()
    if is_casting:
        eq.equipment_mold.save()
        mold.save() 
    eq.save()
    

    if eq.equipment_mold:
        equipment_mold_id = eq.equipment_mold_id
    else:
        equipment_mold_id = eq.id

    # Equipment.objects.filter(equipment_mold_id=equipment_mold_id).update(is_trip_unit=eq.is_trip_unit)
    return HttpResponseRedirect(reverse("equipment", args=(equipment_id, )))


def add_test_results(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})         
    if "add_test_results" in request.FILES:
        eq=Equipment.objects.get(pk=equipment_id)
        eq.test_results=request.FILES["add_test_results"]
        eq.save()

    return HttpResponseRedirect(reverse("equipment", args=(equipment_id, )))

def remove_test_results(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        eq=Equipment.objects.get(pk=equipment_id)
        eq.test_results.delete(save=False)
        eq.save()
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except Equipment.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid File Selection. Contact Admin"})
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def add_nameplate(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    if "add_nameplate" in request.FILES:
        eq=Equipment.objects.get(pk=equipment_id)
        eq.nameplate=request.FILES["add_nameplate"]
        eq.save()
        

        equipment_mold_id = eq.equipment_mold_id or eq.id
        Equipment.objects.filter(Q(id=equipment_mold_id) | Q(equipment_mold_id=equipment_mold_id)).update(nameplate=eq.nameplate)
    return HttpResponseRedirect(reverse("equipment", args=(equipment_id, )))   

def remove_nameplate(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        equipment = Equipment.objects.get(pk=equipment_id)
        equipment_mold_id = equipment.equipment_mold_id or equipment.id
        equipments = Equipment.objects.filter(Q(id=equipment_mold_id) | Q(equipment_mold_id=equipment_mold_id))
        for eq in equipments:
            eq.nameplate.delete(save=False)
            eq.save()
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except Equipment.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid File Selection. Contact Admin"})
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def add_new_equipment_note(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    if "add_note_pk" in request.POST:
        if request.POST["add_note_pk"] and request.POST["add_note_pk"].strip():
            eq_pk = request.POST["add_note_pk"]
            equipment=Equipment.objects.get(pk=eq_pk)
            if "add_new_note" in request.POST:
                if request.POST["add_new_note"] and request.POST["add_new_note"].strip():
                    eq_note_posted = request.POST["add_new_note"]
                    eq_note = EquipmentNotes(note=eq_note_posted, equipment=equipment)
                    eq_note.author = request.user
                    eq_note.save()
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    return render(request, "jobs/error.html", {"message": "Error! Note not added. Must Contain letters, numbers, or symbols"})

def add_equipment_note(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    equipment=Equipment.objects.get(pk=equipment_id)    
    if "add_note" in request.POST:
        if request.POST["add_note"] and request.POST["add_note"].strip():
            eq_note_posted = request.POST["add_note"]
            eq_note = EquipmentNotes(note=eq_note_posted, equipment=equipment)
            eq_note.author = request.user
            eq_note.save()
            #add to notifications
        # try:
            user_properties = UserProperties.objects.get(user=request.user)
            # user_properties.equipment_notifications.add(new_equipment)
            # user_properties.save()
            users_to_notify = equipment.equipment_notifications_user.all()
            note_html = note_notification(eq_note)
            for user_i in users_to_notify:
                t1 = threading.Thread(target=send_notification, args=[user_i.user.email,("Paragon Equipment Note: \""+eq_note.truncated), note_html],daemon=True)
                t1.start()
        # except:
        #     print("failed notification")
            return HttpResponseRedirect(reverse("equipment", args=(equipment_id, )))
    return render(request, "jobs/error.html", {"message": "Error! Note not added. Must Contain letters, numbers, or symbols"})


def remove_model_files(request, file_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        model_file = ModelFolder.objects.get(pk=file_id)
        model_id=model_file.model.id
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except EquipmentFolder.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid File Selection. Contact Admin"})
    model_file.model_file.delete(save=False) #delete file in S3
    model_file.delete() #delete file in django
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


def generate_site_id_copy_name(name, existing_site_ids, copy_number=1):
    # if not request.user.is_authenticated:
    #     return render(request, "jobs/login.html", {"message": None})
    if copy_number == 1:
        new_site_id = f'{name} copy'
    else:
        new_site_id = f'{name} copy {copy_number}'

    if new_site_id in existing_site_ids:
        return generate_site_id_copy_name(name, existing_site_ids, copy_number + 1)
    return new_site_id


def equipment_batch_actions(request, job_id):
    # if not request.user.is_authenticated:
    #     return render(request, "jobs/login.html", {"message": None})
    today = datetime.now()
    now=today.strftime("%b-%d-%Y %H:%M")
    batch_actions = request.POST["batchActionsData"]
    if batch_actions:
        batch=batch_actions.split(',')
        batch_ids = batch[1:]
        batch_ids = list(map(int, batch_ids))
        batch_withsub_duplicates = batch_ids
        equipments = Equipment.objects.filter(id__in=batch_ids)
        for eq in equipments:
            batch_withsub_duplicates.extend(eq.sub_equipment_ids)
        batch_wsubs = list(set(batch_withsub_duplicates))
        
        equipments_wsubs = Equipment.objects.filter(id__in=batch_wsubs)
       
        if batch[0] == 'trash':
            equipments_wsubs.update(trashed=True)
            for eq in equipments_wsubs:
                castings = eq.casting.all()
                if castings.exists():
                    castings.update(trashed=True)
                #the commented out chunk below was originall added to remove the relationship with the job, but we instead added an extra filter in the view for 
                #trashed equipment that filters out any equipment whos mold is trashed.
                # if eq.job_site:
                #     #get all casts
                #     casts = Equipment.objects.filter(equipment_mold = eq)
                #     for cast in casts:
                #         job=Job.objects.get(equipment = cast)
                #         job.equipment.remove(cast)
                eq.site_id = eq.site_id + " [TRASHED on " + now + "]"
                eq.save()
        elif batch[0] == 'complete':
            equipments.update(completion=True)
        elif batch[0] == 'uncomplete':
            equipments.update(completion=False)
        elif batch[0] == 'site_add':
            job_pk = request.POST["selected_job"]
            this_job=Job.objects.get(pk=job_pk)
            #clean will be changed to false if illegal behaviour happens. This will be a redundant check and error page will
            #hopefully never be needed since the front end should prevent submission before adding it
            clean=True
            er_message="Contact Admin. Error Code[equipment-add-error-unknown]"

            for mold in equipments:
                if this_job.equipment.filter(equipment_mold = mold).exists():
                    clean=False   
                    er_message="Failed to add the equipment. Contact Admin. Error Code: [duplicate-mold-error]"
                #check whether eq has a valid parent, either 
                elif mold.parent_equipment:
                    #equipment cannot be added unless it's parent is also in the job or it's parent is also being added to the job.                     
                    if not this_job.equipment.filter(equipment_mold = mold.parent_equipment).exists():
                        if not mold.parent_equipment in equipments:
                            clean=False
                            er_message="Failed to add the equipment. Contact Admin. Error Code: [ancestor-missing]"                          
            if clean:
                child_castings = []
                index = 0
                childindex=0
                for mold in equipments:
                    index += 1
                    mold_pk = mold.pk
                    casting = mold
                    mte = mold.mandatory_test_equipment.all()
                    ote = mold.optional_test_equipment.all()
                    casting.pk=None
                    casting.save()    
                    casting.job_site=None
                    mold=Equipment.objects.get(pk=mold_pk)
                    casting.equipment_mold = mold  
                    casting.save()   
                    casting.mandatory_test_equipment.add(*mte)    
                    casting.mandatory_test_equipment.add(*ote)   
                    this_job.equipment.add(casting)
                    this_job.save()
                    if mold.equipment_type.is_test_sheet: #add a test sheet to the new equipment only if the type has a test sheet
                        #get ts from parent
                        ts = mold.sheet_eq
                        oldts = mold.sheet_eq
                        ts.pk = None
                        ts.eq=casting
                        ts.save()

                    if mold.parent_equipment:
                        child_castings.append(casting)
                for child_casting in child_castings:
                    childindex+=1
                    new_parent = this_job.equipment.get(equipment_mold = child_casting.equipment_mold.parent_equipment )
                    child_casting.parent_equipment = new_parent
                    child_casting.save()                
            
            else:
                return render(request, "jobs/error.html", {"message": er_message})
            return HttpResponseRedirect(reverse("job", args=(job_pk, ))+"#equipment")

        elif batch[0] == 'copy':
            copy_count = int(request.POST.get('copy_count') or 1)
            for i in range(copy_count):
                equipments = equipments.order_by('id')
                job_site = equipments.first().job_site
                existing_site_ids = list(job_site.equipment_jobsite.all().values_list('site_id', flat=True))

                parent_id_changes = {}
                for equipment in equipments:
                    equipment_ts = pydash.get(equipment, 'sheet_eq')
                    mte = equipment.mandatory_test_equipment.all()
                    ote = equipment.optional_test_equipment.all()
                    current_id = None
                    if equipment.sub_equipment_ids:
                        current_id = equipment.id
                    equipment.id = None
                    equipment.pk = None

                    site_id = equipment.site_id
                    copy_number = 1
                    if re.match(r".+ copy [0-9]*$", site_id):
                        copy_number += int(site_id.split(' ')[-1])
                        site_id = site_id.split(' copy')[0]
                    elif re.match(r".+ copy$", site_id):
                        copy_number += 1
                        site_id = site_id.split(' copy')[0]
                    equipment.site_id = generate_site_id_copy_name(site_id, existing_site_ids, copy_number)
                    if equipment.parent_equipment:
                        equipment.parent_equipment_id = parent_id_changes.get(f'{equipment.parent_equipment_id}')

                    if job_site.equipment_jobsite.filter(site_id__exact=equipment.site_id).exists():
                        return render(request, "jobs/error.html", {"message": "Found duplicate site ids"})

                    existing_site_ids.append(equipment.site_id)
                    equipment.scope = ''
                    equipment.completion = False
                    equipment.test_results = None
                    equipment.nameplate = None
                    equipment.serial_number = None
                    equipment.save()

                    if equipment_ts:
                        new_test_sheet = (
                            TestSheet.objects.create(
                                eq=equipment,
                                eq_type=equipment.equipment_type,
                                eq_model=equipment.equipment_model,
                                control_voltage=equipment_ts.control_voltage,
                                is_dc_control_voltage=equipment_ts.is_dc_control_voltage,
                                trip_coil_voltage=equipment_ts.trip_coil_voltage,
                                is_dc_trip_voltage=equipment_ts.is_dc_trip_voltage,
                                fuse_manufacturer=equipment_ts.fuse_manufacturer,
                                fuse_type=equipment_ts.fuse_type,
                                fuse_size=equipment_ts.fuse_size,
                                frame_size=equipment_ts.frame_size,
                                mount_style=equipment_ts.mount_style,
                                trip_unit_model=equipment_ts.trip_unit_model,
                                trip_unit_manufacturer=equipment_ts.trip_unit_manufacturer,
                                trip_unit_serial_number=equipment_ts.trip_unit_serial_number,
                                trip_unit_rating_plug=equipment_ts.trip_unit_rating_plug,
                                trip_unit_curve=equipment_ts.trip_unit_curve,
                                trip_unit_phase_ct_high=equipment_ts.trip_unit_phase_ct_high,
                                trip_unit_phase_ct_low=equipment_ts.trip_unit_phase_ct_low,
                                is_primary_injection=equipment_ts.trip_unit_phase_ct_low,
                                protection_device=equipment_ts.protection_device,
                                trip_device=equipment_ts.trip_device,
                                power_rating=equipment_ts.power_rating,
                                power_rating_units=equipment_ts.power_rating_units,
                                primary_winding_config=equipment_ts.primary_winding_config,
                                primary_voltage=equipment_ts.primary_voltage,
                                secondary_voltage=equipment_ts.secondary_voltage,
                                temp_rise=equipment_ts.temp_rise,
                                temp_rise_units=equipment_ts.temp_rise_units,
                                impedance=equipment_ts.impedance,
                                impedance_at=equipment_ts.impedance_at,
                                tap_qty=equipment_ts.tap_qty,
                                tap_position=equipment_ts.tap_position,
                                tap_one_volts=equipment_ts.tap_one_volts,
                                ttr_upper_tolerance=equipment_ts.ttr_upper_tolerance,
                                ttr_lower_tolerance=equipment_ts.ttr_lower_tolerance,
                                tap_two_volts=equipment_ts.tap_two_volts,
                                tap_three_volts=equipment_ts.tap_three_volts,
                                tap_four_volts=equipment_ts.tap_four_volts,
                                tap_five_volts=equipment_ts.tap_five_volts,
                                tap_six_volts=equipment_ts.tap_six_volts,
                                tap_seven_volts=equipment_ts.tap_seven_volts,
                                is_oil_sample_required=equipment_ts.is_oil_sample_required,
                                fluid_type=equipment_ts.fluid_type,
                                fluid_capacity=equipment_ts.fluid_capacity,
                                fluid_capacity_units=equipment_ts.fluid_capacity_units,
                                cable_voltage_rating=equipment_ts.cable_voltage_rating,
                                operating_cable_voltage=equipment_ts.operating_cable_voltage,
                                cable_amp_rating=equipment_ts.cable_amp_rating,
                                cable_insulation_type=equipment_ts.cable_insulation_type,
                                cable_insulation_thickness=equipment_ts.cable_insulation_thickness,
                                cable_size=equipment_ts.cable_size,
                                cable_conductor_material=equipment_ts.cable_conductor_material,
                                is_cable_shielded=equipment_ts.is_cable_shielded,
                            )
                        )

                    if current_id:
                        parent_id_changes[f'{current_id}'] = equipment.id

                    equipment.mandatory_test_equipment.add(*mte)
                    equipment.optional_test_equipment.add(*ote)
        else:
            #this is a move function
            jobs = None
            job_site = JobSite.objects.get(pk = job_id)
            jobs = Job.objects.filter(job_site=job_site)
            
            if int(batch[0])>0:
                new_parent = Equipment.objects.get(id = batch[0])

                if int(batch[0]) in batch_ids:
                    eq_names = list(Equipment.objects.filter(id__in=batch_ids).values_list('site_id', flat=True))
                    return render(
                        request,
                        "jobs/error.html",
                        {
                            "message": f"{', '.join(eq_names)} equipment(s) are not valid to be moved under {new_parent.site_id}. The sub equipment would be left without a parent."
                        }
                    )

                if new_parent.equipment_mold:
                    site_eq = False
                else:
                    site_eq = True
            else:
                new_parent = None

            #loop through selection and remove anything except the highest ancestor
            x=0

            for eq in equipments:
                new_child = get_ancestor(eq, batch_ids)
                # print(new_parent)
                if new_child:
                    new_child.parent_equipment = new_parent
                    for job in jobs:              
                        if new_parent:
                            print(new_parent)
                            job_parent_eq = job.equipment.filter(equipment_mold=new_parent).first()
                        elif new_parent == None:
                            print("a;ldskfj")
                            job_parent_eq = None
                        job.equipment.filter(equipment_mold=new_child).update(parent_equipment=job_parent_eq)
                    
                    new_child.save()
            
     

    headers = request.META
    origin = headers.get('HTTP_ORIGIN')
    referer = headers.get('HTTP_REFERER')
    path = referer.replace(origin, '').split('/')[1]
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    # return HttpResponseRedirect(reverse(path, args=(job_id, ))+"#equipment")




#function takes a target equipment and list of ids and returns the Top level ancestor OF THE SELECTED. 
#i.e. if you have 5 levels of depth and you select levels 3, 4, and 5, this will return '3'. It doesn't return the highest level '1' because that wasn't in the selected. 
def get_ancestor(eq, batch):
    # if not request.user.is_authenticated:
    #     return render(request, "jobs/login.html", {"message": None})
    if eq.parent_equipment:
        if eq.parent_equipment.pk in batch:
            get_ancestor(eq.parent_equipment, batch)
        else:
            return eq
    else:
        return eq
        

def edit_site_id(request, job_id= None):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    batch_actions = request.POST["batchSiteIdData"]
    if batch_actions:
        batch=batch_actions.split(',')
        
        site = None  
        job=None
        try:
            job=Job.objects.get(pk=job_id)
        except:
            job = None
        if job:
            site = job.job_site
        else:
            site = Equipment.objects.get(pk=batch[0]).job_site #grab the jobsite from the first equipment 
        id_checks = Equipment.objects.filter(job_site = site).values_list('site_id', flat = True)
        failed_equipments=[]
        passed_equipments=[]
        i=0
        while i <len(batch):
            equipment=Equipment.objects.get(pk=batch[i]) #get equipment using primary key
            #check 
            if (batch[i+1] in id_checks):#checks to see if equipment name already exists
                if not equipment.site_id == batch[i+1]:
                    if Equipment.objects.filter(site_id = equipment.site_id).exists(): #checks to see if you changed the name
                        failed_equipments.append(equipment.site_id)#adds to a list of equipments to be passed out to the user on completion
                    
            else:
                equipment.site_id=batch[i+1] #add the new name
                equipment.save() #save the changes
                if job:
                    equipment.equipment_mold.site_id = batch[i+1]
                    equipment.equipment_mold.save()
                else:
                    Equipment.objects.filter(equipment_mold = equipment).update(site_id = batch[i+1])
                
                passed_equipments.append(equipment.site_id)
                
            i+=2 #increment to next equipment
        if len(failed_equipments)>0 and len(passed_equipments)>0:
            #renders error page if any equipment failed. Successful changes did save
            return render(request, "jobs/warning.html", {"message": "Some site ID's were not changed because they are already in use. The Site ID's that were not changed are: "+" ; ".join(failed_equipments)+". These Site ID's were set succesfully: "+" ; ".join(passed_equipments)+"." })
        elif len(failed_equipments)>0:
            return render(request, "jobs/warning.html", {"message": "The selected site ID(s) already exist for this jobsite: "+site.name+". Each site ID must be unique on the site" })
        elif job:
            return HttpResponseRedirect(reverse("job", args=(job_id, ))+"#equipment") #returns to the equipment tab
        else:
            return HttpResponseRedirect(reverse("job_site", args=(site.id, ))+"#equipment") #returns to the equipment tab
        
    else:
        return render(request, "jobs/error.html", {"message": "Contact Admin. Error Code: batchSiteIdData-no-back-actions "})

def create_type(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})

    #extract type name from form
    type_name = request.POST["type_name"].strip()

    if Type.objects.filter(name=type_name).exists():
        return render(request, "jobs/error.html", {"message": "Already Exists! No need to create this type. If the details of the existing one are incorrect, select Types from side menu to open and edit there. "})
    
    #create new type object
    new_type=Type.objects.create_type(type_name)
    #fill in all properties that were submitted in the form
    #Need to try these things, but any failure needs to delete the object
    # try:
    if 'test_sheet' in request.FILES:
        new_type.test_sheet = request.FILES["test_sheet"]
    if 'type_notes' in request.POST:
        new_type.type_notes = request.POST["type_notes"]
    user_properties=UserProperties.objects.get(user=request.user)
    if "is_private" in request.POST:
            new_type.is_private=True
            try:
                company=user_properties.company
                new_type.company = company
            except:
                company=None        
    mandatory_equipment_list=request.POST["mandatory_test_equipment"]
    optional_equipment_list=request.POST["optional_test_equipment"]
    if mandatory_equipment_list:
        mandatory_equipment_list=mandatory_equipment_list.split(',')
    else:
        mandatory_equipment_list=[]
    if optional_equipment_list:
        optional_equipment_list=optional_equipment_list.split(',')
    else:
        optional_equipment_list=[]
    #if mandatory equipment was added, this adds them to the type
    if mandatory_equipment_list:
        for equipment_name in mandatory_equipment_list:
            
            try:
                equipment=TestEquipment.objects.get(name=equipment_name)
                new_type.mandatory_type_test_equipment.add(equipment.pk)
            except TestEquipment.DoesNotExist:
                return HttpResponseRedirect(reverse("jobs"))
    #if optional equipment was added, this adds them to the type
    if optional_equipment_list:
        for equipment_name in optional_equipment_list:
            try:
                equipment=TestEquipment.objects.get(name=equipment_name)
                new_type.optional_type_test_equipment.add(equipment.pk)
            except TestEquipment.DoesNotExist:
                return HttpResponseRedirect(reverse("jobs"))

    new_type.save()

    return HttpResponseRedirect(reverse("eq_type", args=(new_type.pk, )))


def edit_type(request, type_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    type = Type.objects.filter(id=type_id).first()
    if not type:
        return render(request, "jobs/error.html", {"message": "Type not found"})

    # copy_type = deepcopy(type)
    has_changed = False

    current_ote_ids = []
    current_mte_ids = []

    ts_ids = None
    if "ts_ids" in request.POST:
        has_changed = True
        ts_ids = request.POST["ts_ids"].split('-')

    if "mandatory_type_test_equipment" in request.POST:
        mandatory_test_equipments = request.POST['mandatory_type_test_equipment'].split(',')
        mandatory_test_equipments = TestEquipment.objects.filter(name__in=mandatory_test_equipments).order_by('id')
        mandatory_test_equipments_ids = mandatory_test_equipments.values_list('id', flat=True)
        current_mte_ids = list(type.mandatory_type_test_equipment.values_list('id', flat=True).order_by('id'))
        if current_mte_ids != list(mandatory_test_equipments_ids):
            has_changed = True

            type.mandatory_type_test_equipment.clear()
            type.mandatory_type_test_equipment.add(*mandatory_test_equipments)

    if "optional_type_test_equipment" in request.POST:
        optional_model_test_equipments = request.POST['optional_type_test_equipment'].split(',')
        optional_model_test_equipments = TestEquipment.objects.filter(name__in=optional_model_test_equipments).order_by('id')
        optional_model_test_equipments_ids = optional_model_test_equipments.values_list('id', flat=True)
        current_ote_ids = list(type.optional_type_test_equipment.values_list('id', flat=True).order_by('id'))
        if current_ote_ids != list(optional_model_test_equipments_ids):
            has_changed = True

            type.optional_type_test_equipment.clear()
            type.optional_type_test_equipment.add(*optional_model_test_equipments)

    if "name" in request.POST:
        if type.name != request.POST['name'].strip():
            has_changed = True
            type.name = request.POST['name'].strip()
    if "ts_name" in request.POST:
        if type.name != request.POST['ts_name'].strip():
            has_changed = True
            type.name = request.POST['name'].strip()
    if "name" in request.POST:
        if type.name != request.POST['name'].strip():
            has_changed = True
            type.name = request.POST['name'].strip()
    if "name" in request.POST:
        if type.name != request.POST['name'].strip():
            has_changed = True
            type.name = request.POST['name'].strip()
    if "is_insulation_resistance" in request.POST:
        if type.is_insulation_resistance != request.POST['is_insulation_resistance']:
            has_changed = True
            type.is_insulation_resistance = request.POST['is_insulation_resistance'] == 'True'
    if "is_contact_resistance" in request.POST:
        if type.is_contact_resistance != request.POST['is_contact_resistance']:
            has_changed = True
            type.is_contact_resistance = request.POST['is_contact_resistance'] == 'True'
    if "is_trip_unit" in request.POST:
        if type.is_trip_unit != request.POST['is_trip_unit']:
            has_changed = True
            type.is_trip_unit = request.POST['is_trip_unit'] == 'True'
    if "is_primary_injection" in request.POST:
        if type.is_primary_injection != request.POST['is_primary_injection']:
            has_changed = True
            type.is_primary_injection = request.POST['is_primary_injection'] == 'True'
    if "is_secondary_injection" in request.POST:
        if type.is_secondary_injection != request.POST['is_secondary_injection']:
            has_changed = True
            type.is_secondary_injection = request.POST['is_secondary_injection'] == 'True'
    if "is_power_fused" in request.POST:
        if type.is_power_fused != request.POST['is_power_fused']:
            has_changed = True
            type.is_power_fused = request.POST['is_power_fused'] == 'True'
    if "is_breaker" in request.POST:
        if type.is_breaker != request.POST['is_breaker']:
            has_changed = True
            type.is_breaker = request.POST['is_breaker'] == 'True'
    if "is_hipot" in request.POST:
        if type.is_hipot != request.POST['is_hipot']:
            has_changed = True
            type.is_hipot = request.POST['is_hipot'] == 'True'
    if "is_inspection" in request.POST:
        if type.is_inspection != request.POST['is_inspection']:
            has_changed = True
            type.is_inspection = request.POST['is_inspection'] == 'True'
    if "is_transformer" in request.POST:
        if type.is_transformer != request.POST['is_transformer']:
            has_changed = True
            type.is_transformer = request.POST['is_transformer'] == 'True'
    if "is_winding_resistance" in request.POST:
        if type.is_winding_resistance != request.POST['is_winding_resistance']:
            has_changed = True
            type.is_winding_resistance = request.POST['is_winding_resistance'] == 'True'
    if "is_liquid_type" in request.POST:
        if type.is_liquid_type != request.POST['is_liquid_type']:
            has_changed = True
            type.is_liquid_type = request.POST['is_liquid_type'] == 'True'
    if "is_cable" in request.POST:
        if type.is_cable != request.POST['is_cable']:
            has_changed = True
            type.is_cable = request.POST['is_cable'] == 'True'
    if "is_cable_vlf_withstand_test" in request.POST:
        if type.is_cable_vlf_withstand_test != request.POST['is_cable_vlf_withstand_test']:
            has_changed = True
            type.is_cable_vlf_withstand_test = request.POST['is_cable_vlf_withstand_test'] == 'True'
    if "is_ttr" in request.POST:
        if type.is_ttr != request.POST['is_ttr']:
            has_changed = True
            type.is_ttr = request.POST['is_ttr'] == 'True'
    if "is_xfmr_insulation_resistance" in request.POST:
        if type.is_xfmr_insulation_resistance != request.POST['is_xfmr_insulation_resistance']:
            has_changed = True
            type.is_xfmr_insulation_resistance = request.POST['is_xfmr_insulation_resistance'] == 'True'

    # if has_changed:
    #     copy_type.pk = None
    #     copy_type.id = None
    #     copy_type.status = STATUS_PENDING
    #     copy_type.save()
    #     if current_mte_ids:
    #         mte = TestEquipment.objects.filter(id__in=current_mte_ids)
    #         copy_type.mandatory_type_test_equipment.add(*mte)
    #     if current_ote_ids:
    #         ote = TestEquipment.objects.filter(id__in=current_ote_ids)
    #         copy_type.optional_type_test_equipment.add(*ote)

    #     for guide in TypeTestGuide.objects.filter(eq_type=type):
    #         guide.id = None
    #         guide.pk = None
    #         guide.eq_type = copy_type
    #         guide.save()

    #     for model in Model.objects.filter(model_type=type).order_by("id"):
    #         model.id = None
    #         model.pk = None
    #         model.model_type = copy_type
    #         model.save()

    #     for folder in TypeFolder.objects.filter(eq_type=type).order_by('id'):
    #         folder.id = None
    #         folder.pk = None
    #         folder.eq_type = copy_type
    #         folder.save()

    #     parent_id_changes = {}
    #     for note in type.note_model.order_by('created_at'):
    #         current_id = None
    #         if note.sub_notes.exists():
    #             current_id = note.id
    #         note.id = None
    #         note.pk = None
    #         note.eq_type = copy_type
    #         note.save()

    #         if current_id:
    #             parent_id_changes[f'{current_id}'] = note.id

    #         if note.parent_note:
    #             note.parent_note_id = parent_id_changes[f'{note.parent_note_id}']
    #             note.save()


    #     test_type_standards = TypeTestStandards.objects.filter(ts_type=type)
    #     for ts in test_type_standards:
    #         ts.id = None
    #         ts.pk = None
    #         ts.ts_type = copy_type
    #         ts.save()

    #     if ts_ids:
    #         test_type_standards = TypeTestStandards.objects.filter(id__in=ts_ids)
    #         for ts in test_type_standards:
    #             if f"ts_name-{ts.id}" in request.POST:
    #                 ts.ts_name = request.POST[f"ts_name-{ts.id}"]

    #             if f"ts_standard-{ts.id}" in request.POST:
    #                 ts.ts_standard = request.POST[f"ts_standard-{ts.id}"]

    #             if f"ts_description-{ts.id}" in request.POST:
    #                 ts.ts_description = request.POST[f"ts_description-{ts.id}"]

    #             ts.save()

    type.save()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))



def copy_private_type(request, type_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    #get type id from form and then pull up the type object
    try:
        eq_type = Type.objects.get(pk=type_id)
    except KeyError:
        return render(request, "jobs/error.html", {"message": "Error code: copy_ptype_KEYerror"})
    except Type.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "copy_ptype_DNEerror"})

    properties=UserProperties.objects.get(user=request.user)
    type_copy = eq_type
    type_copy.pk = None
    type_copy.save()
    type_copy.is_private = True
    type_copy.company = properties.company
    type_copy.save()
    return HttpResponseRedirect(reverse("eq_type", args=(type_copy.id, )))

def create_test_eq(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})

    #extract type name from form
    te_name = request.POST["test_eq_name"]
    test_equipment=TestEquipment(name=te_name)
    if "notes" in request.POST:
        test_equipment.notes=request.POST["notes"]

    test_equipment.save()

    return HttpResponseRedirect(reverse("test_equipment"))

def edit_test_eq(request, te_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    te = TestEquipment.objects.filter(id=te_id).first()
    if not te:
        return render(request, "jobs/error.html", {"message": "Test set not found"})

    if request.POST.get("name"):
        te.name = request.POST["name"].strip()

    if request.POST.get("notes"):
        te.notes = request.POST["notes"].strip()

    te.save()
    return HttpResponseRedirect(reverse("test_equipment"))

def delete_test_eq(request, te_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    te = TestEquipment.objects.filter(id=te_id)
    if not te:
        return render(request, "jobs/error.html", {"message": "Test set not found"})

    te.delete()
    return HttpResponseRedirect(reverse("test_equipment"))

def create_manufacturer(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})

    #extract type name from form
    manufacturer_name = request.POST["manufacturer_name"].strip()
    if Manufacturer.objects.filter(name=manufacturer_name).exists():
        return render(request, "jobs/error.html", {"message": "Already Exists! No need to create this manufacturer. If the details of the existing one are incorrect, select Manufacturers and edit there. "})
    #create new type object
    new_manufacturer=Manufacturer.objects.create_manufacturer(manufacturer_name)
    new_manufacturer.customer_support = request.POST["customer_support"]
    #save to database
    new_manufacturer.save()
    return HttpResponseRedirect(reverse("eq_manufacturer", args=(new_manufacturer.pk, )))

def create_model(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})

    #extract model name from form
    req_fields = ["model_name", "model_type", "manufacturer"]
    if all(elem in request.POST  for elem in req_fields):
        model_name = request.POST["model_name"].strip()   
        print(Model.objects.filter(name=model_name).count())         
        if Model.objects.filter(name=model_name).exists():
            return render(request, "jobs/error.html", {"message": "Already Exists! No need to create this model. If the details of the existing one are incorrect, select Models and edit there. "})
        manufacturer_name = request.POST["manufacturer"].strip()
        manufacturer=Manufacturer.objects.get(name=manufacturer_name)
        eq_type_name = request.POST["model_type"].strip()
        eq_type = Type.objects.get(name=eq_type_name)
        #create new model object
        new_model=Model.objects.create_model(model_name)
        new_model.model_type = eq_type
        new_model.model_manufacturer = manufacturer
        new_model.save()
        #fill in all properties that were submitted in the form
        user_properties=UserProperties.objects.get(user=request.user)
        if "is_private" in request.POST:
            new_model.is_private=True
            try:
                company=user_properties.company
                new_model.company = company
            except:
                company=None        
        if "model_test_sheet" in request.FILES:
            new_model.model_test_sheet = request.FILES["model_test_sheet"]
        if "model_manual" in request.FILES:
            new_model.model_manual = request.FILES["model_manual"]
        mandatory_equipment_list=request.POST["mandatory_test_equipment"]
        if mandatory_equipment_list:
            mandatory_equipment_list=mandatory_equipment_list.split(',')
        else:
            mandatory_equipment_list=[]
        optional_equipment_list=request.POST["optional_test_equipment"]
        if optional_equipment_list:
            optional_equipment_list=optional_equipment_list.split(',')
        else:
            optional_equipment_list=[]


        #if mandatory equipment was added, this adds them to the model
        if mandatory_equipment_list:
            for equipment_name in mandatory_equipment_list:
                
                
                try:
                    equipment=TestEquipment.objects.get(name=equipment_name)
                    new_model.mandatory_model_test_equipment.add(equipment.pk)
                except TestEquipment.DoesNotExist:
                    new_model.delete()
                    return render(request, "jobs/error.html", {"message": "Error with selected Mandatory Equipment"})

        #if mandatory equipment was added, this adds them to the model
        if optional_equipment_list:
            for equipment_name in optional_equipment_list:
                
                try:
                    equipment=TestEquipment.objects.get(name=equipment_name)
                    new_model.optional_model_test_equipment.add(equipment.pk)
                except TestEquipment.DoesNotExist:
                    new_model.delete()
                    return render(request, "jobs/error.html", {"message": "Error with selected Optional Equipment"})

        new_model.save()
        return HttpResponseRedirect(reverse("eq_model", args=(new_model.pk, )))
    else:
        return render(request, "jobs/error.html", {"message": "Requires Name, Type, and Manufacturer"})


def edit_manufacturer(request, manufacturer_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    man = Manufacturer.objects.filter(id=manufacturer_id).first()
    if not man:
        return render(request, "jobs/error.html", {"message": "Error, Contact Admin. Error Code: Manufacturer not found"})
    has_changed=False
    if "customer_support" in request.POST:
        if man.customer_support != request.POST['customer_support'].strip():
            has_changed = True
            man.customer_support = request.POST['customer_support'].strip()
    if "name" in request.POST:
        if man.name != request.POST['name'].strip():
            has_changed = True
            man.name = request.POST['name'].strip()
    if has_changed:
        man.save()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


#ERRORS WERE ARISING WHEN ADDING EQUIPMENT IF THE EQUIPMENT HAD AN ANCESTOR. THIS ENTIRE FUNCTION HAS BEEN TEMPORARILY SUSPENDED
def add_jobsite_eq(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    job = Job.objects.get(pk=job_id)
    eq_name = request.POST['adding_name']
    if Equipment.objects.filter(equipments = job, site_id = eq_name).exists():
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        
    if Equipment.objects.filter(job_site = job.job_site, site_id = eq_name).exists():
        if Equipment.objects.filter(job_site = job.job_site, site_id = eq_name).count()>1:
            return render(request, "jobs/error.html", {"message": "Error multiple equipment with same ID. Contact Admin"})
        
    parent_equipment = None
    # if "parent_equipment_id" in request.POST:
    #     if request.POST["parent_equipment_id"]!="":
    #         try:
    #             parent_equipment_id = request.POST["parent_equipment_id"]
    #             parent_equipment = Equipment.objects.get(pk=parent_equipment_id)
    #         except:
    #             return render(request, "jobs/error.html", {"message": "Error Parent equipment not loaded. Contact Admin"})

    try:
        mold = Equipment.objects.get(job_site = job.job_site, site_id = eq_name)
        #ERRORS WERE ARISING WHEN ADDING EQUIPMENT IF THE EQUIPMENT HAD AN ANCESTOR. THIS ENTIRE FUNCTION HAS BEEN TEMPORARILY SUSPENDED
        # if mold.parent_equipment:
            #in this case there is a parent equipment so we need to get 
            # try:
            #     parent_equipment = mold.parent
        temp = mold
        temp.pk = None
        eq = temp
        eq.save()
        equipment_type = eq.equipment_type
        if equipment_type.is_test_sheet == True:
            eq.is_insulation_resistance=equipment_type.is_insulation_resistance
            eq.is_contact_resistance=equipment_type.is_contact_resistance
            eq.is_trip_unit=equipment_type.is_trip_unit
            eq.is_primary_injection=equipment_type.is_primary_injection
            eq.is_secondary_injection=equipment_type.is_secondary_injection
            eq.is_xfmr_insulation_resistance=equipment_type.is_xfmr_insulation_resistance
            eq.is_power_fused=equipment_type.is_power_fused
            eq.is_breaker=equipment_type.is_breaker
            eq.is_hipot=equipment_type.is_hipot
            eq.is_inspection=equipment_type.is_inspection
            eq.is_transformer=equipment_type.is_transformer
            eq.is_winding_resistance=equipment_type.is_winding_resistance
            eq.is_ttr=equipment_type.is_ttr
            eq.is_liquid_type=equipment_type.is_liquid_type
            eq.is_cable=equipment_type.is_cable
            eq.is_cable_vlf_withstand_test=equipment_type.is_cable_vlf_withstand_test
            eq.save()
        try:
            blank_test_sheet=TestSheet(eq=eq)
            blank_test_sheet.save()
        except:
            blank_test_sheet = None
            equipment_type.is_test_sheet = False
            equipment_type.save()
            return render(request, "jobs/error.html", {"message": "Error generating new test sheet. 'is_test_sheet' set to 'False' for eq type."})


        eq.equipment_mold = mold
        eq.save()
        job.equipment.add(eq)
        job.save()
        if parent_equipment:
            eq.parent_equipment = parent_equipment
            eq.save()

        eq.save()

    except:
        return render(request, "jobs/error.html", {"message": "get eq failed. contact admin"})
        
    
    return HttpResponseRedirect(reverse("job", args=(job_id, )))
        


def edit_model(request, model_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    model = Model.objects.filter(id=model_id).first()
    if not model:
        return render(request, "jobs/error.html", {"message": "Model not found"})

    # copy_model = deepcopy(model)
    has_changed = False

    current_ote_ids = []
    current_mte_ids = []
    if "mandatory_model_test_equipment" in request.POST:
        mandatory_test_equipments = request.POST['mandatory_model_test_equipment'].split(',')
        mandatory_test_equipments = TestEquipment.objects.filter(name__in=mandatory_test_equipments).order_by('id')
        mandatory_test_equipments_ids = mandatory_test_equipments.values_list('id', flat=True)
        current_mte_ids = list(model.mandatory_model_test_equipment.values_list('id', flat=True).order_by('id'))
        if current_mte_ids != list(mandatory_test_equipments_ids):
            has_changed = True

            model.mandatory_model_test_equipment.clear()
            model.mandatory_model_test_equipment.add(*mandatory_test_equipments)

    if "optional_model_test_equipment" in request.POST:
        optional_model_test_equipments = request.POST['optional_model_test_equipment'].split(',')
        optional_model_test_equipments = TestEquipment.objects.filter(name__in=optional_model_test_equipments).order_by('id')
        optional_model_test_equipments_ids = optional_model_test_equipments.values_list('id', flat=True)
        current_ote_ids = list(model.optional_model_test_equipment.values_list('id', flat=True).order_by('id'))
        if current_ote_ids != list(optional_model_test_equipments_ids):
            has_changed = True

            model.optional_model_test_equipment.clear()
            model.optional_model_test_equipment.add(*optional_model_test_equipments)

    if "name" in request.POST:
        if model.name != request.POST['name'].strip():
            has_changed = True
            model.name = request.POST['name'].strip()
    if "model_id" in request.POST:
        if model.model_id != request.POST['model_id'].strip():
            has_changed = True
            model.model_id = request.POST['model_id'].strip()
    if "model_customer_support" in request.POST:
        if model.model_customer_support != request.POST['model_customer_support'].strip():
            has_changed = True
            model.model_customer_support = request.POST['model_customer_support'].strip()
    if "model_type" in request.POST:
        if model.model_type.id != int(request.POST['model_type'].strip()):
            has_changed = True

            model_type = Type.objects.filter(id=request.POST['model_type']).first()
            if not model_type:
                return render(request, "jobs/error.html", {"message": "Model type not found"})
            model.model_type = model_type

    if "model_manufacturer" in request.POST:
        if model.model_manufacturer.id != int(request.POST['model_manufacturer'].strip()):
            has_changed = True
            model_manufacturer = Manufacturer.objects.filter(id=request.POST['model_manufacturer']).first()
            if not model_manufacturer:
                return render(request, "jobs/error.html", {"message": "Model manufacturer not found"})
            model.model_manufacturer = model_manufacturer
    if "model_manual" in request.FILES:
        model_manual = request.FILES["model_manual"]
        has_changed=True
        if model.model_manual:
            model.model_manual.delete(save=False)   
        model.model_manual = model_manual

    # if has_changed:
    #     copy_model.pk = None
    #     copy_model.id = None
    #     copy_model.status = STATUS_PENDING
    #     copy_model.save()
    #     if current_mte_ids:
    #         mte = TestEquipment.objects.filter(id__in=current_mte_ids)
    #         copy_model.mandatory_model_test_equipment.add(*mte)
    #     if current_ote_ids:
    #         ote = TestEquipment.objects.filter(id__in=current_ote_ids)
    #         copy_model.optional_model_test_equipment.add(*ote)

    #     for model_folder in ModelFolder.objects.filter(model=model).order_by('id'):
    #         model_folder.id = None
    #         model_folder.pk = None
    #         model_folder.model = copy_model
    #         model_folder.save()

    #     for model_test_guide in ModelTestGuide.objects.filter(model=model).order_by('id'):
    #         model_test_guide.id = None
    #         model_test_guide.pk = None
    #         model_test_guide.model = copy_model
    #         model_test_guide.save()

    #     parent_id_changes = {}
    #     for note in model.note_model.order_by('created_at'):
    #         current_id = None
    #         if note.sub_notes.exists():
    #             current_id = note.id
    #         note.id = None
    #         note.pk = None
    #         note.model = copy_model
    #         note.save()

    #         if current_id:
    #             parent_id_changes[f'{current_id}'] = note.id

    #         if note.parent_note:
    #             note.parent_note_id = parent_id_changes[f'{note.parent_note_id}']
    #             note.save()

    model.save()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def copy_private_model(request, model_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    #get model id from form and then pull up the model object
    try:
        eq_model = Model.objects.get(pk=model_id)
    except KeyError:
        return render(request, "jobs/error.html", {"message": "Error code: copy_pmodel_KEYerror"})
    except Model.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "copy_pmodel_DNEerror"})

    properties=UserProperties.objects.get(user=request.user)
    model_copy = eq_model
    model_copy.pk = None
    model_copy.save()
    model_copy.is_private = True
    model_copy.company = properties.company
    model_copy.save()
    return HttpResponseRedirect(reverse("eq_model", args=(model_copy.id, )))

def add_test_standard(request, type_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        eq_type = Type.objects.get(pk=type_id)
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except Type.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid Type Selection."})
    type_ts = TypeTestStandards(ts_type=eq_type)
    filled_out=False
    if 'ts_add' in request.FILES:
        ts_add = request.FILES["ts_add"]
        type_ts.ts_file = ts_add
        filled_out=True
    elif 'ts_url' in request.POST:
        url_add = request.POST["ts_url"]  
        type_ts.ts_url = url_add   
        filled_out=True
    if 'ts_name' in request.POST:
            if request.POST["ts_name"].strip() != "":
                type_ts.ts_name = request.POST["ts_name"]
    if 'ts_standard' in request.POST:
            if request.POST["ts_standard"].strip() != "":
                type_ts.ts_standard = request.POST["ts_standard"]
    if 'ts_description' in request.POST:
            if request.POST["ts_description"].strip() != "":
                type_ts.ts_description = request.POST["ts_description"]
    if filled_out:
        type_ts.created_by = UserProperties.objects.get(user=request.user)
        type_ts.save()
    else:
        type_ts.delete()
        
    return HttpResponseRedirect(reverse("eq_type", args=(type_id, )))

def add_type_files(request, type_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        eq_type = Type.objects.get(pk=type_id)
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except Type.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid Type Selection."})
    type_folder=TypeFolder(eq_type=eq_type)
    filled_out = False #initialization of variable to determine if a file has been added
    if 'file_add' in request.FILES:
        file_add = request.FILES["file_add"]
        type_folder.type_file = file_add
        filled_out=True
    elif 'file_url' in request.POST:
        url_add = request.POST["file_url"]  
        type_folder.file_url = url_add   
        filled_out=True
    if 'file_name' in request.POST:
            if request.POST["file_name"].strip() != "":
                type_folder.file_name = request.POST["file_name"]
    if filled_out:
        type_folder.save()
    else:
        type_folder.delete()
        
    return HttpResponseRedirect(reverse("eq_type", args=(type_id, )))


def remove_type_files(request, file_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        type_file = TypeFolder.objects.get(pk=file_id)
        type_id=type_file.eq_type.id
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except TypeFolder.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid File Selection. Contact Admin"})
    type_file.type_file.delete(save=False) #delete file in S3
    type_file.delete() #delete file in django
    return HttpResponseRedirect(reverse("eq_type", args=(type_id, )))

def add_type_video(request, type_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        eq_type = Type.objects.get(pk=type_id)
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except Type.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid Type Selection."})
    type_test_guide=TypeTestGuide(eq_type=eq_type)
    # filled_out = False #initialization of variable to determine if a file has been added

    if 'video_name' in request.POST:
            if request.POST["video_name"].strip() != "":
                type_test_guide.title = request.POST["video_name"]
    else:
        return render(request, "jobs/error.html", {"message": "No title entered for video. Try again"})
        
    if 'video_add' in request.FILES:
        video_add = request.FILES["video_add"]
        type_test_guide.type_test_guide = video_add
        # filled_out=True
    else:
        return render(request, "jobs/error.html", {"message": "Valid video file not selected."})
    type_test_guide.save()
    return HttpResponseRedirect(reverse("eq_type", args=(type_id, )))

def add_model_files(request, model_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        model = Model.objects.get(pk=model_id)
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except Model.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid Model Selection."})
    model_folder=ModelFolder(model=model)
    filled_out = False
    if 'file_add' in request.FILES:
        file_add = request.FILES["file_add"]
        model_folder.model_file = file_add
        filled_out=True
    if 'file_url' in request.POST:
        url_add = request.POST["file_url"]  
        model_folder.file_url = url_add   
        filled_out=True
    if 'file_name' in request.POST:
            if request.POST["file_name"].strip() != "":
                model_folder.file_name = request.POST["file_name"]
    if filled_out:
        model_folder.save()
    else:
        model_folder.delete()
        
    return HttpResponseRedirect(reverse("eq_model", args=(model_id, )))

def remove_model_files(request, file_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        model_file = ModelFolder.objects.get(pk=file_id)
        model_id=model_file.model.id
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except ModelFolder.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid File Selection. Contact Admin"})
    model_file.model_file.delete(save=False) #delete file in S3
    model_file.delete() #delete file in django
    return HttpResponseRedirect(reverse("eq_model", args=(model_id, )))

def add_model_video(request, model_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        model=Model.objects.get(pk=model_id)
    except KeyError:
        return render(request, "jobs/error.html", {"message": "No Selection"})
    except Model.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Invalid. Contact admin; error code: PC add_model_video Model DNE"})
    video_guide=ModelTestGuide(model=model)

    if 'video_name' in request.POST:
            if request.POST["video_name"].strip() != "":
                video_guide.title = request.POST["video_name"]
    else:
        return render(request, "jobs/error.html", {"message": "No title entered for video. Try again"})
        
    if 'video_add' in request.FILES:
        video_add = request.FILES["video_add"]
        video_guide.model_test_guide = video_add
    else:
        return render(request, "jobs/error.html", {"message": "Valid video file not selected."})
    video_guide.save()

    return HttpResponseRedirect(reverse("eq_model", args=(model_id, )))

def trash_equipment(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:

       equipment = Equipment.objects.get(pk=equipment_id)
       job = Job.objects.get(equipment=equipment)
       job_id=job.id
       equipment.trashed = True
       today = datetime.now()
       now=today.strftime("%b-%d-%Y %H:%M")
       equipment.site_id = equipment.site_id + " [TRASHED on " + now + "]"
       equipment.save()
    except Equipment.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Equipment Removal Error"})
    except Job.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Associated Job not Identified."})

    return HttpResponseRedirect(reverse("job", args=(job_id, ))+'#equipment' )

def untrash_equipment(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        # -*- coding: utf-8 -*-
        equipment = Equipment.objects.get(pk=equipment_id)
        job = Job.objects.get(equipment=equipment)
        job_id = job.id
        equipment.trashed = False
        equipment.site_id = equipment.site_id[:-31]
        equipment.save()

    except Equipment.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Equipment Restore Error"})
    return HttpResponseRedirect(reverse("trashed_equipment", args=(job_id, )))

def restore_equipments(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    body = json.loads(request.body)
    equipment_ids = body.get('equipment_ids')

    job = Job.objects.filter(id=job_id).first()
    if not job:
        return JsonResponse({
            'message': 'Job does not exists',
        }, status=HTTPStatus.BAD_REQUEST)

    equipments = job.equipment.filter(id__in=equipment_ids)
    for equipment in equipments:
        parent_equipment = equipment.parent_equipment
        if parent_equipment and parent_equipment.trashed and str(parent_equipment.id) not in equipment_ids:
            return JsonResponse({
                'message': f'You must also restore the parent equipment: {parent_equipment.site_id}',
            }, status=HTTPStatus.BAD_REQUEST)

        restore_sub_equipments(equipment)

    return JsonResponse({
        'message': 'Successfully restored equipments',
    }, status=HTTPStatus.OK)

def restore_job_site_equipments(request, job_site_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    body = json.loads(request.body)
    equipment_ids = body.get('equipment_ids')

    job_site = JobSite.objects.filter(id=job_site_id).first()
    if not job_site:
        return JsonResponse({
            'message': 'Job site does not exists',
        }, status=HTTPStatus.BAD_REQUEST)

    equipments = job_site.equipment_jobsite.filter(id__in=equipment_ids).order_by('id')
    for equipment in equipments:
        parent_equipment = equipment.parent_equipment
        if parent_equipment and parent_equipment.trashed and parent_equipment.id not in equipment_ids:
            return JsonResponse({
                'message': f'You must also restore the parent equipment: {parent_equipment.site_id}',
            }, status=HTTPStatus.BAD_REQUEST)

        restore_sub_equipments(equipment)

    return JsonResponse({
        'message': 'Successfully restored equipments',
    }, status=HTTPStatus.OK)


def restore_sub_equipments(equipment):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    equipment.trashed = False
    equipment.site_id = equipment.site_id.split(' [TRASHED')[0]
    equipment.save()

    # Just uncomment below codes if in case we want to auto restore sub equipments
    # sub_equipments = equipment.sub_equipments.filter(trashed=True)
    # for sub in sub_equipments:
    #     restore_sub_equipments(sub)


def edit_equipment(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    update_all = {}
    equipment = Equipment.objects.filter(id=equipment_id).first()
    if not equipment:
        return render(
            request,
            "jobs/error.html",
            {"message": "Equipment / Job not found"}
        )

    job_site = equipment.job_site or equipment.equipment_mold.job_site
    equipment_mold_id = equipment.equipment_mold_id or equipment.id

    site_id = request.POST.get("site_id").strip()
    if equipment.site_id != site_id:
        site_id = request.POST["site_id"].strip()
        if job_site.equipment_jobsite.filter(site_id__iexact=site_id).exists():
            return render(
                request,
                "jobs/error.html",
                {"message": "Site ID already exists"}
            )
        update_all['site_id'] = site_id

    # if ' equipment_type' in request.POST:
    #     update_all['equipment_type_id'] = request.POST.get('equipment_type')
    
    if 'selected_model' in request.POST:
        if request.POST.get('selected_model') != "":
            eq_model = Model.objects.get(pk = request.POST.get('selected_model'))            
            update_all['equipment_model'] = eq_model        
            update_all['equipment_type'] = eq_model.model_type
    if 'serial_number' in request.POST:
        update_all['serial_number'] = request.POST.get('serial_number')
    if 'equipment_location' in request.POST:
        update_all['equipment_location'] = request.POST.get('equipment_location')
    if 'scope' in request.POST:
        equipment.scope = request.POST['scope'].strip()
    if 'completion' in request.POST:
        if request.POST['completion'] == 'True':
            equipment.completion = True
        elif request.POST['completion'] == 'False':
            equipment.completion = False
    if "mandatory_test_equipment" in request.POST:
        mandatory_test_equipments = request.POST['mandatory_test_equipment'].split(',')
        mandatory_test_equipments = TestEquipment.objects.filter(name__in=mandatory_test_equipments)

        equipment.mandatory_test_equipment.clear()
        equipment.mandatory_test_equipment.add(*mandatory_test_equipments)

    if "optional_test_equipment" in request.POST:
        optional_test_equipments = request.POST['optional_test_equipment'].split(',')
        optional_test_equipments = TestEquipment.objects.filter(name__in=optional_test_equipments)

        equipment.optional_test_equipment.clear()
        equipment.optional_test_equipment.add(*optional_test_equipments)

    equipment.save()
    if update_all:
        Equipment.objects.filter(Q(id=equipment_mold_id) | Q(equipment_mold_id=equipment_mold_id)).update(**update_all)

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


#probably need to remove
def create_equipment(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    manufacturer = request.POST["manufacturer"]
    model = request.POST["equipmentmmodel"]
    equipmentmanual = request.POST["equipmentmanual"]
    tips = request.POST["tips"]
    voltage = request.POST["voltage"]
    new_equipment=Equipment(manufacturer=manufacturer, equipmentmodel=equipmentmodel, equipmentmanual=equipmentmanual, tips=tips, voltage=voltage)
    return HttpResponseRedirect(reverse("templates", args=(None, )))

#probably need to remove
def complete_equipment(request, equipment_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    equipment=Equipment.objects.get(pk=equipment_id)
    equipment.completion=True
    equipment.save()
    job=Job.objects.get(equipment=equipment)
    return HttpResponseRedirect(reverse("job", args=(job.id, )))

def archive_job(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        job=Job.objects.get(pk=job_id)
        job.archived=True
        job.trashed=False
        job.save()
    except Job.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Associated Job not Identified."})
    return HttpResponseRedirect(request.META.get('HTTP_REFERER')+"#job_info")

def unarchive_job(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        job=Job.objects.get(pk=job_id)
        job.archived=False
        job.trashed=False
        job.save()
    except Job.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Associated Job not Identified."})
    return HttpResponseRedirect(request.META.get('HTTP_REFERER')+"#job_info")

def trash_job(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        job=Job.objects.get(pk=job_id)
        job.trashed=True
        job.archived=False
        job.save()
    except Job.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Associated Job not Identified."})
    return HttpResponseRedirect(request.META.get('HTTP_REFERER')+"#job_info")

def untrash_job(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        job=Job.objects.get(pk=job_id)
        job.trashed=False
        job.archived=False
        job.save()
    except Job.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Associated Job not Identified."})
    return HttpResponseRedirect(request.META.get('HTTP_REFERER')+"#job_info")

def complete_job(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        job=Job.objects.get(pk=job_id)
        job.completion=True
        job.save()
    except Job.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Associated Job not Identified."})
    return HttpResponseRedirect(reverse("job", args=(job_id, ))+"#job_info")

def uncomplete_job(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        job=Job.objects.get(pk=job_id)
        job.completion=False
        job.save()
    except Job.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Associated Job not Identified."})
    return HttpResponseRedirect(reverse("job", args=(job_id, ))+"#job_info")

def copy_job(request, job_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    try:
        job=Job.objects.get(pk=job_id)
    except Job.DoesNotExist:
        return render(request, "jobs/error.html", {"message": "Contact Admin. Error Code: 'copy_job-JOB DNE'"})
    new_job=job
    new_job.job_name=job.job_name + ' (duplicate)'
    if request.POST["new_job_number"]:
        new_job.job_number = request.POST["new_job_number"]
    else:
        return render(request, "jobs/error.html", {"message": "A new job number must be specified. Please try again and enter a unique job number."})    
    #get user properties
    user_properties=UserProperties.objects.get(user=request.user)
    #get user company
    try:
        company=user_properties.company
    except:
        company=None
    #verify uniqueness of job within this company
    if Job.objects.filter(job_number=new_job.job_number, company = company):
        return render(request, "jobs/error.html", {"message": "Job Number is In Use. Enter A Unique One."})    
    
    new_job.pk= None
    new_job.save()

    return HttpResponseRedirect(reverse("job", args=(new_job.pk, )))

def add_feedback_note(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    
    if "add_note" in request.POST:
        if request.POST["add_note"] and request.POST["add_note"].strip():
            note_posted = request.POST["add_note"]
            note = FeedbackNote(note=note_posted)
            note.save()
            return HttpResponseRedirect(reverse("feedback"))
    return render(request, "jobs/error.html", {"message": "Failure! Note not added. Must contain at least one letter, number, or symbol"})

def working_update(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    
    if "note" in request.POST:
        if request.POST["note"] and request.POST["note"].strip():
            text = request.POST["note"]
            note = WorkingNote(note=text)
            note.pk = 1
            note.save()
            return HttpResponseRedirect(reverse("working_page"))
    return render(request, "jobs/error.html", {"message": "Failure! Note not added. Must contain at least one letter, number, or symbol"})

def edit_feedback_note(request, note_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    feedback_note=FeedbackNote.objects.get(pk=note_id)
    input_id="edit_note_input"+str(note_id)
    if input_id in request.POST:
        if request.POST[input_id] and request.POST[input_id].strip():
            feedback_note.note = request.POST[input_id]
            feedback_note.save()
            return HttpResponseRedirect(reverse("feedback"))

    if "edit_note" in request.POST:
        if request.POST["add_note"] and request.POST["add_note"].strip():
            note_posted = request.POST["add_note"]
            note = FeedbackNote(note=note_posted)
            note.save()
            return HttpResponseRedirect(reverse("feedback"))
    return render(request, "jobs/error.html", {"message": "Failure! Note not added. Must contain at least one letter, number, or symbol"})

def add_feedback_file(request):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    
    if 'file_add' in request.FILES:
        file_add = request.FILES["file_add"]
        feedback_file=FeedbackFile()
        feedback_file.feedback_file = file_add
        feedback_file.save()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def ag_new_well(request):    
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    if 'new_well_name' in request.POST:
        if request.POST['new_well_name'] and request.POST["new_well_name"].strip():
            new_well = Well(name = request.POST["new_well_name"])
            new_well.save()    
    return HttpResponseRedirect(reverse("agwells"))

def ag_new_maint(request, well_id):    
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    well = Well.objects.get(pk=well_id)
    if 'new_maint_name' in request.POST:
        if request.POST['new_maint_name'] and request.POST["new_maint_name"].strip():
            new_maint = MaintEvent(title = request.POST["new_maint_name"])
            new_maint.save()
            new_maint.well = well
            new_maint.save()
    return HttpResponseRedirect(reverse("agwell", args=(well_id, )))

def ag_edit_maint(request, maint_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    maint = MaintEvent.objects.get(pk=maint_id)    
    if 'edit_maint_title' in request.POST:
        if request.POST['edit_maint_title'] and request.POST["edit_maint_title"].strip():
            maint.title=request.POST["edit_maint_title"]
    if 'edit_maint_description' in request.POST:
        if request.POST['edit_maint_description'] and request.POST["edit_maint_description"].strip():
            maint.description=request.POST["edit_maint_description"]
    if 'hours' in request.POST:
        if request.POST['hours']:
            maint.hours=request.POST["hours"]
    if "oil_change_date" in request.POST:
        if request.POST["oil_change_date"]!="":
            maint.oil_change_date = request.POST["oil_change_date"]
    maint.save()
    return HttpResponseRedirect(reverse("agwell", args=(maint.well.pk, )))

def ag_edit_well(request, well_id):    
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    well = Well.objects.get(pk=well_id)
    if 'edit_well_name' in request.POST:
        if request.POST['edit_well_name'] and request.POST["edit_well_name"].strip():
            well.name=request.POST["edit_well_name"]
    if 'location' in request.POST:
        if request.POST['location'] and request.POST["location"].strip():
            well.location=request.POST["location"]   
    if 'motor' in request.POST:
        if request.POST['motor'] and request.POST["motor"].strip():
            well.motor=request.POST["motor"]
    
    if 'oil_type' in request.POST:
        if request.POST['oil_type'] and request.POST["oil_type"].strip():
            well.oil_type=request.POST["oil_type"]
    
    if 'fuel_type' in request.POST:
        if request.POST['fuel_type'] and request.POST["fuel_type"].strip():
            well.fuel_type=request.POST["fuel_type"]
    
    if 'oil_filter' in request.POST:
        if request.POST['oil_filter'] and request.POST["oil_filter"].strip():
            well.oil_filter=request.POST["oil_filter"]
    
    if 'air_filter' in request.POST:
        if request.POST['air_filter'] and request.POST["air_filter"].strip():
            well.air_filter=request.POST["air_filter"]
    
    if 'oil_capacity' in request.POST:
        if request.POST['oil_capacity']:
            well.oil_capacity=request.POST["oil_capacity"]
    if 'nav_link' in request.POST:
        if request.POST['nav_link']:
            well.nav_link = request.POST['nav_link']    
    well.save()
    return HttpResponseRedirect(reverse("agwells"))

def add_maint_file(request, maint_id):
    if not request.user.is_authenticated:
        return render(request, "jobs/login.html", {"message": None})
    request_str = 'file_add'+ str(maint_id)
    maint = MaintEvent.objects.get(pk=maint_id)
    if request_str in request.FILES:
        file_add = request.FILES[request_str]
        maint_file=MaintFile(maint_event=maint)
        maint_file.save()
        maint_file.maint_file = file_add
        maint_file.save()
        if request_str in request.FILES:
            maint_file.file_name = request.FILES[request_str]
            maint_file.save()
        
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
